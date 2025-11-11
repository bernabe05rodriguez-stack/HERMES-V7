# -*- coding: utf-8 -*-
"""
HΞЯMΞS V1 - Envío automático de mensajes de WhatsApp
Autor: Berna - 2025
Con procesador de Excel/CSV integrado

Código limpiado y optimizado.
(Incluye Tooltip en el título principal y corrección de encoding)

--- MODIFICADO (para incluir MODO GRUPO en Fidelizado) ---
--- FIX 6 (Fix path with spaces issue, Use keyevents for text input) ---
"""

import subprocess
import time
import random
import tkinter as tk
import customtkinter as ctk
import tkinter.font as tkfont
from tkinter import filedialog, messagebox
import os
import threading
from datetime import datetime, timedelta
import sys
import csv
import io
import urllib.parse
import shlex # Import shlex for better command splitting
import tempfile
import shutil
import adbutils
import uiautomator2 as u2
import pytesseract
from PIL import Image
import re
import xml.etree.ElementTree as ET


# --- Función para encontrar archivos en modo compilado ---
def resource_path(relative_path):
    """ Obtiene la ruta absoluta al recurso, funciona para desarrollo y para PyInstaller """
    try:
        # PyInstaller crea una carpeta temporal y guarda la ruta en _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# --- Constante para el directorio base ---
BASE_DIR = resource_path(".")

# --- INICIO: Mapeo de caracteres a keycodes ADB ---
# (Simplificado, solo incluye caracteres comunes. Se puede expandir)
KEYCODE_MAP = {
    '0': 'KEYCODE_0', '1': 'KEYCODE_1', '2': 'KEYCODE_2', '3': 'KEYCODE_3',
    '4': 'KEYCODE_4', '5': 'KEYCODE_5', '6': 'KEYCODE_6', '7': 'KEYCODE_7',
    '8': 'KEYCODE_8', '9': 'KEYCODE_9',
    'a': 'KEYCODE_A', 'b': 'KEYCODE_B', 'c': 'KEYCODE_C', 'd': 'KEYCODE_D',
    'e': 'KEYCODE_E', 'f': 'KEYCODE_F', 'g': 'KEYCODE_G', 'h': 'KEYCODE_H',
    'i': 'KEYCODE_I', 'j': 'KEYCODE_J', 'k': 'KEYCODE_K', 'l': 'KEYCODE_L',
    'm': 'KEYCODE_M', 'n': 'KEYCODE_N', 'o': 'KEYCODE_O', 'p': 'KEYCODE_P',
    'q': 'KEYCODE_Q', 'r': 'KEYCODE_R', 's': 'KEYCODE_S', 't': 'KEYCODE_T',
    'u': 'KEYCODE_U', 'v': 'KEYCODE_V', 'w': 'KEYCODE_W', 'x': 'KEYCODE_X',
    'y': 'KEYCODE_Y', 'z': 'KEYCODE_Z',
    ' ': 'KEYCODE_SPACE', '.': 'KEYCODE_PERIOD', ',': 'KEYCODE_COMMA',
    '!': 'KEYCODE_1', '?': 'KEYCODE_SLASH', # SHIFT + /
    '@': 'KEYCODE_AT', '#': 'KEYCODE_POUND', '$': 'KEYCODE_4', # SHIFT + 4
    '%': 'KEYCODE_5', '^': 'KEYCODE_6', '&': 'KEYCODE_7', '*': 'KEYCODE_8',
    '(': 'KEYCODE_9', ')': 'KEYCODE_0', '-': 'KEYCODE_MINUS', '_': 'KEYCODE_MINUS', # SHIFT + -
    '+': 'KEYCODE_PLUS', '=': 'KEYCODE_EQUALS', '/': 'KEYCODE_SLASH',
    '\\': 'KEYCODE_BACKSLASH', '\n': 'KEYCODE_ENTER', # Nueva línea es Enter
    # Símbolos comunes con SHIFT (esto puede variar según el layout del teclado virtual)
    ':': 'KEYCODE_SEMICOLON', # SHIFT + ;
    '"': 'KEYCODE_APOSTROPHE', # SHIFT + '
    # ... se pueden añadir más según sea necesario
}
NEEDS_SHIFT = "!@#$%^&*()_+?:\"" + "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
# --- FIN: Mapeo ---

# Verificar dependencias
try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("\n"+"="*50+"\nERROR: Falta 'openpyxl'. Ejecuta INSTALAR.bat.\n"+"="*50)
    input("\nEnter para salir...")
    sys.exit(1)
try:
    from PIL import Image, ImageTk
except ImportError:
    print("\n"+"="*50+"\nERROR: Falta 'Pillow'. Ejecuta INSTALAR.bat.\n"+"="*50)
    input("\nEnter para salir...")
    sys.exit(1)
try:
    import customtkinter
except ImportError:
    print("\n"+"="*50+"\nERROR: Falta 'customtkinter'. Ejecuta INSTALAR.bat.\n"+"="*50)
    input("\nEnter para salir...")
    sys.exit(1)

# --- Funciones de color ---
def _clamp(value):
    """Asegura que un valor esté entre 0 y 255."""
    return max(0, min(255, int(value)))

def lighten_color(color, factor=0.1):
    """Aclara un color hexadecimal."""
    color = color.lstrip('#')
    if len(color) == 3:
        color = "".join([c*2 for c in color])
    if len(color) != 6:
        return color
    try:
        r, g, b = int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
    except ValueError:
        return color
    r = _clamp(r + (255 - r) * factor)
    g = _clamp(g + (255 - g) * factor)
    b = _clamp(b + (255 - b) * factor)
    return f"#{int(r):02x}{int(g):02x}{int(b):02x}"

def darken_color(color, factor=0.1):
    """Oscurece un color hexadecimal."""
    color = color.lstrip('#')
    if len(color) == 3:
        color = "".join([c*2 for c in color])
    if len(color) != 6:
        return color
    try:
        r, g, b = int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
    except ValueError:
        return color
    r = _clamp(r * (1 - factor))
    g = _clamp(g * (1 - factor))
    b = _clamp(b * (1 - factor))
    return f"#{int(r):02x}{int(g):02x}{int(b):02x}"

# --- INICIO MODIFICACIÓN: Clase para Tooltips (CORREGIDA) ---
class Tooltip:
    """
    Crea un tooltip (mensaje flotante) para un widget de CustomTkinter.
    Se instancia como: Tooltip(widget, "Texto del tooltip", font_info)
    """
    def __init__(self, widget, text, font_info):
        self.widget = widget
        self.text = text
        self.font_info = font_info
        self.tooltip_window = None
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)

    def show_tooltip(self, event): # <--- CORRECCIÓN: Se usa el 'event'
        if self.tooltip_window:
            return

        # Crear la ventana Toplevel
        self.tooltip_window = tk.Toplevel(self.widget)
        self.tooltip_window.wm_overrideredirect(True) # Sin bordes/barra de título

        # --- INICIO DE LA CORRECCIÓN ---
        # Usar las coordenadas del mouse (event.x_root) en lugar de
        # las coordenadas del widget (widget.winfo_rootx()), que
        # pueden ser incorrectas (0,0) si la ventana se está iniciando.
        # Se posiciona 15px a la derecha y 10px abajo del cursor.
        x = event.x_root + 15
        y = event.y_root + 10
        # --- FIN DE LA CORRECCIÓN ---

        # Ajustar si se sale de la pantalla (simple check)
        if x + 400 > self.widget.winfo_screenwidth():
            x = self.widget.winfo_screenwidth() - 410

        self.tooltip_window.wm_geometry(f"+{int(x)}+{int(y)}")

        # Añadir el label de CustomTkinter dentro
        label = ctk.CTkLabel(self.tooltip_window,
                             text=self.text,
                             font=self.font_info,
                             fg_color=("#333333", "#444444"), # Color oscuro
                             text_color="white",
                             corner_radius=6,
                             justify='left',
                             wraplength=400, # Ancho máximo del texto
                             padx=10, pady=10)
        label.pack()

        self.tooltip_window.update_idletasks()
        self.tooltip_window.lift() # Asegurarse de que esté al frente

    def hide_tooltip(self, event): # <--- CORRECCIÓN: Se usa el 'event'
        if self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None

# --- Clase principal de la aplicación ---
class Hermes:
    def __init__(self, root):
        self.root = root
        self.root.title("HΞЯMΞS V1")
        self.root.state('zoomed')
        self.root.resizable(True, True)
        self.root.minsize(1500, 900)
        self.center_window(1500, 900)

        # Variables de estado
        self.adb_path = tk.StringVar(value="")
        self.delay_min = tk.IntVar(value=10)
        self.delay_max = tk.IntVar(value=15)
        self.wait_after_open = tk.IntVar(value=15)
        self.wait_after_first_enter = tk.IntVar(value=10)

        self.excel_file = ""
        self.links = []
        self.devices = []

        self.is_running = False
        self.is_paused = False
        self.should_stop = False
        self.pause_lock = threading.Lock()

        self.total_messages = 0
        self.sent_count = 0
        self.failed_count = 0
        self.current_index = 0
        self.start_time = None
        self.task_times = []  # Lista de tiempos de cada tarea para promediado
        self.last_task_time = None  # Tiempo de inicio de la última tarea

        # --- INICIO MODIFICACIÓN: Variables de Fidelizado (Modo Bucles Blast V2) ---
        self.manual_inputs_numbers = [] # Almacena números
        self.manual_inputs_groups = [] # Almacena links de grupo
        self.manual_paired_messages = [] # Almacena los mensajes pareados (para Modo Grupo y Mixto)
        self.manual_messages_numbers = [] # Almacena los mensajes .txt para números
        self.manual_messages_groups = [] # Almacena los mensajes .txt para grupos
        
        self.manual_mode = False # Flag general de Fidelizado
        self.group_mode = False # Flag para MODO GRUPO (puro)
        
        # Estado: "NUMEROS", "GRUPOS", "MIXTO" o None (modo tradicional Excel/CSV)
        self.fidelizado_mode = None 
        self.mixto_variant = tk.IntVar(value=1)  # Variante del modo mixto: 1, 2 o 3
        
        # Índice de inicio aleatorio para rotación de mensajes
        self.mensaje_start_index = 0
        
        # --- INICIO MODIFICACIÓN: Variables para el nuevo Modo Números Automático ---
        self.fidelizado_numeros_mode = tk.StringVar(value="Uno a uno")
        self.detected_phone_lines = [] # Almacenará {"device": str, "type": "WA/WB", "number": str}
        self.manual_cycles_var = tk.IntVar(value=1) # NUEVO: Para los ciclos del modo "Uno a uno"
        # --- FIN MODIFICACIÓN ---

        self.manual_loops = 1

        self.fidelizado_delay_min = tk.IntVar(value=10)
        self.fidelizado_delay_max = tk.IntVar(value=15)
        # --- FIX: Añadir variables separadas para el retardo entre envíos ---
        self.fidelizado_send_delay_min = tk.IntVar(value=10)
        self.fidelizado_send_delay_max = tk.IntVar(value=15)
        
        # Variables de tiempo para Modo Grupos Dual
        self.wait_after_write = tk.IntVar(value=2)  # Tiempo después de escribir antes del primer Enter
        self.wait_between_enters = tk.IntVar(value=3)  # Tiempo entre el primer y segundo Enter
        self.wait_between_messages = tk.IntVar(value=2)  # Tiempo entre Business y Normal
        self.write_speed = tk.StringVar(value="Normal")  # Velocidad de escritura: Lento, Normal, Rápido
        self.whatsapp_mode = tk.StringVar(value="Todas")  # Qué WhatsApp usar: Normal, Business, Ambos
        self.traditional_send_mode = tk.StringVar(value="Business")  # Modo de envío tradicional: Business, Normal, Ambos, TODOS

        self.raw_data = []
        self.columns = []
        self.selected_columns = []
        self.phone_columns = []

        self.fidelizado_unlocked = True
        self.fidelizado_unlock_btn = None
        self.dark_mode = False  # Estado del modo oscuro

        # Paleta de colores
        self.colors_light = {
            'blue': '#4285F4', 'green': '#1DB954', 'orange': '#FB923C',
            'bg': '#e8e8e8', 'bg_card': '#ffffff', 'bg_header': '#ffffff',
            'bg_log': '#282c34',
            'log_text': '#abb2bf', 'log_success': '#98c379', 'log_error': '#e06c75',
            'log_warning': '#d19a66', 'log_info': '#61afef',
            'text': '#202124', 'text_light': '#5f6368', 'text_header_buttons': '#ffffff', 'text_header': '#000000', 'log_title_color': '#ffffff',
            'action_detect': '#2563EB', 'action_excel': '#F97316',
            'action_fidelizador': '#111827', 'action_start': '#16A34A',
            'action_pause': '#FB923C', 'action_cancel': '#DC2626'
        }
        
        self.colors_dark = {
            'blue': '#5B9FFF', 'green': '#1ED760', 'orange': '#FFA45C',
            'bg': '#000000', 'bg_card': '#1a1a1a', 'bg_header': '#1a1a1a',
            'bg_log': '#1a1a1a',
            'log_text': '#ffffff', 'log_success': '#98c379', 'log_error': '#e06c75',
            'log_warning': '#d19a66', 'log_info': '#61afef',
            'text': '#ffffff', 'text_light': '#cccccc', 'text_header_buttons': '#ffffff', 'text_header': '#ffffff', 'log_title_color': '#ffffff',
            'action_detect': '#5B9FFF', 'action_excel': '#FFA45C',
            'action_fidelizador': '#e4e6eb', 'action_start': '#22C55E',
            'action_pause': '#FFA45C', 'action_cancel': '#EF4444'
        }
        
        self.colors = self.colors_light.copy()

        self.hover_colors = {k: darken_color(v, 0.18) for k, v in self.colors.items() if k.startswith('action_')}

        # Fuentes
        self.fonts = {
            'header': ('Big Russian', 76, 'bold'),
            'card_title': ('Inter', 20, 'bold'),
            'button': ('Inter', 13, 'bold'),
            'button_small': ('Inter', 12, 'bold'),
            'stat_value': ('Inter', 40, 'bold'),
            'stat_label': ('Inter', 13),
            'setting_label': ('Inter', 12),
            'log_title': ('Inter', 16, 'bold'),
            'log_text': ('Consolas', 12),
            'progress_label': ('Inter', 12, 'bold'),
            'progress_value': ('Inter', 20, 'bold'),
            'time_label': ('Inter', 10),
            'dialog_title': ('Inter', 16, 'bold'),
            'dialog_text': ('Inter', 12)
        }

        self.auto_detect_adb()
        self.setup_ui()

    def center_window(self, width, height):
        self.root.update_idletasks()
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        x = (sw // 2) - (width // 2)
        y = (sh // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')

    def setup_ui(self):
        # Configurar fondo de la ventana principal
        self.root.configure(fg_color=self.colors['bg'])
        
        # 1. Header
        header = ctk.CTkFrame(self.root, fg_color=self.colors['bg_header'], height=140, corner_radius=30)
        header.pack(fill=tk.X, pady=(10, 10), padx=10)
        header.pack_propagate(False)

        hc = ctk.CTkFrame(header, fg_color=self.colors['bg_header'])
        hc.pack(expand=True, fill=tk.X, padx=40)

        # Logo Izquierdo
        try:
            l_img_path = os.path.join(BASE_DIR, 'logo_left.png')
            l_img = Image.open(l_img_path).resize((150, 150), Image.Resampling.LANCZOS)
            l_pho = ctk.CTkImage(light_image=l_img, dark_image=l_img, size=(150, 150))
            ctk.CTkLabel(hc, image=l_pho, text="").pack(side=tk.LEFT, padx=(0, 20))
        except Exception as e:
            print(f"Error cargando logo_left: {e}")
            ctk.CTkLabel(hc, text="🦶", font=('Inter', 60), fg_color="transparent").pack(side=tk.LEFT, padx=(0, 20))

        # Logo Derecho
        try:
            r_img_path = os.path.join(BASE_DIR, 'logo_right.png')
            r_img = Image.open(r_img_path).resize((150, 150), Image.Resampling.LANCZOS)
            r_pho = ctk.CTkImage(light_image=r_img, dark_image=r_img, size=(150, 150))
            ctk.CTkLabel(hc, image=r_pho, text="").pack(side=tk.RIGHT, padx=(20, 0))
        except Exception as e:
            print(f"Error cargando logo_right: {e}")
            ctk.CTkLabel(hc, text="🦶", font=('Inter', 60), fg_color="transparent").pack(side=tk.RIGHT, padx=(20, 0))

        # Título
        title_label = ctk.CTkLabel(hc, text="HΞЯMΞS", font=self.fonts['header'],
                                   fg_color="transparent",
                                   text_color=self.colors['text_header'],
                                   cursor="hand2") # Añadir cursor para indicar que es interactivo
        title_label.pack(side=tk.LEFT, fill=tk.X, expand=True, anchor='center')

        # Tooltip para el título
        tooltip_text = (
            "Hermes fue el mensajero de los dioses en la mitología griega. \n"
            "Hijo de Zeus, simbolizaba la comunicación, la rapidez y el ingenio. \n"
            "También protegía a los viajeros y guiaba las almas al inframundo.\n\n"
            "Programa pensado y creado por \n"
            "BERNABE GABRIEL RODRIGUEZ, y FRANCISCO JOSE RODRIGUEZ."
        )
        tooltip_font = self.fonts.get('dialog_text', ('Inter', 12))
        self.hermes_tooltip = Tooltip(widget=title_label, text=tooltip_text, font_info=tooltip_font)

        # 2. Contenedor principal scrollable
        mc = ctk.CTkFrame(self.root, fg_color="transparent")
        mc.pack(fill=tk.BOTH, expand=True, padx=40, pady=(0, 20))

        self.scroll_frame = ctk.CTkScrollableFrame(mc, fg_color="transparent")
        self.scroll_frame.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        self.scroll_frame.grid_columnconfigure(0, weight=618, uniform='main_panels')
        self.scroll_frame.grid_columnconfigure(1, weight=382, uniform='main_panels')
        self.scroll_frame.grid_rowconfigure(0, weight=1)
        self.main_layout = self.scroll_frame

        # 3. Paneles
        left = ctk.CTkFrame(self.main_layout, fg_color="transparent")
        right = ctk.CTkFrame(self.main_layout, fg_color="transparent")
        self.left_panel = left
        self.right_panel = right
        self._current_main_layout = None

        self.root.bind("<Configure>", self._on_main_configure)
        self.setup_right(right) # FIX: Inicializar el panel derecho primero para que exista el log_text
        self.setup_left(left)
        self.root.update_idletasks()
        self._update_main_layout(self.root.winfo_width())

    def _on_main_configure(self, event):
        self._update_main_layout(self.root.winfo_width())

    def _update_main_layout(self, width=None):
        """Cambia entre vista de 2 columnas o 1 columna (apilada) si la ventana es muy angosta."""
        if not hasattr(self, 'left_panel') or not hasattr(self, 'right_panel'):
            return
        if not width:
            width = self.root.winfo_width() - 80 # 80 por el padding

        mode = 'stacked' if width < 1100 else 'columns'

        if self._current_main_layout == mode:
            return

        self.left_panel.update_idletasks()
        self.right_panel.update_idletasks()

        if mode == 'columns':
            self.main_layout.grid_columnconfigure(0, weight=618, uniform='main_panels', minsize=0)
            self.main_layout.grid_columnconfigure(1, weight=382, uniform='main_panels', minsize=0)
            self.main_layout.grid_rowconfigure(1, weight=0)
            self.left_panel.grid(row=0, column=0, sticky='nsew', padx=(0, 10), pady=0)
            self.right_panel.grid(row=0, column=1, sticky='nsew', padx=(10, 0), pady=0)
        else: # mode == 'stacked'
            self.main_layout.grid_columnconfigure(0, weight=1, uniform='main_panels', minsize=0)
            self.main_layout.grid_columnconfigure(1, weight=0, minsize=0)
            self.main_layout.grid_rowconfigure(1, weight=1)
            self.left_panel.grid(row=0, column=0, sticky='nsew', padx=0, pady=0)
            self.right_panel.grid(row=1, column=0, sticky='nsew', padx=0, pady=0)

        self._current_main_layout = mode

    def _iniciar_ver_pantalla(self):
        """Handles the logic for the 'Ver Pantalla' button."""
        if not self.devices:
            messagebox.showerror("Error", "No hay dispositivos detectados. Por favor, detecta los dispositivos primero.", parent=self.root)
            return

        if len(self.devices) == 1:
            self._lanzar_scrcpy(self.devices[0])
        else:
            self._open_device_selection_window()

    def _open_device_selection_window(self):
        """Opens a window to select a device for screen mirroring."""
        selection_window = ctk.CTkToplevel(self.root)
        selection_window.title("Seleccionar Dispositivo")
        selection_window.transient(self.root)
        selection_window.grab_set()

        # Center the window
        width, height = 300, len(self.devices) * 50 + 40
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (width // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (height // 2)
        selection_window.geometry(f"{width}x{height}+{x}+{y}")

        ctk.CTkLabel(selection_window, text="Selecciona un dispositivo:", font=self.fonts['button']).pack(pady=10)

        def on_select(device_id):
            selection_window.destroy()
            self._lanzar_scrcpy(device_id)

        for device_id in self.devices:
            btn = ctk.CTkButton(selection_window, text=device_id, command=lambda d=device_id: on_select(d),
                                font=self.fonts['button'], height=40)
            btn.pack(fill=tk.X, padx=20, pady=5)

    def _lanzar_scrcpy(self, device_id):
        """Launches scrcpy for a specific device."""
        scrcpy_path = os.path.join(BASE_DIR, "scrcpy-win64-v3.2", "scrcpy.exe")

        if not os.path.exists(scrcpy_path):
            self.log(f"Error: No se encontró scrcpy.exe en la ruta esperada: {scrcpy_path}", 'error')
            messagebox.showerror("Error", "No se encontró 'scrcpy.exe'.\nAsegúrate de que la carpeta 'scrcpy-win64-v3.2' está en el directorio del programa.", parent=self.root)
            return

        self.log(f"Iniciando scrcpy para el dispositivo: {device_id}", 'info')

        try:
            # Use Popen to run scrcpy in a non-blocking way
            command = [scrcpy_path, "-s", device_id]
            subprocess.Popen(command)
        except Exception as e:
            self.log(f"Error al iniciar scrcpy: {e}", 'error')
            messagebox.showerror("Error", f"No se pudo iniciar scrcpy:\n{e}", parent=self.root)

    def _ejecutar_perfil(self, profile_name):
        """Inicia la ejecución de un perfil de comandos en un hilo separado."""
        if not self.devices:
            messagebox.showerror("Error", "No hay dispositivos detectados. Por favor, detecta los dispositivos primero.", parent=self.root)
            return

        self.log(f"Iniciando ejecución del {profile_name.replace('_', ' ').title()} en {len(self.devices)} dispositivo(s)...", 'info')

        # Iniciar la ejecución en un hilo para no bloquear la UI
        threading.Thread(target=self._run_profile_thread, args=(profile_name,), daemon=True).start()

    def _run_profile_thread(self, profile_name):
        """
        Hilo de trabajo que ejecuta un perfil de comandos en todos los dispositivos simultáneamente.
        """
        # Estructura de datos con los perfiles de comandos
        profiles = {
            'perfil_1': {
                'rapido_1': [("shell input keyevent KEYCODE_BACK", 5),
                             ("shell am force-stop com.android.settings", 1),
                             ("shell am force-stop com.whatsapp", 1),
                             ("shell am force-stop com.whatsapp.w4b", 1)],
                'lento_1': [("shell cmd statusbar expand-settings", 1),
                            ("shell input keyevent KEYCODE_TAB", 13),
                            ("shell input keyevent KEYCODE_ENTER", 1),
                            ("shell input keyevent KEYCODE_TAB", 2),
                            ("shell input keyevent KEYCODE_ENTER", 1)],
                'pausa_1': [("sleep", 8)],
                'rapido_2': [("shell input keyevent KEYCODE_BACK", 5),
                             ("shell am force-stop com.android.settings", 1),
                             ("shell am force-stop com.whatsapp", 1),
                             ("shell am force-stop com.whatsapp.w4b", 1)]
            },
            'perfil_2': {
                'rapido_1': [("shell input keyevent KEYCODE_BACK", 5),
                             ("shell am force-stop com.android.settings", 1),
                             ("shell am force-stop com.whatsapp", 1),
                             ("shell am force-stop com.whatsapp.w4b", 1)],
                'lento_1': [("shell cmd statusbar expand-settings", 1),
                            ("shell input keyevent KEYCODE_TAB", 13),
                            ("shell input keyevent KEYCODE_ENTER", 1),
                            ("shell input keyevent KEYCODE_TAB", 3),
                            ("shell input keyevent KEYCODE_ENTER", 1)],
                'pausa_1': [("sleep", 8)],
                'rapido_2': [("shell input keyevent KEYCODE_BACK", 5),
                             ("shell am force-stop com.android.settings", 1),
                             ("shell am force-stop com.whatsapp", 1),
                             ("shell am force-stop com.whatsapp.w4b", 1)]
            },
            'perfil_3': {
                'rapido_1': [("shell input keyevent KEYCODE_BACK", 5),
                             ("shell am force-stop com.android.settings", 1),
                             ("shell am force-stop com.whatsapp", 1),
                             ("shell am force-stop com.whatsapp.w4b", 1)],
                'lento_1': [("shell cmd statusbar expand-settings", 1),
                            ("shell input keyevent KEYCODE_TAB", 13),
                            ("shell input keyevent KEYCODE_ENTER", 1),
                            ("shell input keyevent KEYCODE_TAB", 4),
                            ("shell input keyevent KEYCODE_ENTER", 1)],
                'pausa_1': [("sleep", 8)],
                'rapido_2': [("shell input keyevent KEYCODE_BACK", 5),
                             ("shell am force-stop com.android.settings", 1),
                             ("shell am force-stop com.whatsapp", 1),
                             ("shell am force-stop com.whatsapp.w4b", 1)]
            },
            'perfil_4': {
                'rapido_1': [("shell input keyevent KEYCODE_BACK", 5),
                             ("shell am force-stop com.android.settings", 1),
                             ("shell am force-stop com.whatsapp", 1),
                             ("shell am force-stop com.whatsapp.w4b", 1)],
                'lento_1': [("shell cmd statusbar expand-settings", 1),
                            ("shell input keyevent KEYCODE_TAB", 13),
                            ("shell input keyevent KEYCODE_ENTER", 1),
                            ("shell input keyevent KEYCODE_TAB", 5),
                            ("shell input keyevent KEYCODE_ENTER", 1)],
                'pausa_1': [("sleep", 8)],
                'rapido_2': [("shell input keyevent KEYCODE_BACK", 5),
                             ("shell am force-stop com.android.settings", 1),
                             ("shell am force-stop com.whatsapp", 1),
                             ("shell am force-stop com.whatsapp.w4b", 1)]
            }
        }

        command_sequence = profiles.get(profile_name)
        if not command_sequence:
            self.log(f"Error: Perfil '{profile_name}' no encontrado.", 'error')
            return

        for block_name, commands in command_sequence.items():
            if 'rapido' in block_name:
                delay = 0.3
            elif 'lento' in block_name:
                delay = 1.2
            else: # Es un bloque de pausa
                delay = 0

            for command_str, repetitions in commands:
                if command_str == "sleep":
                    self.log(f"Esperando {repetitions} segundos...", 'info')
                    time.sleep(repetitions)
                    continue

                for i in range(repetitions):
                    # Usar hilos para ejecutar el comando en todos los dispositivos a la vez
                    device_threads = []
                    for device in self.devices:
                        # self._run_adb_command espera una lista de argumentos
                        args = ['-s', device] + command_str.split()
                        thread = threading.Thread(target=self._run_adb_command, args=(args,), daemon=True)
                        device_threads.append(thread)
                        thread.start()

                    # Esperar a que todos los dispositivos terminen el comando actual
                    for thread in device_threads:
                        thread.join()

                    time.sleep(delay) # Pausa después de cada ejecución de comando

        self.log(f"Perfil {profile_name.replace('_', ' ').title()} completado.", 'success')


    def setup_left(self, parent):
        # Contenedor principal para las vistas
        self.views_container = ctk.CTkFrame(parent, fg_color="transparent")
        self.views_container.pack(fill=tk.BOTH, expand=True)

        # --- Vista Tradicional ---
        self.traditional_view_frame = ctk.CTkFrame(self.views_container, fg_color="transparent")
        self.setup_traditional_view(self.traditional_view_frame)

        # --- Vista Fidelizado (inicialmente vacía) ---
        self.fidelizado_view_frame = ctk.CTkFrame(self.views_container, fg_color="transparent")
        self.setup_fidelizado_view(self.fidelizado_view_frame) # <-- LLAMAR AL MÉTODO DE CONSTRUCCIÓN

        # Mostrar la vista tradicional por defecto
        self.show_traditional_view()

    def show_traditional_view(self):
        """Guarda el estado de la vista Fidelizado y muestra la tradicional."""
        # Guardar datos de los textboxes para persistencia
        if hasattr(self, 'fidelizado_groups_text'): # Comprobar si los widgets existen
            self.manual_inputs_groups = [line.strip() for line in self.fidelizado_groups_text.get("1.0", tk.END).splitlines() if line.strip()]
            # Los mensajes se gestionan al cargar el archivo, no se guardan desde un widget.
            # Asumir que los mensajes de grupo son los mismos
            self.manual_messages_groups = self.manual_messages_numbers

        self.fidelizado_view_frame.pack_forget()
        self.traditional_view_frame.pack(fill=tk.BOTH, expand=True)

    def show_fidelizado_view(self):
        """Muestra la vista de Fidelizado, repoblando los datos, y oculta las demás."""
        self._populate_fidelizado_inputs() # Repoblar datos al mostrar la vista
        self.traditional_view_frame.pack_forget()
        self.fidelizado_view_frame.pack(fill=tk.BOTH, expand=True)

    def setup_traditional_view(self, parent):
        # Bloque 1: Configuración de Tiempo
        cc = ctk.CTkFrame(parent, fg_color=self.colors['bg_card'], corner_radius=30)
        cc.pack(fill=tk.X, pady=(0, 30), padx=10)

        ctf = ctk.CTkFrame(cc, fg_color="transparent")
        ctf.pack(fill=tk.X, pady=(25, 15), padx=25)
        ctk.CTkLabel(ctf, text="⚙", font=('Inter', 20), fg_color="transparent").pack(side=tk.LEFT, padx=(0, 10))
        ctk.CTkLabel(ctf, text="Configuración de Tiempo", font=self.fonts['card_title'], fg_color="transparent", text_color=self.colors['text']).pack(side=tk.LEFT)

        ctk.CTkFrame(cc, fg_color=self.colors['text_light'], height=1).pack(fill=tk.X, pady=(0, 20), padx=25)

        s = ctk.CTkFrame(cc, fg_color="transparent")
        s.pack(fill=tk.X, pady=(0, 25), padx=25)
        self.create_setting(s, "Delay (seg):", self.delay_min, self.delay_max, 0)
        self.create_setting(s, "Espera Abrir (seg):", self.wait_after_open, None, 1)
        self.create_setting(s, "Espera Enter (seg):", self.wait_after_first_enter, None, 2)

        # Bloque 2: Acciones
        ac = ctk.CTkFrame(parent, fg_color=self.colors['bg_card'], corner_radius=30)
        ac.pack(fill=tk.X, pady=(0, 30), padx=10)

        atf = ctk.CTkFrame(ac, fg_color="transparent")
        atf.pack(fill=tk.X, pady=(25, 15), padx=25)
        ctk.CTkLabel(atf, text="Acciones", font=self.fonts['card_title'], fg_color="transparent", text_color=self.colors['text']).pack(side=tk.LEFT)

        # Botón desplegable para mostrar/ocultar funciones adicionales
        self.actions_expanded = False
        self.toggle_actions_btn = ctk.CTkButton(atf, text="▼", command=self.toggle_additional_actions,
                                               fg_color=self.colors['bg_card'], text_color=self.colors['text'],
                                               hover_color=self.colors["bg"],
                                               font=('Inter', 16, 'bold'),
                                               cursor='hand2', width=40, height=40, corner_radius=10,
                                               border_width=1, border_color=self.colors["text_light"])
        self.toggle_actions_btn.pack(side=tk.LEFT, padx=(12, 0))
        
        # Frame contenedor para los botones adicionales (inicialmente oculto) - organizado en grid
        self.additional_actions_frame = ctk.CTkFrame(atf, fg_color="transparent")
        
        # Configurar grid para organizar botones en filas
        self.additional_actions_frame.grid_columnconfigure(0, weight=0)
        self.additional_actions_frame.grid_columnconfigure(1, weight=0)
        self.additional_actions_frame.grid_columnconfigure(2, weight=0)
        self.additional_actions_frame.grid_columnconfigure(3, weight=0) # New column for layout
        
        # Primera fila de botones
        # Botón Fidelizado
        self.fidelizado_unlock_btn = ctk.CTkButton(self.additional_actions_frame, text="Fidelizado", command=self.handle_fidelizado_access,
                                                   fg_color=self.colors['bg_card'], text_color=self.colors['text'],
                                                   hover_color=self.colors["bg"],
                                                   font=('Inter', 13),
                                                   cursor='hand2', width=130, height=38, corner_radius=10, state=tk.NORMAL,
                                                   border_width=1, border_color=self.colors["text_light"])
        self.fidelizado_unlock_btn.grid(row=0, column=0, padx=(12, 8), pady=4)
        
        # Botón Inyector ADB
        self.adb_injector_btn = ctk.CTkButton(self.additional_actions_frame, text="Inyector ADB", command=self.open_adb_injector,
                                              fg_color=self.colors['bg_card'], text_color=self.colors['text'],
                                              hover_color=self.colors["bg"],
                                              font=('Inter', 13),
                                              cursor='hand2', width=130, height=38, corner_radius=10, state=tk.NORMAL,
                                              border_width=1, border_color=self.colors["text_light"])
        self.adb_injector_btn.grid(row=0, column=1, padx=8, pady=4)

        # Botón Ver Pantalla
        self.ver_pantalla_btn = ctk.CTkButton(self.additional_actions_frame, text="Ver Pantalla", command=self._iniciar_ver_pantalla,
                                              fg_color=self.colors['bg_card'], text_color=self.colors['text'],
                                              hover_color=self.colors["bg"],
                                              font=('Inter', 13),
                                              cursor='hand2', width=130, height=38, corner_radius=10, state=tk.NORMAL,
                                              border_width=1, border_color=self.colors["text_light"])
        self.ver_pantalla_btn.grid(row=0, column=2, padx=8, pady=4)
        
        # Botón Detectar Números
        self.detect_numbers_btn = ctk.CTkButton(self.additional_actions_frame, text="Detectar Números", command=self.detect_phone_numbers_thread,
                                              fg_color=self.colors['bg_card'], text_color=self.colors['text'],
                                              hover_color=self.colors["bg"],
                                              font=('Inter', 13),
                                              cursor='hand2', width=130, height=38, corner_radius=10, state=tk.NORMAL,
                                              border_width=1, border_color=self.colors["text_light"])
        self.detect_numbers_btn.grid(row=1, column=0, padx=(12, 8), pady=4)

        # Botón Cambiar Cuenta WhatsApp
        self.switch_account_btn = ctk.CTkButton(self.additional_actions_frame, text="Cambiar Cuenta", command=self.switch_whatsapp_account,
                                               fg_color=self.colors['bg_card'], text_color=self.colors['text'],
                                               hover_color=self.colors["bg"],
                                               font=('Inter', 13),
                                               cursor='hand2', width=130, height=38, corner_radius=10, state=tk.NORMAL,
                                               border_width=1, border_color=self.colors["text_light"])
        # self.switch_account_btn.grid(row=0, column=2, padx=8, pady=4)
        
        # Segunda fila de botones
        # Botón Cambiador
        self.cambiador_btn = ctk.CTkButton(self.additional_actions_frame, text="Cambiador", command=self.run_cambiador,
                                          fg_color=self.colors['bg_card'], text_color=self.colors['text'],
                                          hover_color=self.colors["bg"],
                                          font=('Inter', 13),
                                          cursor='hand2', width=130, height=38, corner_radius=10, state=tk.NORMAL,
                                          border_width=1, border_color=self.colors["text_light"])
        # self.cambiador_btn.grid(row=1, column=0, padx=(12, 8), pady=4)
        
        # Botón Modo Oscuro
        self.dark_mode_btn = ctk.CTkButton(self.additional_actions_frame, text="Modo Oscuro", command=self.toggle_dark_mode,
                                          fg_color=self.colors['bg_card'], text_color=self.colors['text'],
                                          hover_color=self.colors["bg"],
                                          font=('Inter', 13),
                                          cursor='hand2', width=130, height=38, corner_radius=10, state=tk.NORMAL,
                                          border_width=1, border_color=self.colors["text_light"])
        self.dark_mode_btn.grid(row=0, column=3, padx=8, pady=4)

        # Botones de Perfil
        self.perfil_1_btn = ctk.CTkButton(self.additional_actions_frame, text="Perfil 1", command=lambda: self._ejecutar_perfil('perfil_1'),
                                           fg_color=self.colors['bg_card'], text_color=self.colors['text'],
                                           hover_color=self.colors["bg"], font=('Inter', 13),
                                           cursor='hand2', width=130, height=38, corner_radius=10, state=tk.NORMAL,
                                           border_width=1, border_color=self.colors["text_light"])
        # self.perfil_1_btn.grid(row=1, column=0, padx=(12, 8), pady=4)

        self.perfil_2_btn = ctk.CTkButton(self.additional_actions_frame, text="Perfil 2", command=lambda: self._ejecutar_perfil('perfil_2'),
                                           fg_color=self.colors['bg_card'], text_color=self.colors['text'],
                                           hover_color=self.colors["bg"], font=('Inter', 13),
                                           cursor='hand2', width=130, height=38, corner_radius=10, state=tk.NORMAL,
                                           border_width=1, border_color=self.colors["text_light"])
        # self.perfil_2_btn.grid(row=1, column=1, padx=8, pady=4)

        self.perfil_3_btn = ctk.CTkButton(self.additional_actions_frame, text="Perfil 3", command=lambda: self._ejecutar_perfil('perfil_3'),
                                           fg_color=self.colors['bg_card'], text_color=self.colors['text'],
                                           hover_color=self.colors["bg"], font=('Inter', 13),
                                           cursor='hand2', width=130, height=38, corner_radius=10, state=tk.NORMAL,
                                           border_width=1, border_color=self.colors["text_light"])
        # self.perfil_3_btn.grid(row=1, column=2, padx=8, pady=4)

        self.perfil_4_btn = ctk.CTkButton(self.additional_actions_frame, text="Perfil 4", command=lambda: self._ejecutar_perfil('perfil_4'),
                                           fg_color=self.colors['bg_card'], text_color=self.colors['text'],
                                           hover_color=self.colors["bg"], font=('Inter', 13),
                                           cursor='hand2', width=130, height=38, corner_radius=10, state=tk.NORMAL,
                                           border_width=1, border_color=self.colors["text_light"])
        # self.perfil_4_btn.grid(row=1, column=3, padx=8, pady=4)

        ctk.CTkFrame(ac, fg_color=self.colors['text_light'], height=1).pack(fill=tk.X, pady=(0, 25), padx=25)

        acts = ctk.CTkFrame(ac, fg_color="transparent")
        acts.pack(fill=tk.X, pady=(0, 25), padx=25)

        btn_frames_data = [
            (1, "🔍  Detectar Dispositivos", self.detect_devices, 'action_detect'),
            (2, "📄  Cargar y Procesar Excel", self.load_and_process_excel, 'action_excel'),
        ]

        for num, text, cmd, color_key in btn_frames_data:
            bfc = ctk.CTkFrame(acts, fg_color="transparent")
            bfc.pack(fill=tk.X, pady=(0, 15))
            bfc.grid_columnconfigure(0, weight=0)
            bfc.grid_columnconfigure(1, weight=1)
            bfc.grid_rowconfigure(0, weight=1)

            num_lbl = ctk.CTkLabel(bfc, text=str(num), font=self.fonts['progress_value'], fg_color="transparent", text_color=self.colors['text'], width=40)
            num_lbl.grid(row=0, column=0, padx=(0, 15))

            btn = ctk.CTkButton(bfc, text=text, command=cmd,
                                fg_color=self.colors[color_key],
                                hover_color=self.hover_colors[color_key],
                                text_color=self.colors['text_header_buttons'], font=self.fonts['button'],
                                corner_radius=10, height=50)
            btn.grid(row=0, column=1, sticky='nsew')

            if num == 1: self.btn_detect = btn
            elif num == 2: self.btn_load = btn

        # Selector de Modo de Envío (Simple/Doble/Triple) - SOLO para modo tradicional
        mode_frame = ctk.CTkFrame(acts, fg_color="transparent")
        mode_frame.pack(fill=tk.X, pady=(0, 15))
        mode_frame.grid_columnconfigure(0, weight=0)
        mode_frame.grid_columnconfigure(1, weight=1)
        
        num_lbl_mode = ctk.CTkLabel(mode_frame, text="3", font=self.fonts['progress_value'], fg_color="transparent", text_color=self.colors['text'], width=40)
        num_lbl_mode.grid(row=0, column=0, padx=(0, 15))
        
        mode_selector_frame = ctk.CTkFrame(mode_frame, fg_color=self.colors['bg_card'], corner_radius=10, height=50)
        mode_selector_frame.grid(row=0, column=1, sticky='nsew')
        mode_selector_frame.grid_columnconfigure(0, weight=1)
        mode_selector_frame.grid_rowconfigure(0, weight=1)
        
        mode_label = ctk.CTkLabel(mode_selector_frame, text="Modo de Envío:", font=self.fonts['button'], text_color=self.colors['text'])
        mode_label.grid(row=0, column=0, padx=(20, 10), sticky='w')

        self.mode_selector = ctk.CTkSegmentedButton(mode_selector_frame, variable=self.traditional_send_mode,
                                                     values=["Business", "Normal", "Business/Normal", "B/N.1/N.2"],
                                                     font=('Inter', 10, 'bold'),
                                                     height=35,
                                                     corner_radius=8,
                                                     fg_color=self.colors['bg'],
                                                     selected_color=self.colors['action_excel'],
                                                     selected_hover_color=self.hover_colors['action_excel'],
                                                     unselected_color=self.colors['bg_card'],
                                                     unselected_hover_color=self.colors['bg'],
                                                     text_color=self.colors['text'])
        self.mode_selector.grid(row=0, column=1, padx=(10, 20))
        mode_selector_frame.grid_columnconfigure(1, weight=1)
        self.traditional_send_mode.trace_add('write', self.update_per_whatsapp_stat)
        
        # Botón 3: Iniciar Envío
        btn_frame_3 = ctk.CTkFrame(acts, fg_color="transparent")
        btn_frame_3.pack(fill=tk.X, pady=(0, 15))
        btn_frame_3.grid_columnconfigure(0, weight=0)
        btn_frame_3.grid_columnconfigure(1, weight=1)
        
        num_lbl_3 = ctk.CTkLabel(btn_frame_3, text="4", font=self.fonts['progress_value'], fg_color="transparent", text_color=self.colors['text'], width=40)
        num_lbl_3.grid(row=0, column=0, padx=(0, 15))
        
        self.btn_start = ctk.CTkButton(btn_frame_3, text="▶  INICIAR ENVÍO", command=self.start_sending,
                            fg_color=self.colors['action_start'],
                            hover_color=self.hover_colors['action_start'],
                            text_color=self.colors['text_header_buttons'], font=self.fonts['button'],
                            corner_radius=10, height=50)
        self.btn_start.grid(row=0, column=1, sticky='nsew')


        # Botones de control (Pausar/Cancelar)
        ctrls = ctk.CTkFrame(acts, fg_color="transparent")
        ctrls.pack(fill=tk.X, pady=(10, 0))
        self.btn_pause = ctk.CTkButton(ctrls, text="⏸  PAUSAR", command=self.pause_sending, fg_color=self.colors['action_pause'], hover_color=self.hover_colors['action_pause'], text_color=self.colors['text_header_buttons'], text_color_disabled='#ffffff', font=self.fonts['button_small'], corner_radius=20, height=40, state=tk.DISABLED)
        self.btn_pause.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        self.btn_stop = ctk.CTkButton(ctrls, text="⏹  CANCELAR", command=self.stop_sending, fg_color=self.colors['action_cancel'], hover_color=self.hover_colors['action_cancel'], text_color=self.colors['text_header_buttons'], text_color_disabled='#ffffff', font=self.fonts['button_small'], corner_radius=20, height=40, state=tk.DISABLED)
        self.btn_stop.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 0))

    def setup_right(self, parent):
        # Bloque 1: Estado y Progreso
        sc = ctk.CTkFrame(parent, fg_color=self.colors['bg_card'], corner_radius=30)
        sc.pack(fill=tk.X, pady=(0, 30), padx=10)

        t = ctk.CTkFrame(sc, fg_color="transparent")
        t.pack(fill=tk.X, pady=(25, 15), padx=25)
        ctk.CTkLabel(t, text="✓", font=('Inter', 20), fg_color="transparent", text_color=self.colors['green']).pack(side=tk.LEFT, padx=(0, 10))
        ctk.CTkLabel(t, text="Estado y Progreso", font=self.fonts['card_title'], fg_color="transparent", text_color=self.colors['text']).pack(side=tk.LEFT)

        ctk.CTkFrame(sc, fg_color=self.colors['text_light'], height=1).pack(fill=tk.X, pady=(0, 10), padx=25)

        stats = ctk.CTkFrame(sc, fg_color="transparent")
        stats.pack(fill=tk.BOTH, expand=True, pady=(0, 10), padx=25)
        self.stats_frame = stats
        self.create_stat(stats, "Total", "0", self.colors['blue'], 0)
        self.create_stat(stats, "Enviados", "0", self.colors['green'], 1)
        self.create_stat(stats, "Progreso", "0%", self.colors['orange'], 2)

        ctk.CTkLabel(sc, text="Progreso general", font=self.fonts['progress_label'], fg_color="transparent", text_color=self.colors['text_light']).pack(anchor='w', pady=(0, 5), padx=25)
        self.progress_label = ctk.CTkLabel(sc, text="--/--", font=self.fonts['progress_value'], fg_color="transparent", text_color=self.colors['text'])
        self.progress_label.pack(anchor='w', pady=(0, 10), padx=25)

        # Barra de progreso (fondo)
        bbg = ctk.CTkFrame(sc, fg_color=self.colors['text_light'], height=8, corner_radius=4)
        bbg.pack(fill=tk.X, pady=(0, 20), padx=25)
        # Barra de progreso (indicador)
        self.progress_bar = ctk.CTkFrame(bbg, fg_color=self.colors['green'], height=8, corner_radius=4)
        self.progress_bar.place(x=0, y=0, relwidth=0, relheight=1)

        # Tiempos
        self.time_elapsed = ctk.CTkLabel(sc, text="Transcurrido: --:--:--", font=('Inter', 14), fg_color="transparent", text_color=self.colors['text'])
        self.time_elapsed.pack(anchor='w', pady=2, padx=25)
        self.time_remaining = ctk.CTkLabel(sc, text="Restante: --:--:--", font=('Inter', 14), fg_color="transparent", text_color=self.colors['text'])
        self.time_remaining.pack(anchor='w', pady=2, padx=25)

        # Estadística de mensajes por WhatsApp
        self.stat_per_whatsapp = ctk.CTkLabel(sc, text="Mensajes por WhatsApp: --", font=('Inter', 14), fg_color="transparent", text_color=self.colors['text'])
        self.stat_per_whatsapp.pack(anchor='w', pady=(2, 25), padx=25)

        # Bloque 2: Registro de actividad
        lc = ctk.CTkFrame(parent, fg_color=self.colors['bg_log'], corner_radius=30)
        lc.pack(fill=tk.BOTH, expand=True, pady=(0, 30), padx=10)
        lc.grid_columnconfigure(0, weight=1)
        lc.grid_rowconfigure(1, weight=1)

        ltf = ctk.CTkFrame(lc, fg_color="transparent")
        ltf.grid(row=0, column=0, sticky='ew', pady=(25, 15), padx=25)
        ctk.CTkLabel(ltf, text="▶", font=('Inter', 14), fg_color="transparent", text_color=self.colors['log_info']).pack(side=tk.LEFT, padx=(0, 8))
        ctk.CTkLabel(ltf, text="Registro de actividad", font=self.fonts['log_title'], fg_color="transparent", text_color=self.colors['log_title_color']).pack(side=tk.LEFT)

        lco = ctk.CTkFrame(lc, fg_color="transparent")
        lco.grid(row=1, column=0, sticky='nsew', pady=(0, 25), padx=25)
        lco.grid_columnconfigure(0, weight=1)
        lco.grid_rowconfigure(0, weight=1)

        self.log_text = ctk.CTkTextbox(lco, fg_color=self.colors['bg_log'], text_color=self.colors['log_text'], font=self.fonts['log_text'], corner_radius=10, activate_scrollbars=True, border_width=1, border_color="#444851")
        self.log_text.grid(row=0, column=0, sticky="nsew")
        self.log_text.tag_config('success', foreground=self.colors['log_success'])
        self.log_text.tag_config('error', foreground=self.colors['log_error'])
        self.log_text.tag_config('warning', foreground=self.colors['log_warning'])
        self.log_text.tag_config('info', foreground=self.colors['log_info'])

        self.log_text.configure(state=tk.DISABLED)
        self.log("HΞЯMΞS V1 (Modern) iniciado", 'success')
        self.log("Sigue los pasos 1, 2, 3 en orden", 'info')
        if self.adb_path.get():
            self.log("ADB detectado correctamente", 'success')
        else:
            self.log("ADB no detectado automáticamente. Asegúrate de que esté en la carpeta o ejecuta INSTALAR.bat", 'warning')

    def create_stat(self, parent, label, value, color, col):
        """Crea un widget de estadística en el panel de estado."""
        box = ctk.CTkFrame(parent, fg_color="transparent")
        box.grid(row=0, column=col, sticky='nsew', padx=8)
        parent.grid_columnconfigure(col, weight=1)

        ctk.CTkLabel(box, text=label, fg_color="transparent", text_color=self.colors['text_light'], font=self.fonts['stat_label']).pack(pady=(5, 5))
        vl = ctk.CTkLabel(box, text=value, fg_color="transparent", text_color=color, font=self.fonts['stat_value'])
        vl.pack(pady=(0, 5))

        if label == "Total": self.stat_total = vl
        elif label == "Enviados": self.stat_sent = vl
        elif label == "Progreso": self.stat_progress = vl

    def _create_spinbox_widget(self, parent, textvariable, min_val=0, max_val=999, step=1):
        """Crea un widget spinbox personalizado (Entry con botones +/-)."""
        spinbox_frame = ctk.CTkFrame(parent, fg_color="transparent")

        def decrement_callback():
            try:
                val = textvariable.get()
                new_val = max(min_val, val - step)
                textvariable.set(new_val)
            except tk.TclError:
                textvariable.set(min_val)

        btn_decr = ctk.CTkButton(spinbox_frame, text="−", width=30, height=30,
                                 font=(self.fonts['setting_label'][0], 16, 'bold'),
                                 fg_color="transparent", text_color="#495057",
                                 hover_color="#e9ecef",
                                 command=decrement_callback, corner_radius=10)
        btn_decr.pack(side=tk.LEFT, padx=(0, 2))

        entry = ctk.CTkEntry(spinbox_frame, textvariable=textvariable, width=50,
                             font=self.fonts['setting_label'], corner_radius=10,
                             justify='center',
                             border_width=0,
                             fg_color="transparent")
        entry.pack(side=tk.LEFT, padx=2)

        def increment_callback():
            try:
                val = textvariable.get()
                new_val = min(max_val, val + step)
                textvariable.set(new_val)
            except tk.TclError:
                textvariable.set(min_val)

        btn_incr = ctk.CTkButton(spinbox_frame, text="+", width=30, height=30,
                                 font=(self.fonts['setting_label'][0], 16, 'bold'),
                                 fg_color="transparent", text_color="#495057",
                                 hover_color="#e9ecef",
                                 command=increment_callback, corner_radius=10)
        btn_incr.pack(side=tk.LEFT, padx=(2, 0))

        return spinbox_frame

    def create_setting(self, parent, label, var1, var2, row):
        """Crea una fila de configuración en la tarjeta de 'Tiempo'."""
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(parent, text=label, font=self.fonts['setting_label'], fg_color="transparent", text_color=self.colors['text_light']).grid(row=row, column=0, sticky='w', pady=10)

        ctrls = ctk.CTkFrame(parent, fg_color="transparent")
        ctrls.grid(row=row, column=1, sticky='e', pady=10, padx=(10, 0))

        spinbox1_min_val = 1 if label == "Delay (seg):" else 1 # Min 1 seg
        spinbox1 = self._create_spinbox_widget(ctrls, var1, min_val=spinbox1_min_val, max_val=300)
        spinbox1.pack(side=tk.LEFT, padx=(0, 8))

        if var2:
            ctk.CTkLabel(ctrls, text="-", font=self.fonts['setting_label'], fg_color="transparent").pack(side=tk.LEFT, padx=(0, 8))
            spinbox2 = self._create_spinbox_widget(ctrls, var2, min_val=1, max_val=300)
            spinbox2.pack(side=tk.LEFT)

    def log(self, msg, tag='info'):
        """Añade un mensaje al registro de actividad con formato."""
        ts = datetime.now().strftime("[%H:%M:%S]")
        icon = "✓"
        add_space_before = False

        # Asignación de iconos y espaciado
        if tag == 'error':
            icon = "✗"
            add_space_before = True
        elif tag == 'warning':
            icon = "⚠"
        elif tag == 'info':
            icon = "ℹ"
        elif tag == 'success':
            add_space_before = True # Hacer que todos los 'success' tengan espacio antes

        original_msg_key = msg

        # Traducción de mensajes técnicos a mensajes amigables (MOD 26/27)
        if "HΞЯMΞS V1" in msg: msg = "HΞЯMΞS V1 (Modern) iniciado"; add_space_before = False
        elif "Sigue los pasos" in msg: msg = "Sigue los pasos 1, 2, 3 en orden"
        elif "ADB detectado" in msg: msg = "ADB detectado correctamente"
        elif "ADB no detectado" in msg: msg = "ADB no detectado. Revisa la conexión o ejecuta INSTALAR.bat"
        elif "Detectando dispositivos..." in msg: add_space_before = True
        elif "disp:" in msg:
            try:
                count = msg.split()[1]
                devices_list = msg.split(': ')[1]
                msg = f"{count} dispositivo(s) encontrado(s): {devices_list}"
            except: pass
        elif "No encontrados." in msg: msg = "No se encontraron dispositivos conectados o autorizados"
        elif "Timeout ADB." in msg: msg = "Tiempo de espera agotado al comunicar con ADB"; add_space_before = True
        elif "Seleccionando..." in msg: msg = "Selecciona el archivo Excel/CSV"
        elif "Leyendo..." in msg: msg = "Leyendo archivo..."
        elif "Archivo sin datos" in msg: msg = "El archivo seleccionado está vacío o no tiene datos válidos"; add_space_before = True
        elif "Sin col Teléfono/Celular" in msg: msg = "No se encontró una columna llamada 'Telefono' o 'Celular'"; add_space_before = True
        elif "filas." in msg and "Cols Tel:" not in msg:
            try: count = msg.split()[1]; msg = f"{count} filas leídas del archivo"
            except: pass
        elif "Cols Tel:" in msg: msg = f"Columnas de teléfono encontradas: {msg.split(': ')[1]}"
        elif "Procesando..." in msg: msg = "Procesando datos y generando mensajes..."; add_space_before = True
        elif "URLs generados" in msg or "URLs cargados" in msg:
             try:
                 count = msg.split()[1]
                 msg_type = "generados" if "generados" in original_msg_key else "cargados"
                 msg = f"{count} mensajes {msg_type} y listos para enviar"
                 add_space_before = True
             except: pass
        elif "Excel guardado:" in msg: msg = f"Archivo procesado guardado: {os.path.basename(msg.split(': ')[1])}"; add_space_before = True
        
        elif "Fidelizado:" in msg and "generados" in msg: add_space_before = True
        elif "Fidelizado (Bucles Blast) cargado" in msg: add_space_before = True
        elif "Modo Bucles Blast:" in msg: add_space_before = True
        elif "--- Iniciando REPETICIÓN" in msg: add_space_before = True
        elif "Repetición" in msg and "Etapa" in msg: add_space_before = True
        elif "--- Fin REPETICIÓN" in msg: add_space_before = True

        elif "INICIANDO ENVÍO" in msg: msg = "🚀 INICIANDO ENVÍO DE MENSAJES"; add_space_before = True
        elif "Esperando" in msg and "s..." in msg:
            try:
                delay_str = msg.split()[1]
                delay_float = float(delay_str)
                msg = f"⏳ Pausa de {delay_float:.1f}s... {msg.split(')')[1] if ')' in msg else ''}" # Mantener post-tarea
            except: pass
        elif "→" in msg and "Usando" not in msg:
             try:
                 parts = msg.split('→')
                 count_part = parts[0].strip() # FIX: Tomar todo antes de '→'
                 num_part = parts[1].strip()
                 # MOD: Distinguir log de grupo
                 if "Grupo (" in num_part:
                     msg = f"{count_part} → {num_part} (Grupo)"
                 else:
                     msg = f"{count_part} → {num_part} (Número)"
             except: pass
        elif "Abriendo link" in msg: msg = f"Abriendo WhatsApp en {msg.split(' en ')[1]}..."
        elif "Escribiendo mensaje..." in msg: msg = "Escribiendo mensaje en grupo (con keyevents)..."
        elif "Mensaje enviado" in msg: pass # Mantener mensaje simple
        elif "Cerrando apps" in msg: msg = f"🧹 Limpiando aplicaciones en {msg.split(' en ')[1].split('...')[0]}"
        elif "ENVÍO FINALIZADO" in msg: msg = "✅ ENVÍO FINALIZADO"; add_space_before = True
        elif "Resumen:" in msg: msg = f"Resumen: {msg.split('Resumen: ')[1]}"; add_space_before = True
        elif "Reanudado" in msg: msg = "▶ Envío reanudado"
        elif "Pausado" in msg: msg = "⏸ Envío pausado"
        elif "Cancelando..." in msg: msg = "⏹ Cancelando envío..."
        elif "Cancelado" in msg: msg = "⚠ Envío cancelado por el usuario"; add_space_before = True
        
        # Filtrar mensajes de bajo nivel
        if "Traceback" in msg or "ADB stderr:" in msg or "ADB stdout:" in msg:
            if ("ADB stderr:" in original_msg_key or "Error ADB" in original_msg_key) and tag == 'error':
                 # Mostrar el error de ADB si ya está pre-procesado
                 if "Error ADB" in original_msg_key:
                     msg = original_msg_key # Ya está limpio
                 else:
                     # Mostrar errores genéricos de ADB pero con icono de error
                     msg = "Error de comunicación con el dispositivo (ADB)"
                 add_space_before = True
                 icon = "✗"
            else:
                return # Ocultar stdout y tracebacks genéricos

        try:
            self.log_text.configure(state=tk.NORMAL)
            if add_space_before:
                if self.log_text.index("end-1c") != "1.0": # No añadir espacio si es la primera línea
                     self.log_text.insert(tk.END, "\n")
            self.log_text.insert(tk.END, f"{ts} {icon} {msg}\n", tag)
            self.log_text.configure(state=tk.DISABLED)
            self.log_text.see(tk.END)
            self.root.update_idletasks()
        except tk.TclError:
            # Evita crash si la ventana se está cerrando
            pass

    def update_stats(self):
        """Actualiza todos los contadores y barras de progreso en la UI."""
        self.stat_total.configure(text=str(self.total_messages))
        self.stat_sent.configure(text=str(self.sent_count))

        if self.total_messages > 0:
            # Usar sent_count en lugar de current_index para el %
            prog_percent = int((self.sent_count / self.total_messages) * 100)
            
            # current_index (el que se está procesando)
            prog_label_idx = self.current_index
            
            self.stat_progress.configure(text=f"{prog_percent}%")
            self.progress_bar.place(relwidth=(prog_percent / 100))
            self.progress_label.configure(text=f"{self.sent_count}/{self.total_messages}") # Mostrar enviados/total

            if self.start_time and prog_label_idx > 0:
                el = datetime.now() - self.start_time
                self.time_elapsed.configure(text=f"Trans: {str(el).split('.')[0]}")
                
                # Calcular tiempo promedio usando lista de tiempos reales
                if len(self.task_times) > 0:
                    # Usar promedio de los últimos tiempos para mejor precisión
                    recent_times = self.task_times[-min(10, len(self.task_times)):]  # Últimos 10 o menos
                    avg = sum(recent_times) / len(recent_times)
                else:
                    # Fallback al método anterior si no hay datos
                    avg = el.total_seconds() / prog_label_idx
                
                # Calcular tiempo restante
                tasks_remaining = self.total_messages - self.sent_count
                rem_s = avg * tasks_remaining
                rem = timedelta(seconds=int(rem_s))
                self.time_remaining.configure(text=f"Rest: {str(rem).split('.')[0]}")
        else:
            self.stat_progress.configure(text="0%")
            self.progress_bar.place(relwidth=0)
            self.progress_label.configure(text="--/--")
            self.time_elapsed.configure(text="Trans: --:--:--")
            self.time_remaining.configure(text="Rest: --:--:--")

    def toggle_dark_mode(self):
        """Alterna entre modo claro y oscuro."""
        self.dark_mode = not self.dark_mode
        
        # Cambiar paleta de colores
        if self.dark_mode:
            self.colors = self.colors_dark.copy()
            ctk.set_appearance_mode("dark")
        else:
            self.colors = self.colors_light.copy()
            ctk.set_appearance_mode("light")
        
        # Actualizar hover colors
        self.hover_colors = {k: darken_color(v, 0.18) for k, v in self.colors.items() if k.startswith('action_')}
        
        # Recrear la interfaz
        # Destruir widgets existentes
        for widget in self.root.winfo_children():
            widget.destroy()
        
        # Recrear la interfaz
        self.setup_ui()
        
        # Actualizar texto del botón
        if hasattr(self, 'dark_mode_btn') and self.dark_mode_btn:
            if self.dark_mode:
                self.dark_mode_btn.configure(text="Modo Claro")
            else:
                self.dark_mode_btn.configure(text="Modo Oscuro")
        
        self.log(f"Modo {'Oscuro' if self.dark_mode else 'Claro'} activado", 'info')

    def toggle_additional_actions(self):
        """Muestra u oculta los botones adicionales."""
        self.actions_expanded = not self.actions_expanded
        
        if self.actions_expanded:
            # Mostrar los botones en la misma línea
            self.additional_actions_frame.pack(side=tk.LEFT, padx=(0, 0))
            self.toggle_actions_btn.configure(text="▲")
        else:
            # Ocultar los botones
            self.additional_actions_frame.pack_forget()
            self.toggle_actions_btn.configure(text="▼")

    def update_per_whatsapp_stat(self, *args):
        """Calcula y actualiza la estadística de mensajes por cuenta de WhatsApp."""
        num_devices = len(self.devices)
        if not self.links or self.manual_mode or num_devices == 0:
            self.stat_per_whatsapp.configure(text="Mensajes por WhatsApp: --")
            return

        mode = self.traditional_send_mode.get()
        base_links = len(self.links)
        stat_text = "--"

        if mode == "Business":
            per_account = base_links / num_devices
            stat_text = f"~{round(per_account)} (Business)"
        elif mode == "Normal":
            per_account = base_links / num_devices
            stat_text = f"~{round(per_account)} (Normal)"
        elif mode == "Business/Normal":
            # Total messages are split between Business and Normal
            b_total = (base_links + 1) // 2
            n_total = base_links // 2
            # Then distributed among devices
            b_per_account = b_total / num_devices
            n_per_account = n_total / num_devices
            stat_text = f"~{round(b_per_account)} (B) / ~{round(n_per_account)} (N)"
        elif mode == "B/N.1/N.2":
            # Total messages are split among B, N1, N2
            b_total = (base_links + 2) // 3
            n1_total = (base_links + 1) // 3
            n2_total = base_links - b_total - n1_total
            # Then distributed among devices
            b_per_account = b_total / num_devices
            n1_per_account = n1_total / num_devices
            n2_per_account = n2_total / num_devices
            stat_text = f"~{round(b_per_account)}(B), ~{round(n1_per_account)}(N1), ~{round(n2_per_account)}(N2)"

        self.stat_per_whatsapp.configure(text=f"Mensajes por WhatsApp: {stat_text}")

    def auto_detect_adb(self):
        """Busca adb.exe en las carpetas comunes del proyecto."""
        paths = [
            os.path.join(BASE_DIR, "scrcpy-win64-v3.2", "adb.exe"),
            os.path.join(BASE_DIR, "adb.exe")
        ]
        for p in paths:
            if os.path.exists(p):
                self.adb_path.set(p)
                break

    def detect_devices(self):
        """Ejecuta 'adb devices' y actualiza la lista de dispositivos."""
        adb = self.adb_path.get()
        if not adb or not os.path.exists(adb):
            self.auto_detect_adb()
            adb = self.adb_path.get()
        if not adb or not os.path.exists(adb):
            messagebox.showerror("Error", "ADB no encontrado.\nAsegúrate de que 'adb.exe' esté en la carpeta del proyecto o en 'scrcpy-win64-v3.2'.")
            return

        self.log("Detectando dispositivos...", 'info')
        try:
            si = subprocess.STARTUPINFO()
            si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            si.wShowWindow = subprocess.SW_HIDE

            # --- CORRECCIÓN: Usar lista de argumentos para evitar problemas con paths ---
            res = subprocess.run([adb, 'devices'], capture_output=True, text=True, timeout=10, startupinfo=si, check=False)
            # --- FIN CORRECCIÓN ---

            self.devices = [l.split('\t')[0] for l in res.stdout.strip().split('\n')[1:] if '\tdevice' in l]

            # Actualizar las etiquetas de la UI
            self._update_device_labels()

            if self.devices:
                self.log(f"✓ {len(self.devices)} disp: {', '.join(self.devices)}", 'success')
                messagebox.showinfo("Dispositivos", f"{len(self.devices)} dispositivo(s) econtrado(s):\n\n" + "\n".join(self.devices))
            else:
                self.log("No encontrados.", 'error')
                messagebox.showwarning("Dispositivos", "No se encontraron dispositivos.\nAsegúrate de conectar tu teléfono, activar la 'Depuración USB' y autorizar la conexión en el móvil.")
        except subprocess.TimeoutExpired:
            self.log("Timeout ADB.", 'error')
            messagebox.showerror("Error", "Timeout ADB.")
        except Exception as e:
            self.log(f"Error al detectar: {e}", 'error')
            messagebox.showerror("Error", f"Error: {e}")

    # --- Lógica de archivos ---
    def read_csv_file(self, fp):
        """Lee un archivo CSV intentando con múltiples codificaciones y delimitadores."""
        try:
            encs = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252', 'iso-8859-1', 'utf-16']
            for enc in encs:
                try:
                    with open(fp, 'r', encoding=enc, errors='ignore') as f:
                        s = f.read(2048) # Leer una muestra para detectar delimitador
                        f.seek(0)
                        dls = [';', ',', '\t', '|']
                        d = ',' # Default
                        for dl in dls:
                            if dl in s:
                                d = dl
                                break
                        r = csv.DictReader(f, delimiter=d)
                        data = [{k.strip(): (v if v is not None else '') for k, v in rw.items() if k is not None} for rw in r]
                        fn = [n.strip() for n in r.fieldnames if n is not None] if r.fieldnames else []
                        return data, fn
                except Exception:
                    continue
            raise Exception("No se pudo leer el CSV con las codificaciones probadas.")
        except Exception as e:
            raise Exception(f"Error al leer CSV: {e}")

    def read_excel_file(self, fp):
        """Lee un archivo Excel (xlsx/xls) y lo convierte en una lista de diccionarios."""
        try:
            wb = load_workbook(fp, data_only=True) # data_only=True para obtener valores de fórmulas
            sh = wb.active
            hdrs = [str(c.value).strip() if c.value is not None else '' for c in sh[1]] # Fila 1 = cabeceras

            # Mapeo de cabeceras válidas (índice, nombre)
            vh = [(i, h) for i, h in enumerate(hdrs) if h]
            if not vh:
                raise ValueError("No se encontraron cabeceras válidas en la fila 1.")

            data = []
            for r_idx, r in enumerate(sh.iter_rows(min_row=2, values_only=True), start=2):
                rd = {}
                for c_idx, hn in vh:
                    v = r[c_idx]
                    pv = ''
                    if v is None:
                        pv = ''
                    elif isinstance(v, (int, float)):
                        pv = str(v)
                    elif isinstance(v, datetime):
                        pv = v.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        pv = str(v)
                    rd[hn] = pv

                if any(rd.values()): # Solo añadir fila si tiene algún dato
                    data.append(rd)

            vhn = [h for i, h in vh] # Nombres de cabeceras válidas
            return data, vhn
        except Exception as e:
            raise Exception(f"Error al leer Excel: {e}")

    def load_and_process_excel(self):
        """Abre el diálogo para cargar Excel/CSV e inicia el procesamiento."""
        self.log("Seleccionando...", 'info')
        self.manual_mode = False  # Modo tradicional (Excel/CSV)
        self.group_mode = False 
        self.fidelizado_mode = None  # No usar modo fidelizado
        
        fp = filedialog.askopenfilename(
            title="Seleccionar archivo Excel/CSV",
            filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv"), ("Todos", "*.*")]
        )
        if not fp:
            return

        try:
            self.log("Leyendo...", 'info')
            is_csv = fp.lower().endswith('.csv')
            self.raw_data, self.columns = self.read_csv_file(fp) if is_csv else self.read_excel_file(fp)

            if not self.raw_data:
                self.log("Archivo sin datos.", 'warning'); messagebox.showwarning("Vacío", "El archivo seleccionado está vacío o no tiene datos válidos."); return

            # Ordenar por '$ Asig.' si la columna existe
            if '$ Asig.' in self.columns:
                try:
                    # Función auxiliar para convertir el valor monetario a número flotante
                    def to_float(value):
                        if isinstance(value, (int, float)):
                            return float(value)
                        # Limpiar el string de símbolos y comas
                        cleaned_value = str(value).replace('$', '').replace(',', '').strip()
                        return float(cleaned_value) if cleaned_value else 0.0

                    # Ordenar la lista de diccionarios (las filas)
                    self.raw_data.sort(key=lambda row: to_float(row.get('$ Asig.', 0)), reverse=True)
                    self.log("✓ Filas ordenadas por '$ Asig.' de mayor a menor.", 'success')
                except (ValueError, TypeError) as e:
                    self.log(f"Advertencia: No se pudo ordenar por '$ Asig.'. Error: {e}", 'warning')
                    messagebox.showwarning("Error de Ordenamiento",
                                           f"No se pudo ordenar por la columna '$ Asig.'.\n"
                                           f"Asegúrate de que los valores sean numéricos.\n\nError: {e}",
                                           parent=self.root)

            # Caso 1: El archivo ya tiene una columna 'URL'
            if 'URL' in self.columns or 'url' in self.columns:
                uc = 'URL' if 'URL' in self.columns else 'url'
                self.links = [r[uc] for r in self.raw_data if r.get(uc) and r[uc].startswith("http")]
                if self.links:
                    self.total_messages = len(self.links)
                    self.update_stats()
                    self.log(f"✓ {len(self.links)} URLs cargados directamente", 'success')
                    messagebox.showinfo("Cargado", f"Se cargaron {len(self.links)} URLs directamente desde la columna '{uc}'.\nNo se requiere procesamiento.")
                    return

            # Caso 2: El archivo necesita procesamiento
            self.phone_columns = [c for c in self.columns if c and ('telefono' in c.lower() or 'celular' in c.lower())]
            if not self.phone_columns:
                self.log("Sin col Teléfono/Celular.", 'error'); messagebox.showerror("Error", "No se encontró ninguna columna llamada 'Telefono' o 'Celular' en el archivo."); return

            self.log(f"✓ {len(self.raw_data)} filas.", 'success')
            self.log(f"✓ Cols Tel: {', '.join(self.phone_columns)}", 'success')
            self.links = []
            self.total_messages = 0
            self.update_stats()
            self.open_processor_window(fp)

        except Exception as e:
            self.log(f"Error al leer archivo: {e}", 'error'); messagebox.showerror("Error", f"Error al leer el archivo:\n{e}")

    # --- Lógica de Fidelizado (Carga Manual) ---


    def _load_default_messages(self):
        """Carga los mensajes predeterminados desde Grupos.txt si existe."""
        try:
            # Buscar Grupos.txt en el directorio del script
            grupos_path = os.path.join(BASE_DIR, "Grupos.txt")
            if os.path.exists(grupos_path):
                with open(grupos_path, 'r', encoding='utf-8') as f:
                    lines = [ln.strip() for ln in f.read().splitlines() if ln.strip()]
                if lines:
                    self.manual_messages_numbers = lines
                    self.manual_messages_groups = lines
                    # Generar índice de inicio aleatorio
                    self.mensaje_start_index = random.randint(0, len(lines) - 1)
                    self.log(f"Mensajes predeterminados cargados: {len(lines)} mensajes (inicio aleatorio en posición {self.mensaje_start_index + 1})", 'success')
                    return True
        except Exception as e:
            self.log(f"No se pudieron cargar mensajes predeterminados: {e}", 'warning')
        return False

    def _load_fidelizado_messages_from_file(self):
        """Abre un diálogo para cargar un archivo .txt de mensajes para el modo Fidelizado."""
        filepath = filedialog.askopenfilename(
            title="Seleccionar archivo de Mensajes (.txt)",
            filetypes=[("Text Files", "*.txt"), ("All files", "*.*")],
            parent=self.root
        )
        if not filepath:
            return

        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                lines = [line.strip() for line in f.readlines() if line.strip()]

            if not lines:
                messagebox.showwarning("Archivo Vacío", "El archivo seleccionado está vacío.", parent=self.root)
                return

            # Actualizar mensajes para ambos modos
            self.manual_messages_numbers = lines
            self.manual_messages_groups = lines

            # Actualizar la UI
            self.fidelizado_message_count_label.configure(text=f"✅ {len(lines)} mensajes cargados")
            self.log(f"Cargados {len(lines)} nuevos mensajes para Fidelizado desde {os.path.basename(filepath)}", 'success')
            messagebox.showinfo("Éxito", f"Se cargaron {len(lines)} mensajes.", parent=self.root)

        except Exception as e:
            self.log(f"Error al cargar archivo de mensajes: {e}", 'error')
            messagebox.showerror("Error de Lectura", f"No se pudo leer el archivo:\n{e}", parent=self.root)

    def handle_fidelizado_access(self):
        """Manejador del botón de Fidelizado (acceso directo)."""
        # Si la lógica de contraseña sigue siendo necesaria, se puede añadir aquí.
        # Por ahora, simplemente muestra la vista.
        self.show_fidelizado_view()

    def setup_fidelizado_view(self, parent):
        """Construye la interfaz de la vista Fidelizado."""
        # Cargar mensajes predeterminados si es la primera vez
        if not self.manual_messages_numbers and not self.manual_messages_groups:
            self._load_default_messages()

        # Contenedor principal de la vista Fidelizado
        fidelizado_container = ctk.CTkFrame(parent, fg_color="transparent")
        fidelizado_container.pack(fill=tk.BOTH, expand=True)

        # Contenido principal de Fidelizado
        content = ctk.CTkFrame(fidelizado_container, fg_color=self.colors['bg_card'], corner_radius=30)
        content.pack(fill=tk.BOTH, expand=True, padx=0, pady=(10, 20))

        # Layout principal reconfigurado para una columna expandible
        content.grid_columnconfigure(0, weight=1)
        content.grid_rowconfigure(1, weight=1) # Fila de contenido principal se expande

        # --- Fila 0: Header (Volver y Título) ---
        header_frame = ctk.CTkFrame(content, fg_color="transparent")
        header_frame.grid(row=0, column=0, sticky="ew", padx=30, pady=(15, 10))

        back_button = ctk.CTkButton(header_frame, text="Volver al modo Masivos",
                                      command=self.show_traditional_view,
                                      fg_color="transparent",
                                      text_color=self.colors['text_light'],
                                      hover_color=self.colors['bg'])
        back_button.pack(side=tk.LEFT)

        ctk.CTkLabel(header_frame, text="Fidelizado", font=('Inter', 26, 'bold'), text_color=self.colors['text']).pack(side=tk.LEFT, padx=20)


        # --- Fila 1: Contenido Principal ---
        self.fidelizado_main_content_frame = ctk.CTkFrame(content, fg_color="transparent")
        self.fidelizado_main_content_frame.grid(row=1, column=0, sticky="nsew", padx=30, pady=(0, 20))
        self.fidelizado_main_content_frame.grid_columnconfigure(0, weight=55) # Columna de inputs
        self.fidelizado_main_content_frame.grid_columnconfigure(1, weight=45) # Columna de controles
        self.fidelizado_main_content_frame.grid_rowconfigure(0, weight=1)

        # --- Columna Izquierda: Inputs (Números y Grupos) ---
        self.fidelizado_inputs_col = ctk.CTkFrame(self.fidelizado_main_content_frame, fg_color="transparent")
        self.fidelizado_inputs_col.grid(row=0, column=0, sticky="nsew", padx=(0, 20))
        self.fidelizado_inputs_col.grid_rowconfigure(0, weight=1) # Permitir que el frame interno crezca
        self.fidelizado_inputs_col.grid_columnconfigure(0, weight=1)

        # Frame para los textboxes
        self.fidelizado_inputs_container = ctk.CTkFrame(self.fidelizado_inputs_col, fg_color="transparent")
        self.fidelizado_inputs_container.grid(row=0, column=0, sticky="nsew")
        self.fidelizado_inputs_container.grid_rowconfigure(0, weight=1)
        self.fidelizado_inputs_container.grid_rowconfigure(1, weight=1) # Los textboxes se expanden

        # Widgets de Números (MODIFICADO para el nuevo modo automático)
        self.fidelizado_numbers_frame = ctk.CTkFrame(self.fidelizado_inputs_container, fg_color="transparent")
        self.fidelizado_numbers_frame.grid_columnconfigure(0, weight=1)

        # Widgets de Grupos (Label y Textbox)
        self.fidelizado_groups_frame = ctk.CTkFrame(self.fidelizado_inputs_container, fg_color="transparent")
        self.fidelizado_groups_frame.grid_columnconfigure(0, weight=1) # Hacer que el textbox se expanda horizontalmente
        ctk.CTkLabel(self.fidelizado_groups_frame, text="Links de Grupos (https://...)", font=('Inter', 16, 'bold'), text_color=self.colors['text']).grid(row=0, column=0, sticky='w', pady=(0, 10))
        self.fidelizado_groups_text = ctk.CTkTextbox(self.fidelizado_groups_frame, font=('Inter', 14), corner_radius=10, border_width=1, border_color="#cccccc", wrap=tk.WORD, height=150) # Altura más controlada
        self.fidelizado_groups_text.grid(row=1, column=0, sticky="ew")

        # --- Columna Derecha: Controles de Envío ---
        self.fidelizado_controls_col = ctk.CTkFrame(self.fidelizado_main_content_frame, fg_color="transparent")
        self.fidelizado_controls_col.grid(row=0, column=1, sticky="nsew", padx=(20, 0))
        self.fidelizado_controls_col.grid_columnconfigure(0, weight=1)

        # Card para Detección de Dispositivos
        device_card = ctk.CTkFrame(self.fidelizado_controls_col, fg_color=self.colors['bg'], corner_radius=15)
        device_card.grid(row=0, column=0, sticky="ew", pady=(0, 20))
        ctk.CTkLabel(device_card, text="📱 Dispositivos", font=('Inter', 16, 'bold'), text_color=self.colors['text']).pack(anchor='w', padx=20, pady=(15, 10))
        device_container = ctk.CTkFrame(device_card, fg_color="transparent")
        device_container.pack(fill="x", padx=20, pady=(0, 20))

        self.fidelizado_detect_btn = ctk.CTkButton(device_container, text="🔍 Detectar Dispositivos",
                                                  command=self.detect_devices,
                                                  font=self.fonts['button'],
                                                  fg_color=self.colors['action_detect'],
                                                  hover_color=self.hover_colors['action_detect'],
                                                  height=40)
        self.fidelizado_detect_btn.pack(fill='x', pady=(0, 15))

        self.fidelizado_device_list_label = ctk.CTkLabel(device_container, text="No hay dispositivos detectados.",
                                                        font=self.fonts['setting_label'],
                                                        text_color=self.colors['text_light'],
                                                        wraplength=350,
                                                        justify='left')
        self.fidelizado_device_list_label.pack(anchor='w')

        # Card para Configuración
        config_card = ctk.CTkFrame(self.fidelizado_controls_col, fg_color=self.colors['bg'], corner_radius=15)
        config_card.grid(row=1, column=0, sticky="ew", pady=(0, 20))
        ctk.CTkLabel(config_card, text="⚙️ Configuración", font=('Inter', 16, 'bold'), text_color=self.colors['text']).pack(anchor='w', padx=20, pady=(15, 10))

        config_grid = ctk.CTkFrame(config_card, fg_color="transparent")
        config_grid.pack(fill=tk.X, padx=20, pady=(0, 20))
        config_grid.grid_columnconfigure([0, 1], weight=1)

        # Fila 0: Modo de envío
        mode_container = ctk.CTkFrame(config_grid, fg_color="transparent")
        mode_container.grid(row=0, column=0, columnspan=2, sticky='ew', pady=(0, 15))
        ctk.CTkLabel(mode_container, text="Modo de envío:", font=self.fonts['button'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(0, 10))
        fidelizado_modes = ["Modo Números", "Modo Grupos"]
        mode_map_to_ui = {"NUMEROS": "Modo Números", "GRUPOS": "Modo Grupos", "MIXTO": "Modo Mixto"}
        current_mode_ui = mode_map_to_ui.get(self.fidelizado_mode, "Modo Números")
        self.fidelizado_mode_var = tk.StringVar(value=current_mode_ui)
        mode_menu = ctk.CTkOptionMenu(mode_container, variable=self.fidelizado_mode_var, values=fidelizado_modes, font=self.fonts['button'], dropdown_font=self.fonts['setting_label'], fg_color=self.colors['bg_card'], button_color=self.colors['blue'], button_hover_color=darken_color(self.colors['blue'], 0.15), text_color=self.colors['text'], height=35)
        mode_menu.pack(side=tk.LEFT, expand=True, fill=tk.X)

        # Fila 1: Controles para el modo "Uno a uno" vs "Uno a muchos"
        self.numeros_mode_container = ctk.CTkFrame(config_grid, fg_color="transparent")
        ctk.CTkLabel(self.numeros_mode_container, text="Modo de Conversación:", font=self.fonts['button'], text_color=self.colors['text']).pack(side=tk.LEFT, anchor="w", padx=(0,10))
        ctk.CTkRadioButton(self.numeros_mode_container, text="Uno a uno", variable=self.fidelizado_numeros_mode, value="Uno a uno", font=self.fonts['setting_label'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(15, 10))
        ctk.CTkRadioButton(self.numeros_mode_container, text="Uno a muchos", variable=self.fidelizado_numeros_mode, value="Uno a muchos", font=self.fonts['setting_label'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=10)

        # Fila 2: Bucles y Ciclos
        self.loops_container = ctk.CTkFrame(config_grid, fg_color="transparent")
        self.loops_container.grid(row=2, column=0, sticky='ew', pady=(0, 15), padx=(0, 10))
        ctk.CTkLabel(self.loops_container, text="Bucle:", font=self.fonts['button'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(0, 10))
        self.manual_loops_var = tk.IntVar(value=max(1, self.manual_loops))
        spinbox_loops = self._create_spinbox_widget(self.loops_container, self.manual_loops_var, min_val=1, max_val=100)
        spinbox_loops.pack(side=tk.LEFT, expand=True, fill=tk.X)

        # NUEVO: Contenedor para Ciclos (inicialmente oculto)
        self.cycles_container = ctk.CTkFrame(config_grid, fg_color="transparent")
        ctk.CTkLabel(self.cycles_container, text="Ciclos:", font=self.fonts['button'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(0, 10))
        spinbox_cycles = self._create_spinbox_widget(self.cycles_container, self.manual_cycles_var, min_val=1, max_val=100)
        spinbox_cycles.pack(side=tk.LEFT, expand=True, fill=tk.X)

        # Fila 3: Velocidad y WhatsApp
        speed_container = ctk.CTkFrame(config_grid, fg_color="transparent")
        speed_container.grid(row=3, column=0, columnspan=2, sticky='ew', pady=(0, 15))
        ctk.CTkLabel(speed_container, text="Velocidad escritura:", font=self.fonts['button'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(0, 10))
        speed_menu = ctk.CTkSegmentedButton(speed_container, variable=self.write_speed, values=["Lento", "Normal", "Rápido"], font=self.fonts['button_small'], height=35, fg_color=self.colors['bg_card'], selected_color=self.colors['blue'], selected_hover_color=darken_color(self.colors['blue'], 0.15), unselected_color=self.colors['bg_card'], unselected_hover_color=self.colors["bg"], text_color=self.colors['text'])
        speed_menu.pack(side=tk.LEFT, expand=True, fill=tk.X)

        whatsapp_container = ctk.CTkFrame(config_grid, fg_color="transparent")
        whatsapp_container.grid(row=4, column=0, columnspan=2, sticky='ew', pady=(0, 15))
        ctk.CTkLabel(whatsapp_container, text="WhatsApp a usar:", font=self.fonts['button'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(0, 10))
        self.fidelizado_whatsapp_menu = ctk.CTkSegmentedButton(whatsapp_container, variable=self.whatsapp_mode, values=["Normal", "Business", "Ambas", "Todas"], font=self.fonts['button_small'], height=35, fg_color=self.colors['bg_card'], selected_color=self.colors['green'], selected_hover_color=darken_color(self.colors['green'], 0.15), unselected_color=self.colors['bg_card'], unselected_hover_color=self.colors["bg"], text_color=self.colors['text'])
        self.fidelizado_whatsapp_menu.pack(side=tk.LEFT, expand=True, fill=tk.X)

        # --- Fila 5: Retardo entre envíos ---
        delay_container = ctk.CTkFrame(config_grid, fg_color="transparent")
        delay_container.grid(row=5, column=0, columnspan=2, sticky='ew', pady=(0, 15))
        ctk.CTkLabel(delay_container, text="Retardo Envíos (s):", font=self.fonts['button'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(0, 10))
        spinbox_frame = ctk.CTkFrame(delay_container, fg_color="transparent")
        spinbox_frame.pack(side=tk.RIGHT)
        spinbox_max = self._create_spinbox_widget(spinbox_frame, self.fidelizado_send_delay_max, min_val=1, max_val=300)
        spinbox_max.pack(side=tk.RIGHT)
        ctk.CTkLabel(spinbox_frame, text="-", font=self.fonts['setting_label'], fg_color="transparent").pack(side=tk.RIGHT, padx=8)
        spinbox_min = self._create_spinbox_widget(spinbox_frame, self.fidelizado_send_delay_min, min_val=1, max_val=300)
        spinbox_min.pack(side=tk.RIGHT)

        # Card para Mensajes
        self.messages_card = ctk.CTkFrame(self.fidelizado_controls_col, fg_color=self.colors['bg'], corner_radius=15)
        self.messages_card.grid(row=2, column=0, sticky="ew", pady=(0, 20))
        ctk.CTkLabel(self.messages_card, text="✍️ Mensajes", font=('Inter', 16, 'bold'), text_color=self.colors['text']).pack(anchor='w', padx=20, pady=(15, 10))
        self.fidelizado_messages_container = ctk.CTkFrame(self.messages_card, fg_color="transparent")
        self.fidelizado_messages_container.pack(fill="x", padx=20, pady=(0, 20))
        self.fidelizado_messages_container.grid_columnconfigure(1, weight=1)

        load_messages_btn = ctk.CTkButton(self.fidelizado_messages_container, text="Cargar Archivo",
                                          command=self._load_fidelizado_messages_from_file,
                                          font=self.fonts['button'],
                                          fg_color=self.colors['blue'],
                                          hover_color=darken_color(self.colors['blue'], 0.15),
                                          height=35)
        load_messages_btn.grid(row=0, column=0, sticky='w')

        self.fidelizado_message_count_label = ctk.CTkLabel(self.fidelizado_messages_container, text="", font=self.fonts['setting_label'], text_color=self.colors['text'])
        self.fidelizado_message_count_label.grid(row=0, column=1, sticky='w', padx=15)

        initial_message_count = len(self.manual_messages_numbers)
        if initial_message_count > 0:
            self.fidelizado_message_count_label.configure(text=f"✅ {initial_message_count} mensajes cargados")
        else:
            self.fidelizado_message_count_label.configure(text="⚠️ No hay mensajes cargados")

        # --- Controles de Variante Mixto (inicialmente ocultos) ---
        self.mixto_variant_container = ctk.CTkFrame(config_grid, fg_color="transparent")
        ctk.CTkLabel(self.mixto_variant_container, text="Variante Modo Mixto:", font=self.fonts['setting_label'], text_color=self.colors['text']).pack(anchor='w', pady=(0, 8))
        mixto_radio_frame = ctk.CTkFrame(self.mixto_variant_container, fg_color="transparent")
        mixto_radio_frame.pack(anchor='w')
        ctk.CTkRadioButton(mixto_radio_frame, text="1G:1N", variable=self.mixto_variant, value=1, font=self.fonts['setting_label'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(0, 15))
        ctk.CTkRadioButton(mixto_radio_frame, text="2G:1N", variable=self.mixto_variant, value=2, font=self.fonts['setting_label'], text_color=self.colors['text']).pack(side=tk.LEFT, padx=(0, 15))
        ctk.CTkRadioButton(mixto_radio_frame, text="3G:1N", variable=self.mixto_variant, value=3, font=self.fonts['setting_label'], text_color=self.colors['text']).pack(side=tk.LEFT)

        # --- Botones de Acción ---
        self.actions_frame = ctk.CTkFrame(self.fidelizado_controls_col, fg_color="transparent")
        self.actions_frame.grid(row=3, column=0, sticky="sew", pady=(15, 0))
        self.actions_frame.grid_columnconfigure(0, weight=1)
        self.actions_frame.grid_columnconfigure(1, weight=1)

        self.fidelizado_btn_start = ctk.CTkButton(self.actions_frame, text="▶ INICIAR ENVÍO FIDELIZADO", command=self.start_fidelizado_sending, fg_color=self.colors['action_start'], hover_color=self.hover_colors['action_start'], text_color=self.colors['text_header_buttons'], font=self.fonts['button'], corner_radius=10, height=50)
        self.fidelizado_btn_start.grid(row=0, column=0, sticky='ew', padx=(0, 10))

        self.unirse_grupos_btn = ctk.CTkButton(self.actions_frame, text="🔗 UNIRSE A GRUPOS", command=self.start_unirse_grupos, fg_color=self.colors['action_detect'], hover_color=self.hover_colors['action_detect'], text_color=self.colors['text_header_buttons'], font=self.fonts['button'], corner_radius=10, height=50)
        self.unirse_grupos_btn.grid(row=0, column=1, sticky='ew', padx=(10, 0))

        # --- Botones de Control (Pausa/Cancelar) ---
        self.control_buttons_frame = ctk.CTkFrame(self.fidelizado_controls_col, fg_color="transparent")
        self.control_buttons_frame.grid(row=4, column=0, sticky="ew", pady=(10, 0))
        self.control_buttons_frame.grid_columnconfigure(0, weight=1)
        self.control_buttons_frame.grid_columnconfigure(1, weight=1)

        self.fidelizado_btn_pause = ctk.CTkButton(self.control_buttons_frame, text="⏸  PAUSAR", command=self.pause_sending, fg_color=self.colors['action_pause'], hover_color=self.hover_colors['action_pause'], text_color=self.colors['text_header_buttons'], font=self.fonts['button'], corner_radius=10, height=45)
        self.fidelizado_btn_pause.grid(row=0, column=0, sticky='ew', padx=(0, 10))

        self.fidelizado_btn_stop = ctk.CTkButton(self.control_buttons_frame, text="⏹  CANCELAR", command=self.stop_sending, fg_color=self.colors['action_cancel'], hover_color=self.hover_colors['action_cancel'], text_color=self.colors['text_header_buttons'], font=self.fonts['button'], corner_radius=10, height=45)
        self.fidelizado_btn_stop.grid(row=0, column=1, sticky='ew', padx=(10, 0))

        # Ocultar el frame de botones de control inicialmente
        self.control_buttons_frame.grid_remove()

        # --- Lógica de la Vista ---
        self.fidelizado_mode_var.trace_add('write', self._update_fidelizado_ui_mode)
        self.fidelizado_numeros_mode.trace_add('write', self._update_fidelizado_ui_mode) # NUEVO: trace para los radio buttons
        self._update_fidelizado_ui_mode()
        self._populate_fidelizado_inputs()

    def _update_fidelizado_ui_mode(self, *args):
        """Muestra u oculta los widgets y reorganiza el layout según el modo Fidelizado seleccionado."""
        mode_ui = self.fidelizado_mode_var.get()
        numeros_submode = self.fidelizado_numeros_mode.get()

        mode_map_from_ui = {"Modo Números": "NUMEROS", "Modo Grupos": "GRUPOS", "Modo Mixto": "MIXTO"}
        self.fidelizado_mode = mode_map_from_ui.get(mode_ui)

        show_numeros_mode = self.fidelizado_mode == "NUMEROS"
        show_mixto_variant = self.fidelizado_mode == "MIXTO"
        show_cycles = self.fidelizado_mode == "NUMEROS" and numeros_submode == "Uno a uno"

        # --- 1. Reorganización del Layout Principal ---
        if self.fidelizado_mode == "NUMEROS":
            # Layout de 1 columna: Solo mostrar los controles en la parte superior.
            self.fidelizado_main_content_frame.grid_columnconfigure(0, weight=1)
            self.fidelizado_main_content_frame.grid_columnconfigure(1, weight=0) # Anular segunda columna
            self.fidelizado_main_content_frame.grid_rowconfigure(0, weight=0) # Fila de controles no se expande
            self.fidelizado_main_content_frame.grid_rowconfigure(1, weight=0) # Fila de abajo tampoco

            self.fidelizado_inputs_col.grid_forget() # Ocultar la columna de inputs que ya no se usa
            self.fidelizado_controls_col.grid(row=0, column=0, sticky="ew", padx=0) # Mostrar solo los controles arriba
        elif self.fidelizado_mode == "GRUPOS":
            # Layout de 1 columna: Apilar controles arriba y inputs (links de grupo) abajo
            self.fidelizado_main_content_frame.grid_columnconfigure(0, weight=1)
            self.fidelizado_main_content_frame.grid_columnconfigure(1, weight=0) # Anular segunda columna
            self.fidelizado_main_content_frame.grid_rowconfigure(0, weight=0) # Fila de controles (arriba) no se expande
            self.fidelizado_main_content_frame.grid_rowconfigure(1, weight=0) # Fila de inputs (abajo) tampoco

            self.fidelizado_controls_col.grid(row=0, column=0, sticky="ew", padx=0, pady=(0, 20)) # Controles arriba
            self.fidelizado_inputs_col.grid(row=1, column=0, sticky="ew", padx=0) # Inputs (links) abajo
        else: # MODO MIXTO
            # Layout de 2 columnas: Inputs a la izquierda, controles a la derecha
            self.fidelizado_main_content_frame.grid_columnconfigure(0, weight=55)
            self.fidelizado_main_content_frame.grid_columnconfigure(1, weight=45)
            self.fidelizado_main_content_frame.grid_rowconfigure(1, weight=0) # Anular segunda fila

            self.fidelizado_inputs_col.grid(row=0, column=0, sticky="nsew", padx=(0, 20))
            self.fidelizado_controls_col.grid(row=0, column=1, sticky="nsew", padx=(20, 0))

        # --- 2. Visibilidad de Widgets de Input ---
        if self.fidelizado_mode == "NUMEROS":
            self.fidelizado_numbers_frame.grid(row=0, column=0, sticky="nsew")
            self.fidelizado_groups_frame.grid_forget()
        elif self.fidelizado_mode == "GRUPOS":
            self.fidelizado_numbers_frame.grid_forget()
            self.fidelizado_groups_frame.grid(row=0, column=0, sticky="nsew")
        elif self.fidelizado_mode == "MIXTO":
            self.fidelizado_inputs_container.grid_rowconfigure(0, weight=1)
            self.fidelizado_inputs_container.grid_rowconfigure(1, weight=1)
            self.fidelizado_numbers_frame.grid(row=0, column=0, sticky="nsew", pady=(0, 10))
            self.fidelizado_groups_frame.grid(row=1, column=0, sticky="nsew", pady=(10, 0))


        # --- 3. Visibilidad de Configuración Adicional ---
        if show_numeros_mode:
            self.numeros_mode_container.grid(row=1, column=0, columnspan=2, sticky='w', pady=(0, 15))
        else:
            self.numeros_mode_container.grid_remove()

        if show_mixto_variant:
            self.mixto_variant_container.grid(row=6, column=0, columnspan=2, sticky='w', pady=(0, 15))
        else:
            self.mixto_variant_container.grid_remove()

        if show_cycles:
            self.loops_container.grid(row=2, column=0, sticky='ew', pady=(0, 15), padx=(0, 10))
            self.cycles_container.grid(row=2, column=1, sticky='ew', pady=(0, 15), padx=(10, 0))
        else:
            self.cycles_container.grid_remove()
            self.loops_container.grid(row=2, column=0, columnspan=2, sticky='ew', pady=(0, 15), padx=0)


        # --- 4. Visibilidad de Botones de Acción ---
        if self.fidelizado_mode == "GRUPOS":
            self.fidelizado_btn_start.grid(row=0, column=0, columnspan=1, sticky='ew', padx=(0, 5))
            self.unirse_grupos_btn.grid(row=0, column=1, sticky='ew', padx=(5, 0))
            self.actions_frame.grid_columnconfigure(1, weight=1)
            self.fidelizado_btn_start.configure(text="▶ INICIAR ENVÍO A GRUPOS")
        else:
            self.unirse_grupos_btn.grid_remove()
            self.fidelizado_btn_start.grid(row=0, column=0, columnspan=2, sticky='ew', padx=0)
            self.actions_frame.grid_columnconfigure(1, weight=0)
            self.fidelizado_btn_start.configure(text="▶ INICIAR ENVÍO FIDELIZADO")

        # --- 5. Actualización del Menú de WhatsApp ---
        if self.fidelizado_mode == "NUMEROS":
            self.fidelizado_whatsapp_menu.configure(values=["Normal", "Business", "Ambas"])
            if self.whatsapp_mode.get() == "Todas":
                self.whatsapp_mode.set("Ambas")
        else:
            if hasattr(self, 'fidelizado_whatsapp_menu'):
                 self.fidelizado_whatsapp_menu.configure(values=["Normal", "Business", "Ambas", "Todas"])

    def _populate_fidelizado_inputs(self):
        """Limpia y rellena los campos de texto con los datos guardados en las variables."""
        # Limpiar contenido existente
        self.fidelizado_groups_text.delete("1.0", tk.END)

        # Rellenar con datos guardados
        if self.manual_inputs_groups:
            self.fidelizado_groups_text.insert("1.0", "\n".join(self.manual_inputs_groups))

        # Si no hay mensajes de grupo pero sí de número (caso común), usarlos también para grupos
        if self.manual_messages_numbers and not self.manual_messages_groups:
             self.manual_messages_groups = self.manual_messages_numbers

        # --- NUEVO: Actualizar el contador de mensajes ---
        message_count = len(self.manual_messages_numbers)
        if hasattr(self, 'fidelizado_message_count_label'): # Asegurarse de que el widget exista
            if message_count > 0:
                self.fidelizado_message_count_label.configure(text=f"✅ {message_count} mensajes cargados")
            else:
                self.fidelizado_message_count_label.configure(text="⚠️ No hay mensajes cargados")

    def start_fidelizado_sending(self):
        """Función específica para validar y preparar el envío desde la vista Fidelizado."""
        # 1. Guardar los datos de los TextBoxes en las variables de la clase
        self.manual_inputs_groups = [line.strip() for line in self.fidelizado_groups_text.get("1.0", tk.END).splitlines() if line.strip()]

        # 2. Validar los datos según el modo
        if self.fidelizado_mode == "NUMEROS":
            # La validación y el inicio se manejan en un hilo para permitir la detección de números sin congelar la UI.
            threading.Thread(target=self.start_fidelizado_numeros_thread, daemon=True).start()
            return # Salir, ya que el resto del flujo se maneja en el hilo.

        if self.fidelizado_mode == "GRUPOS" and not self.manual_inputs_groups:
            messagebox.showerror("Error", "El 'Modo Grupos' requiere al menos un link de grupo.", parent=self.root)
            return
        if self.fidelizado_mode == "MIXTO" and (not self.manual_inputs_groups):
            messagebox.showerror("Error", "El 'Modo Mixto' requiere grupos.", parent=self.root)
            return
        if not self.manual_messages_numbers:
            messagebox.showerror("Error", "Se requiere al menos un mensaje.", parent=self.root)
            return

        # 3. Marcar el modo manual y llamar a la función de envío principal (SOLO para Grupos y Mixto)
        self.manual_mode = True
        self.group_mode = self.fidelizado_mode == "GRUPOS" # Flag legacy
        self.links = [] # Limpiar links del modo tradicional

        self.start_sending() # Llamar a la lógica de envío compartida

    def start_fidelizado_numeros_thread(self):
        """
        Hilo de trabajo para preparar e iniciar el envío del modo NÚMEROS automático.
        Detecta los números, calcula el total, pide confirmación y luego inicia.
        """
        if not self._detect_and_prepare_phone_lines():
            self.root.after(0, lambda: messagebox.showerror("Error", "No se detectaron suficientes líneas de WhatsApp para iniciar (se requieren al menos 2).", parent=self.root))
            return

        # --- FILTRADO DE LÍNEAS POR TIPO DE WHATSAPP ---
        whatsapp_mode = self.whatsapp_mode.get()
        original_line_count = len(self.detected_phone_lines)

        if whatsapp_mode == "Normal":
            self.detected_phone_lines = [line for line in self.detected_phone_lines if line['type'] == 'WhatsApp']
            self.log(f"Filtrando por 'Normal'. Líneas a usar: {len(self.detected_phone_lines)} de {original_line_count}", 'info')
        elif whatsapp_mode == "Business":
            self.detected_phone_lines = [line for line in self.detected_phone_lines if line['type'] == 'WhatsApp Business']
            self.log(f"Filtrando por 'Business'. Líneas a usar: {len(self.detected_phone_lines)} de {original_line_count}", 'info')
        else: # Ambas
            self.log(f"Modo 'Ambas'. Usando {len(self.detected_phone_lines)} líneas detectadas.", 'info')

        if len(self.detected_phone_lines) < 2 and self.fidelizado_numeros_mode.get() != "Uno a muchos":
             msg = f"Después de filtrar por '{whatsapp_mode}', solo quedan {len(self.detected_phone_lines)} líneas. Se requieren al menos 2 para este modo."
             self.log(msg, 'error')
             self.root.after(0, lambda: messagebox.showerror("Error", msg, parent=self.root))
             return
        # --- FIN FILTRADO ---

        # Calcular total_messages
        num_lines = len(self.detected_phone_lines)
        num_bucles = self.manual_loops_var.get()
        num_ciclos = self.manual_cycles_var.get()

        if self.fidelizado_numeros_mode.get() == "Uno a uno":
            # Cada ciclo crea N/2 parejas, y cada pareja envía 2 mensajes.
            # Esto se repite por el número de ciclos y luego por el número de bucles.
            total = (num_lines // 2) * 2 * num_ciclos * num_bucles
        else: # Uno a muchos
            # Se generan N*(N-1)/2 parejas únicas.
            # Cada pareja tiene una conversación de ida y vuelta (2 mensajes).
            # Esto se repite por el número de bucles.
            num_pairs = num_lines * (num_lines - 1) // 2
            total = num_pairs * 2 * num_bucles

        self.total_messages = total
        self.log(f"Cálculo para Modo '{self.fidelizado_numeros_mode.get()}': {self.total_messages} mensajes en total.", 'info')

        self.root.after(0, self._ask_confirmation_and_start_numeros_mode)

    def _ask_confirmation_and_start_numeros_mode(self):
        """Pide confirmación al usuario y, si se acepta, inicia el envío del modo números."""
        if not messagebox.askyesno("Confirmar Envío", f"Se detectaron {len(self.detected_phone_lines)} líneas.\n\nSe enviarán un total de {self.total_messages} mensajes en modo '{self.fidelizado_numeros_mode.get()}'.\n\n¿Deseas iniciar?", parent=self.root):
            return

        # Iniciar el envío
        self.manual_mode = True
        self.fidelizado_mode = "NUMEROS"
        self.links = []

        self._enter_task_mode()
        self.update_stats()

        threading.Thread(target=self.send_thread, daemon=True).start()

    def start_unirse_grupos(self):
        """Valida e inicia el hilo para unirse a grupos."""
        if not self.devices:
            messagebox.showerror("Error", "Paso 1: Detecta al menos un dispositivo.", parent=self.root)
            return

        grupos = [line.strip() for line in self.fidelizado_groups_text.get("1.0", tk.END).splitlines() if line.strip()]
        if not grupos:
            messagebox.showerror("Error", "Ingresa al menos un link de grupo en la caja de texto.", parent=self.root)
            return

        if not messagebox.askyesno("Confirmar Acción", f"¿Estás seguro de que deseas intentar unirte a {len(grupos)} grupo(s) en {len(self.devices)} dispositivo(s)?", parent=self.root):
            return

        # Iniciar el proceso en un hilo para no bloquear la UI
        threading.Thread(target=self.run_unirse_grupos, args=(grupos,), daemon=True).start()

    def validate_numbers(self, inputs_raw, parent_window):
        """Valida una lista de números. Devuelve lista limpia o None si hay error."""
        inputs_clean_nums = []
        for raw in inputs_raw:
            s = raw.strip()
            norm = ''.join(s.split())
            if not s: continue
            if norm.startswith('+549'): 
                messagebox.showerror("Error", "No incluyas el prefijo '+549' en los números.", parent=parent_window); return None
            if norm.startswith('+'): norm = norm[1:]
            if not norm.isdigit(): 
                messagebox.showerror("Error", f"Número inválido encontrado: {s}", parent=parent_window); return None
            inputs_clean_nums.append(norm)
        return inputs_clean_nums

    def validate_groups(self, inputs_raw, parent_window):
        """Valida una lista de links de grupo. Devuelve lista limpia o None si hay error."""
        inputs_clean_groups = []
        for raw in inputs_raw:
            s = raw.strip()
            if not s: continue
            if not (s.startswith("https://chat.whatsapp.com/") or s.startswith("http://chat.whatsapp.com/")):
                messagebox.showerror("Error", f"Link de grupo inválido encontrado:\n{s}", parent=parent_window); return None
            inputs_clean_groups.append(s)
        return inputs_clean_groups

    def generate_manual_links(self, numbers, messages, loops):
        """Genera la lista de enlaces para el modo Fidelizado (Números)."""
        if not numbers or not messages:
            return []
        
        # Lógica de "loops" para Modo Números:
        # Repite la *lista de mensajes* 'loops' veces.
        # Asigna números rotativamente a esta lista extendida de mensajes.
        
        total_messages_to_send = len(messages) * loops
        final_links = []
        
        for i in range(total_messages_to_send):
            msg = messages[i % len(messages)]
            num = numbers[i % len(numbers)] # Rota los números
            link = f"https://wa.me/549{num}?text={urllib.parse.quote(msg, safe='')}"
            final_links.append(link)
            
        return final_links

    def generate_manual_pairs(self, links_or_nums, messages, loops):
        """Genera pares (link_o_numero, mensaje) para el modo Fidelizado (Grupos)."""
        if not links_or_nums or not messages:
            return []
        
        # Lógica de "loops" para Modo Grupos:
        # Repite la *lista de mensajes* 'loops' veces.
        # Asigna grupos rotativamente a esta lista extendida de mensajes.
        
        total_messages_to_send = len(messages) * loops
        final_pairs = []

        for i in range(total_messages_to_send):
            msg = messages[i % len(messages)]
            target = links_or_nums[i % len(links_or_nums)] # Rota los grupos/números
            final_pairs.append((target, msg))
            
        return final_pairs

    # --- Lógica del Procesador de Excel ---

    def open_processor_window(self, original_file):
        """Abre la ventana para configurar la plantilla de mensajes."""
        proc_window = ctk.CTkToplevel(self.root)
        proc_window.title("Configurar Procesamiento de Excel/CSV")
        proc_window.transient(self.root)

        width, height = 900, 750
        # Centrar en la pantalla
        screen_width = proc_window.winfo_screenwidth()
        screen_height = proc_window.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        proc_window.geometry(f"{width}x{height}+{x}+{y}"); proc_window.after(100, proc_window.focus_force)

        main_cont = ctk.CTkFrame(proc_window, fg_color=self.colors['bg'], corner_radius=0)
        main_cont.pack(fill=tk.BOTH, expand=True)

        # Header
        header = ctk.CTkFrame(main_cont, fg_color=self.colors['blue'], height=80, corner_radius=0)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        ctk.CTkLabel(header, text="Configurar Procesamiento", font=('Inter', 22, 'bold'), text_color=self.colors['text_header']).pack(expand=True)

        # Contenido scrollable
        scroll_f = ctk.CTkScrollableFrame(main_cont, fg_color="transparent", corner_radius=0)
        scroll_f.pack(fill=tk.BOTH, expand=True, padx=20)

        # Tarjeta 1: Info Archivo
        step1_card = ctk.CTkFrame(scroll_f, fg_color=self.colors['bg_card'], corner_radius=15)
        step1_card.pack(fill=tk.X, padx=10, pady=(15, 15))
        ctk.CTkLabel(step1_card, text="Información del Archivo", font=self.fonts['card_title'], text_color=self.colors['text']).pack(anchor='w', padx=20, pady=(15, 10))
        ctk.CTkLabel(step1_card, text=f"📊 {os.path.basename(original_file)}", font=self.fonts['setting_label'], text_color=self.colors['text']).pack(anchor='w', padx=20, pady=(5, 5))
        ctk.CTkLabel(step1_card, text=f"📝 Filas: {len(self.raw_data)}", font=self.fonts['setting_label'], text_color=self.colors['text']).pack(anchor='w', padx=20, pady=5)
        ctk.CTkLabel(step1_card, text=f"📞 Teléfonos: {', '.join(self.phone_columns)}", font=self.fonts['setting_label'], text_color=self.colors['text']).pack(anchor='w', padx=20, pady=(5, 15))

        # Crear acordeón
        steps_data = [
            {"title": "Columnas de Teléfono", "color": self.colors['green'], "id": "step2"},
            {"title": "Columnas para Mensaje", "color": self.colors['orange'], "id": "step3"},
            {"title": "Plantilla de Mensaje", "color": self.colors['blue'], "id": "step4"}
        ]
        toggles = {}

        for i, data in enumerate(steps_data):
            step_card = ctk.CTkFrame(scroll_f, fg_color=self.colors['bg_card'], corner_radius=15)
            step_card.pack(fill=tk.X, padx=10, pady=15)

            tb = ctk.CTkFrame(step_card, fg_color="transparent", cursor='hand2')
            tb.pack(fill=tk.X, padx=20, pady=20)

            hi = ctk.CTkFrame(tb, fg_color="transparent")
            hi.pack(fill=tk.X)

            ctk.CTkLabel(hi, text=str(i+1), font=('Inter', 18, 'bold'), fg_color="transparent", text_color='#202124').pack(side=tk.LEFT, padx=(0, 12))
            ctk.CTkLabel(hi, text=data["title"], font=self.fonts['card_title'], text_color=self.colors['text']).pack(side=tk.LEFT)
            al = ctk.CTkLabel(hi, text="▼", font=('Inter', 16, 'bold'), text_color=data["color"])
            al.pack(side=tk.RIGHT, padx=10)

            cf = ctk.CTkFrame(step_card, fg_color="transparent")
            cf.pack_forget() # Oculto por defecto

            toggles[data["id"]] = {"bar": tb, "header": hi, "arrow": al, "content": cf}

            # Función de toggle
            def create_tf(content_frame, arrow_label):
                def toggle_func(event=None):
                    if content_frame.winfo_ismapped():
                        content_frame.pack_forget()
                        arrow_label.configure(text="▼")
                    else:
                        content_frame.pack(fill=tk.X, pady=(0, 20), padx=20)
                        arrow_label.configure(text="▲")
                return toggle_func

            tf = create_tf(cf, al)
            tb.bind('<Button-1>', tf)
            for w in hi.winfo_children():
                w.bind('<Button-1>', tf)

        # Rellenar Step 2 (Teléfonos)
        step2_c = toggles["step2"]["content"]
        ctk.CTkLabel(step2_c, text="Selecciona qué columnas contienen números de teléfono:", font=self.fonts['setting_label'], text_color=self.colors['text_light']).pack(anchor='w', pady=(0, 10))
        self.phone_vars = {}
        pb_frame = ctk.CTkFrame(step2_c, fg_color="transparent")
        pb_frame.pack(fill=tk.X)
        for i, pc in enumerate(self.phone_columns):
            var = tk.BooleanVar(value=(i==0)) # Marcar solo la primera por defecto
            self.phone_vars[pc] = var
            ctk.CTkCheckBox(pb_frame, text=pc, variable=var, font=self.fonts['setting_label'], text_color=self.colors['text'], border_color=self.colors['text_light'], hover_color=self.colors['bg'], fg_color=self.colors['blue']).pack(anchor='w', pady=4)

        # Rellenar Step 3 (Columnas Mensaje)
        step3_c = toggles["step3"]["content"]
        ctk.CTkLabel(step3_c, text="Selecciona las columnas que usarás en el mensaje:", font=self.fonts['setting_label'], text_color=self.colors['text_light']).pack(anchor='w', pady=(0, 10))
        self.column_vars = {}
        cg = ctk.CTkFrame(step3_c, fg_color="transparent")
        cg.pack(fill=tk.X)
        col_c, row_c = 0, 0
        for col in self.columns:
            if col and col not in self.phone_columns:
                var = tk.BooleanVar(value=False)
                self.column_vars[col] = var
                ctk.CTkCheckBox(cg, text=col, variable=var, font=self.fonts['setting_label'], text_color=self.colors['text'], border_color=self.colors['text_light'], hover_color=self.colors['bg'], fg_color=self.colors['blue']).grid(row=row_c, column=col_c, sticky='w', padx=10, pady=4)
                col_c += 1
                if col_c >= 3: # 3 columnas de checkboxes
                    col_c = 0
                    row_c += 1

        # Rellenar Step 4 (Plantilla)
        step4_c = toggles["step4"]["content"]
        ctk.CTkLabel(step4_c, text="Insertar columna:", font=self.fonts['setting_label'], text_color=self.colors['text_light']).pack(anchor='w')
        bc = ctk.CTkFrame(step4_c, fg_color="transparent") # Contenedor de botones de columna
        bc.pack(fill=tk.X, pady=(5, 10))

        ctk.CTkLabel(step4_c, text="Plantilla de Mensaje (usa {Columna}):", font=self.fonts['setting_label'], text_color=self.colors['text_light']).pack(anchor='w', pady=(10, 5))
        mt = ctk.CTkTextbox(step4_c, height=100, font=self.fonts['setting_label'], corner_radius=10, border_width=1, border_color="#cccccc", wrap=tk.WORD)
        mt.pack(fill=tk.BOTH, expand=True)

        # Previsualización
        pf = ctk.CTkFrame(step4_c, fg_color=self.colors['bg'], corner_radius=10, border_width=1, border_color=self.colors["text_light"])
        pf.pack(fill=tk.BOTH, expand=True, pady=(15, 0))
        ctk.CTkLabel(pf, text="👁️ Previsualización (basada en la primera fila):", font=('Inter', 10, 'bold'), text_color=self.colors['text_light']).pack(anchor='w', padx=10, pady=(8, 5))
        pt = ctk.CTkTextbox(pf, height=70, font=('Inter', 10), fg_color=self.colors['bg_card'], text_color='#333', corner_radius=5, wrap=tk.WORD, border_width=0)
        pt.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 8))
        pt.configure(state=tk.DISABLED)

        def update_preview(*a):
            try:
                cm = mt.get('1.0', tk.END).strip()
                pt.configure(state=tk.NORMAL)
                pt.delete('1.0', tk.END)
                if not cm:
                    pt.insert('1.0', '(Escribe un mensaje para previsualizar)')
                elif self.raw_data:
                    er = self.raw_data[0] # Primera fila de datos
                    pm = cm
                    for c in self.columns:
                        pl = f"{{{c}}}"
                        if pl in pm:
                             v = er.get(c, '')
                             v = '' if v is None else str(v)
                             # Formato especial para valores monetarios
                             if '$ Hist.' in c or '$ Asig.' in c:
                                 try: v = f"${float(str(v).replace(',','').replace('$','').strip()):,.2f}"
                                 except: v = str(v)
                             pm = pm.replace(pl, v)
                    pt.insert('1.0', pm)
                else:
                    pt.insert('1.0', '(No hay datos para previsualizar)')
                pt.configure(state=tk.DISABLED)
            except Exception:
                pass # Evitar errores durante la escritura

        mt.bind('<KeyRelease>', update_preview)
        mt.bind('<ButtonRelease>', update_preview)
        update_preview()

        def update_buttons(*a):
            """Actualiza los botones de inserción rápida."""
            [w.destroy() for w in bc.winfo_children()] # Limpiar botones anteriores
            sel = [c for c, v in self.column_vars.items() if v.get()]
            if not sel:
                ctk.CTkLabel(bc, text="(Selecciona columnas en el Paso 3 para insertarlas)", font=('Inter',10,'italic'), text_color=self.colors['text_light']).pack(anchor='w')
                return

            def ins(fn):
                mt.insert(tk.INSERT, f"{{{fn}}}")
                mt.focus()
                update_preview()

            col, row = 0, 0
            for c in sel:
                ctk.CTkButton(bc, text=c, command=lambda x=c: ins(x),
                              fg_color=self.colors['blue'], hover_color=darken_color(self.colors['blue'],0.18),
                              text_color=self.colors['text_header'], font=('Inter',9,'bold'),
                              height=30, corner_radius=10).grid(row=row, column=col, padx=3, pady=3, sticky='ew')
                col += 1
                if col >= 4: # 4 botones por fila
                    col = 0
                    row += 1

        [v.trace('w', update_buttons) for v in self.column_vars.values()]
        update_buttons()

        # Barra de botones inferior
        button_bar = ctk.CTkFrame(main_cont, fg_color="transparent", corner_radius=0, border_width=0)
        button_bar.pack(fill=tk.X, side=tk.BOTTOM, pady=0)
        btn_inner_frame = ctk.CTkFrame(button_bar, fg_color="transparent")
        btn_inner_frame.pack(fill=tk.X, padx=30, pady=20)

        def process_config():
            sp = [c for c, v in self.phone_vars.items() if v.get()]
            if not sp:
                messagebox.showwarning("Aviso", "Selecciona al menos una columna de Teléfono (Paso 2)", parent=proc_window)
                return

            sc = [c for c, v in self.column_vars.items() if v.get()]
            mtmpl = mt.get("1.0", tk.END).strip()
            if not mtmpl:
                messagebox.showwarning("Aviso", "Escribe una plantilla de Mensaje (Paso 4)", parent=proc_window)
                return

            self.log("Procesando...", 'info')
            # Cerrar ventana primero
            proc_window.destroy()
            self.root.focus_force()
            # Luego procesar datos y mostrar mensaje
            self.process_excel_data(sc, mtmpl, sp)

        def cancel_config():
            proc_window.destroy()
            self.root.focus_force()

        proc_window.protocol("WM_DELETE_WINDOW", cancel_config)
        ctk.CTkButton(btn_inner_frame, text="Cancelar", command=cancel_config, fg_color=self.colors['action_cancel'], hover_color=self.hover_colors['action_cancel'], font=self.fonts['button'], corner_radius=10, height=40).pack(side=tk.RIGHT, padx=(10, 0))
        ctk.CTkButton(btn_inner_frame, text="Procesar y Generar", command=process_config, fg_color=self.colors['action_start'], hover_color=self.hover_colors['action_start'], font=self.fonts['button'], corner_radius=10, height=40).pack(side=tk.RIGHT)

        # Abrir todos los acordeones por defecto
        self.root.update_idletasks()
        for i in range(2, 5):
            toggles[f"step{i}"]["bar"].event_generate("<Button-1>")

    def process_excel_data(self, selected_columns, message_template, selected_phones):
        """Genera la lista de URLs de WhatsApp a partir de los datos y la plantilla."""
        processed_rows = []
        for row in self.raw_data:
            # Obtener todos los números de las columnas de teléfono seleccionadas
            phone_nums = []
            for ph_col in selected_phones:
                ph_val = str(row.get(ph_col, '')) if row.get(ph_col) else ''
                # Soportar números separados por guión (ej. "111-222")
                phone_nums.extend([n.strip() for n in ph_val.split('-') if n.strip()])

            if not phone_nums:
                continue # Sin número en esta fila

            for phone in phone_nums:
                if phone and phone.strip():
                    msg = message_template
                    # Rellenar plantilla
                    for col in selected_columns:
                        pl = f"{{{col}}}"
                        if pl in msg:
                            val = row.get(col, '')
                            val = '' if val is None else str(val)
                            # Formato especial para valores monetarios
                            if '$ Hist.' in col or '$ Asig.' in col:
                                try: val = f"${float(str(val).replace(',', '').replace('$', '').strip()):,.2f}"
                                except: val = str(val)
                            msg = msg.replace(pl, val)

                    ph_clean = phone.strip()
                    enc_msg = urllib.parse.quote(msg, safe='')
                    processed_rows.append(f"https://wa.me/549{ph_clean}?text={enc_msg}")

        self.links = processed_rows
        self.total_messages = len(self.links)
        self.update_stats()
        self.log(f"{len(self.links)} URLs generados", 'success')

        if not self.manual_mode:
            self.save_processed_excel() # Ofrecer guardar solo si no es modo Fidelizado

    def save_processed_excel(self):
        """Ofrece guardar un nuevo archivo Excel solo con las URLs generadas."""
        try:
            self.root.attributes('-topmost', True) # Asegurar que el diálogo esté al frente
            out_path = filedialog.asksaveasfilename(
                parent=self.root,
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                title="Guardar Excel Procesado con URLs"
            )
            self.root.attributes('-topmost', False); self.root.focus_force() # Devolver foco

            if out_path:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "URLs"
                ws['A1'] = 'URL'
                for i, url in enumerate(self.links, 2):
                    ws[f'A{i}'] = url
                wb.save(out_path)
                self.log(f"Excel guardado: {os.path.basename(out_path)}", 'success')
                messagebox.showinfo("Éxito", f"Archivo Excel guardado con éxito.\nSe generaron {len(self.links)} URLs listos para enviar.", parent=self.root)
        except Exception as e:
            self.log(f"Error al guardar Excel: {e}", 'error')
            messagebox.showerror("Error", f"Error al guardar el archivo:\n{e}", parent=self.root)
            self.root.attributes('-topmost', False); self.root.focus_force()

    # --- Lógica de Envío (Threading y ADB) ---

    # --- INICIO MODIFICACIÓN: start_sending (Modo Bucles Blast V2) ---
    def start_sending(self):
        """Valida e inicia el hilo de envío de mensajes."""
        if not self.adb_path.get() or not os.path.exists(self.adb_path.get()):
            messagebox.showerror("Error", "ADB no encontrado.\nVe a la carpeta 'scrcpy' o ejecuta INSTALAR.bat.", parent=self.root); return
        if not self.devices:
            messagebox.showerror("Error", "Paso 1: Detecta al menos un dispositivo.", parent=self.root); return
        
        # --- Validación de Tareas ---
        # Modos que NO requieren self.links (se procesan directamente en el hilo)
        modos_sin_links = ["NUMEROS", "GRUPOS", "MIXTO"]
        
        if not self.links and self.fidelizado_mode not in modos_sin_links:
            messagebox.showerror("Error", "Paso 2 o Fidelizado: Carga datos o genera enlaces.", parent=self.root)
            return
        
        # Validaciones específicas por modo
        if self.fidelizado_mode == "NUMEROS":
            if not self.manual_inputs_numbers:
                messagebox.showerror("Error", "Modo Números requiere números cargados.", parent=self.root); return
            if not self.manual_messages_numbers:
                messagebox.showerror("Error", "Modo Números requiere mensajes cargados.", parent=self.root); return
            
            # Calcular total_messages para Modo Números según WhatsApp seleccionado
            num_dev = len(self.devices)
            num_numeros = len(self.manual_inputs_numbers)
            whatsapp_multiplier = 3 if self.whatsapp_mode.get() == "Todas" else (2 if self.whatsapp_mode.get() == "Ambas" else 1)
            self.total_messages = self.manual_loops * num_numeros * num_dev * whatsapp_multiplier
            wa_mode_str = self.whatsapp_mode.get()
            self.log(f"Modo Números ({wa_mode_str}): {self.total_messages} envíos totales ({self.manual_loops} ciclos x {num_numeros} números x {num_dev} disp. x {whatsapp_multiplier} app(s))", 'info')
        
        elif self.fidelizado_mode == "GRUPOS":
            if not self.manual_inputs_groups:
                messagebox.showerror("Error", "Modo Grupos requiere grupos cargados.", parent=self.root); return
            if not self.manual_messages_groups:
                messagebox.showerror("Error", "Modo Grupos requiere mensajes cargados.", parent=self.root); return
            
            # Calcular total_messages para Modo Grupos según WhatsApp seleccionado
            num_dev = len(self.devices)
            num_grupos = len(self.manual_inputs_groups)
            whatsapp_multiplier = 3 if self.whatsapp_mode.get() == "Todas" else (2 if self.whatsapp_mode.get() == "Ambas" else 1)
            self.total_messages = self.manual_loops * num_grupos * num_dev * whatsapp_multiplier
            wa_mode_str = self.whatsapp_mode.get()
            self.log(f"Modo Grupos ({wa_mode_str}): {self.total_messages} envíos totales ({self.manual_loops} ciclos x {num_grupos} grupos x {num_dev} disp. x {whatsapp_multiplier} app(s))", 'info')
        
        elif self.fidelizado_mode == "MIXTO":
            if not self.manual_inputs_groups or not self.manual_inputs_numbers:
                messagebox.showerror("Error", "Modo Mixto requiere Grupos Y Números cargados.", parent=self.root); return
            if not self.manual_messages_numbers:
                messagebox.showerror("Error", "Modo Mixto requiere mensajes cargados.", parent=self.root); return
            
            # Calcular total_messages para Modo Mixto según WhatsApp seleccionado
            num_dev = len(self.devices)
            num_grupos = len(self.manual_inputs_groups)
            num_numeros = len(self.manual_inputs_numbers)
            whatsapp_multiplier = 3 if self.whatsapp_mode.get() == "Todas" else (2 if self.whatsapp_mode.get() == "Ambas" else 1)
            # Por cada ciclo: cada grupo y cada número recibe mensajes de todos los dispositivos
            tasks_per_ciclo = (num_grupos + num_numeros) * num_dev * whatsapp_multiplier
            
            self.total_messages = self.manual_loops * tasks_per_ciclo
            wa_mode_str = self.whatsapp_mode.get()
            self.log(f"Modo Mixto ({wa_mode_str}): {self.total_messages} envíos totales ({self.manual_loops} ciclos x ({num_grupos} grupos + {num_numeros} nums) x {num_dev} disp. x {whatsapp_multiplier} app(s))", 'info')
        
        # (total_messages para otros modos ya está calculado)
        # --- Fin Validación ---

        if self.is_running:
            return

        if not messagebox.askyesno("Confirmar Envío", f"¿Estás seguro de que deseas iniciar el envío de {self.total_messages} mensajes?", parent=self.root):
            return

        
        # Calcular total_messages para modo tradicional según Simple/Doble/Triple
        if not self.manual_mode:
            mode = self.traditional_send_mode.get()
            base_links = len(self.links)
            if mode == "Simple":
                self.total_messages = base_links
            elif mode == "Doble":
                self.total_messages = base_links * 2
            elif mode == "Triple":
                self.total_messages = base_links * 3
            self.log(f"Modo Tradicional ({mode}): {self.total_messages} envíos totales", 'info')
        
        # Limpieza de flags
        if not self.manual_mode:
            # Modo tradicional (Excel/CSV)
            self.group_mode = False
            self.fidelizado_mode = None  # No usar modo fidelizado
            self.manual_paired_messages = []

        self._enter_task_mode()
        self.update_stats() # Actualizar UI con el total

        # Iniciar hilo
        threading.Thread(target=self.send_thread, daemon=True).start()

    def pause_sending(self):
        """Pausa o reanuda el envío."""
        with self.pause_lock:
            if self.is_paused:
                self.is_paused = False
                self.btn_pause.configure(text="⏸  PAUSAR")
                if hasattr(self, 'fidelizado_btn_pause'):
                    self.fidelizado_btn_pause.configure(text="⏸  PAUSAR")
                self.log("Reanudado", 'success')
            else:
                self.is_paused = True
                self.btn_pause.configure(text="▶  REANUDAR")
                if hasattr(self, 'fidelizado_btn_pause'):
                    self.fidelizado_btn_pause.configure(text="▶  REANUDAR")
                self.log("Pausado", 'warning')

    def stop_sending(self):
        """Solicita la detención del hilo de envío."""
        if messagebox.askyesno("Confirmar Cancelación", "¿Estás seguro de que deseas cancelar el envío actual?", parent=self.root):
            self.should_stop = True
            self.log("Cancelando...", 'warning')

    def _show_completion_dialog(self):
        """Muestra la ventana personalizada de finalización (MOD 28)."""
        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Envío Completado")
        dialog.transient(self.root); dialog.grab_set(); dialog.attributes('-topmost', True)
        dialog.resizable(False, False)

        width, height = 400, 200
        self.root.update_idletasks()
        root_x, root_y = self.root.winfo_x(), self.root.winfo_y()
        root_w, root_h = self.root.winfo_width(), self.root.winfo_height()
        x, y = root_x + (root_w // 2) - (width // 2), root_y + (root_h // 2) - (height // 2)
        dialog.geometry(f"{width}x{height}+{x}+{y}")
        dialog.after(100, dialog.focus_force)

        main_frame = ctk.CTkFrame(dialog, fg_color=self.colors['bg_card'])
        main_frame.pack(expand=True, fill=tk.BOTH, padx=20, pady=20)
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_rowconfigure(0, weight=1)
        main_frame.grid_rowconfigure(1, weight=0)

        content_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        content_frame.grid(row=0, column=0, pady=(10, 20))

        try:
            logo_p = os.path.join(BASE_DIR, 'logo_left.png')
            logo_i = Image.open(logo_p).resize((60, 60), Image.Resampling.LANCZOS)
            logo_img = ctk.CTkImage(light_image=logo_i, dark_image=logo_i, size=(60, 60))
            ctk.CTkLabel(content_frame, image=logo_img, text="").pack(pady=(0, 10))
        except Exception as e:
            print(f"Error cargando logo para diálogo: {e}")

        ctk.CTkLabel(content_frame,
                     text="Hermes entregó tus mensajes correctamente.",
                     font=self.fonts['dialog_text'],
                     text_color=self.colors['text'],
                     wraplength=300).pack()

        def close_dialog(e=None):
            dialog.grab_release()
            dialog.destroy()
            self.root.focus_force()

        ok_button = ctk.CTkButton(main_frame, text="OK", command=close_dialog,
                                  font=self.fonts['button'],
                                  fg_color=self.colors['action_start'],
                                  hover_color=self.hover_colors['action_start'],
                                  width=100)
        ok_button.grid(row=1, column=0, pady=(0, 10))

        dialog.bind('<Return>', close_dialog)
        dialog.protocol("WM_DELETE_WINDOW", close_dialog)
        self.root.wait_window(dialog)

    # --- INICIO MODIFICACIÓN: send_thread (Refactorizado) ---
    def send_thread(self):
        """Hilo de trabajo que envía los mensajes uno por uno."""
        try:
            self.log("INICIANDO ENVÍO", 'success')

            # Limpieza inicial
            for dev in self.devices:
                if self.should_stop: break
                self.close_all_apps(dev)

            if self.should_stop: self.log("Cancelado", 'warning'); return
            self.log("Pausa inicial de 3s...", 'info'); time.sleep(3)
            if self.should_stop: self.log("Cancelado", 'warning'); return

            # --- Lógica de envío (depende del modo) ---
            if self.fidelizado_mode == "GRUPOS":
                self.run_grupos_dual_whatsapp_thread()
            elif self.fidelizado_mode == "NUMEROS":
                if self.fidelizado_numeros_mode.get() == "Uno a uno":
                    self.run_uno_a_uno_thread()
                else: # Uno a muchos
                    self.run_uno_a_muchos_thread()
            elif self.fidelizado_mode == "MIXTO":
                self.run_mixto_dual_whatsapp_thread()
            else:
                self.run_default_thread()
            # --- Fin Lógica de envío ---

            # Finalización
            if not self.should_stop:
                self.log("ENVÍO FINALIZADO", 'success')
                self.log(f"Resumen: Enviados: {self.sent_count} | Fallidos: {self.failed_count}", 'info')
                self.root.after(100, self._show_completion_dialog)

        except Exception as e:
            self.log(f"Error CRÍTICO en el hilo de envío: {e}", 'error')
            import traceback
            traceback_str = traceback.format_exc()
            print(f"ERROR THREAD ENVIO:\n{traceback_str}")
            self.root.after(100, lambda: messagebox.showerror("Error Crítico", f"Ocurrió un error inesperado durante el envío:\n{e}\n\nRevise el log para más detalles.", parent=self.root))
        finally:
            # Siempre reestablecer la UI
            self.root.after(100, self._finalize_sending)
    
    def run_single_task(self, device, link, message_to_send, task_index, whatsapp_package="com.whatsapp.w4b"):
        """
        Ejecuta una única tarea de envío (abrir link, enviar, esperar), gestionando la conexión de uiautomator2.
        """
        ui_device = None
        # --- MODIFICACIÓN: Siempre intentar conectar uiautomator2 ---
        try:
            ui_device = u2.connect(device)
            ui_device.unlock()
        except Exception as e:
            self.log(f"No se pudo conectar uiautomator2 a {device}: {e}", "warning")
            # En este nuevo enfoque, un fallo aquí debería ser crítico.
            return False

        # Bucle de pausa
        while self.is_paused and not self.should_stop:
            time.sleep(0.1)
        if self.should_stop: return False

        self.current_index = task_index
        self.root.after(0, self.update_stats)

        self.close_all_apps(device)
        if self.should_stop: return False

        # Enviar mensaje, pasando el objeto ui_device
        success = self.send_msg(device, link, task_index, self.total_messages, message_to_send, whatsapp_package, ui_device)
        
        # --- Importante: Actualizar contadores DESPUÉS de send_msg ---
        if success:
            self.sent_count += 1
        else:
            self.failed_count += 1

        # Actualizar UI (contadores y barra de progreso)
        self.root.after(0, self.update_stats)
        # --- Fin actualización contadores ---

        # Espera entre mensajes (solo si no es la última tarea)
        if task_index < self.total_messages and not self.should_stop:
            if self.manual_mode:
                # FIX: Usar las nuevas variables de retardo de envío
                delay = random.uniform(self.fidelizado_send_delay_min.get(), self.fidelizado_send_delay_max.get())
            else:
                delay = random.uniform(self.delay_min.get(), self.delay_max.get())
            self.log(f"Esperando {delay:.1f}s... (Post-tarea {task_index})", 'info')
            elapsed = 0
            while elapsed < delay and not self.should_stop:
                while self.is_paused and not self.should_stop: time.sleep(0.1)
                if self.should_stop: break
                time.sleep(0.1); elapsed += 0.1
        
        return success

    def run_default_thread(self):
        """
        Lógica de envío tradicional (Excel/CSV) con soporte para Simple/Doble/Triple.
        """
        if not self.links:
            self.log("Error: No hay links para enviar (modo tradicional)", 'error')
            return
        
        mode = self.traditional_send_mode.get()
        self.log(f"Modo de envío: {mode}", 'info')
        
        if mode == "Business":
            self._run_simple_mode()
        elif mode == "Normal":
            # Reutiliza _run_simple_mode pero cambiando el paquete de WA
            self._run_simple_mode(whatsapp_package="com.whatsapp")
        elif mode == "Business/Normal":
            self._run_doble_mode()
        elif mode == "B/N.1/N.2":
            self._run_triple_mode()
    
    def _run_simple_mode(self, whatsapp_package="com.whatsapp.w4b"):
        """Modo Simple: 1 URL por teléfono, usando el paquete de WhatsApp especificado."""
        log_msg = "Ejecutando Modo Business..." if whatsapp_package == "com.whatsapp.w4b" else "Ejecutando Modo Normal..."
        self.log(log_msg, 'info')
        idx = 0  # Índice del dispositivo a usar
        
        for i, link in enumerate(self.links):
            if self.should_stop:
                self.log("Cancelado en bucle", 'warning')
                break
            
            device = self.devices[idx]
            idx = (idx + 1) % len(self.devices)
            
            # Ejecutar tarea con el paquete de WA especificado
            self.run_single_task(device, link, None, i + 1, whatsapp_package=whatsapp_package)

    def _run_doble_mode(self):
        """Modo Doble: Rota secuencialmente entre dispositivos y cuentas Business/Normal."""
        self.log("Ejecutando Modo Doble (Rotación Correcta)...", 'info')

        # 1. Crear la lista de todas las combinaciones de envío
        envio_combinations = []
        for device in self.devices:
            envio_combinations.append({"device": device, "wa_name": "Business", "wa_package": "com.whatsapp.w4b"})
            envio_combinations.append({"device": device, "wa_name": "Normal", "wa_package": "com.whatsapp"})

        num_combinations = len(envio_combinations)
        if num_combinations == 0:
            self.log("Error: No hay dispositivos para el modo Doble.", "error")
            return

        # 2. Iterar una vez sobre los links, rotando las combinaciones
        for i, link in enumerate(self.links):
            if self.should_stop:
                self.log("Cancelado en bucle", 'warning')
                break
            
            self.last_task_time = time.time()

            # Seleccionar la combinación de envío (dispositivo + cuenta)
            combination = envio_combinations[i % num_combinations]
            device = combination["device"]
            wa_name = combination["wa_name"]
            wa_package = combination["wa_package"]

            self.log(f"[{device}] Enviando con {wa_name}", 'info')
            self.run_single_task(device, link, None, i + 1, whatsapp_package=wa_package)

    def _run_triple_mode(self):
        """Modo Triple: Rota secuencialmente entre dispositivos y las 3 cuentas."""
        self.log("Ejecutando Modo Triple (Rotación Correcta)...", 'info')
        is_normal_account_2 = {dev_id: False for dev_id in self.devices}

        # 1. Crear la lista de todas las combinaciones de envío
        envio_combinations = []
        for device in self.devices:
            envio_combinations.append({"device": device, "wa_name": "Business", "wa_package": "com.whatsapp.w4b", "needs_switch": False})
            envio_combinations.append({"device": device, "wa_name": "Normal (Cuenta 1)", "wa_package": "com.whatsapp", "needs_switch": False})
            envio_combinations.append({"device": device, "wa_name": "Normal (Cuenta 2)", "wa_package": "com.whatsapp", "needs_switch": True})

        num_combinations = len(envio_combinations)
        if num_combinations == 0:
            self.log("Error: No hay dispositivos para el modo Triple.", "error")
            return

        # 2. Iterar una vez sobre los links, rotando las combinaciones
        for i, link in enumerate(self.links):
            if self.should_stop:
                self.log("Cancelado en bucle", 'warning')
                break

            self.last_task_time = time.time()

            # Seleccionar la combinación de envío (dispositivo + cuenta)
            combination = envio_combinations[i % num_combinations]
            device = combination["device"]
            wa_name = combination["wa_name"]
            wa_package = combination["wa_package"]
            needs_switch_to_acc2 = combination["needs_switch"]
            
            # Gestionar cambio de cuenta si es necesario
            if "Normal" in wa_name:
                currently_is_acc2 = is_normal_account_2.get(device, False)
                if needs_switch_to_acc2 and not currently_is_acc2:
                    self.log(f"Cambiando a Cuenta 2 en {device}...", 'info')
                    self._switch_whatsapp_account(device)
                    time.sleep(1)
                    is_normal_account_2[device] = True
                elif not needs_switch_to_acc2 and currently_is_acc2:
                    self.log(f"Restaurando a Cuenta 1 en {device}...", 'info')
                    self._switch_whatsapp_account(device)
                    time.sleep(1)
                    is_normal_account_2[device] = False

            if self.should_stop: break

            self.log(f"[{device}] Enviando con {wa_name}", 'info')
            self.run_single_task(device, link, None, i + 1, whatsapp_package=wa_package)

        # Dejar todas las cuentas Normal en el estado inicial (Cuenta 1)
        self.log("Finalizando y restaurando cuentas a estado inicial...", 'info')
        for dev, is_acc2 in is_normal_account_2.items():
            if is_acc2:
                self.log(f"Restaurando a Cuenta 1 en {dev}...", 'info')
                self._switch_whatsapp_account(dev)
                time.sleep(1)
    
    
    def _get_filtered_devices(self):
        """
        Filtra self.devices según el tipo de WhatsApp seleccionado.
        Devuelve una lista de series de dispositivos a utilizar.
        """
        whatsapp_mode = self.whatsapp_mode.get()

        if whatsapp_mode == "Ambas":
            self.log("Modo 'Ambas' seleccionado, usando todos los dispositivos.", 'info')
            return self.devices

        if not self.detected_phone_lines:
            self.log("No se han detectado líneas de WhatsApp. No se puede filtrar. Usando todos los dispositivos.", 'warning')
            return self.devices

        target_type = "WhatsApp" if whatsapp_mode == "Normal" else "WhatsApp Business"

        allowed_serials = {line['device'] for line in self.detected_phone_lines if line['type'] == target_type}

        filtered_devices = [device for device in self.devices if device in allowed_serials]

        self.log(f"Filtrando por '{whatsapp_mode}'. Dispositivos a usar: {len(filtered_devices)} de {len(self.devices)}", 'info')

        return filtered_devices

    def _get_whatsapp_apps_to_use(self):
        """
        Retorna una lista de tuplas (nombre, package) según la selección del usuario.
        Opciones: Business, Normal, Ambas (Business + Normal), Todas (Business + Normal + Normal con cambio de cuenta)
        """
        wa_mode = self.whatsapp_mode.get()
        
        if wa_mode == "Normal":
            return [("Normal", "com.whatsapp")]
        elif wa_mode == "Business":
            return [("Business", "com.whatsapp.w4b")]
        elif wa_mode == "Ambas":
            return [("Business", "com.whatsapp.w4b"), ("Normal", "com.whatsapp")]
        else:  # "Todas"
            # Business + Normal (cuenta 1) + Normal (cuenta 2 después de cambio automático)
            return [("Business", "com.whatsapp.w4b"), ("Normal", "com.whatsapp"), ("Normal", "com.whatsapp")]
    
    
    def _send_to_target_with_whatsapp(self, device, target_link, wa_name, wa_package, mensaje, task_counter):
        """
        Envía un mensaje a un target, gestionando la conexión de uiautomator2 y usando el método de envío unificado.
        """
        # Conectar con uiautomator2 al inicio de la tarea
        ui_device = None
        try:
            ui_device = u2.connect(device)
            ui_device.unlock()
        except Exception as e:
            self.log(f"No se pudo conectar uiautomator2 a {device}: {e}", "warning")
            ui_device = None

        # Llamar a send_msg, que ahora contiene toda la lógica de envío y fallback
        success = self.send_msg(device, target_link, task_counter, self.total_messages, message_to_send=mensaje, whatsapp_package=wa_package, ui_device=ui_device)
        
        if success:
            self.sent_count += 1
        else:
            self.failed_count += 1
        self.root.after(0, self.update_stats)
        
        # Si es Normal y el modo es "Todas", ejecutar el cambio de cuenta DESPUÉS de enviar
        if wa_name == "Normal" and self.whatsapp_mode.get() == "Todas":
            self.log(f"[{device}] Cerrando WhatsApp Normal después de enviar...", 'info')
            close_cmd = ['-s', device, 'shell', 'am', 'force-stop', 'com.whatsapp']
            self._run_adb_command(close_cmd, timeout=5)
            time.sleep(1)
            
            self.log(f"[{device}] Reabriendo WhatsApp Normal...", 'info')
            open_cmd = ['-s', device, 'shell', 'am', 'start', '-n', 'com.whatsapp/.Main']
            self._run_adb_command(open_cmd, timeout=5)
            time.sleep(3)  # Esperar 3 segundos para que WhatsApp se abra completamente
            
            self.log(f"[{device}] Cambiando de cuenta...", 'info')
            self._switch_account_for_device(device)
            time.sleep(1)
            
            self.log(f"[{device}] Cerrando WhatsApp Normal después de cambiar cuenta...", 'info')
            close_cmd = ['-s', device, 'shell', 'am', 'force-stop', 'com.whatsapp']
            self._run_adb_command(close_cmd, timeout=5)
            time.sleep(1)
            
            self.log(f"[{device}] Reabriendo WhatsApp Normal con nueva cuenta...", 'info')
            open_cmd = ['-s', device, 'shell', 'am', 'start', '-n', 'com.whatsapp/.Main']
            self._run_adb_command(open_cmd, timeout=5)
            time.sleep(2)
        
        return True
    
    def run_mixto_dual_whatsapp_thread(self):
        """
        Lógica de envío para MODO MIXTO con 3 variantes.
        Variante 1: G1→N1→G2→N2 (alternar 1 a 1)
        Variante 2: G1→G2→N1→G3→G4→N2 (2 grupos por número)
        Variante 3: G1→G2→G3→N1→G4→G5→G6→N2 (3 grupos por número)
        
        Los grupos y números se repiten en bucle si hay menos de los necesarios.
        Todas las líneas (dispositivos) siguen la misma secuencia.
        """
        # --- FILTRADO DE DISPOSITIVOS ---
        active_devices = self._get_filtered_devices()
        if not active_devices:
            self.log("No hay dispositivos que cumplan con el filtro seleccionado.", "error")
            self.root.after(0, lambda: messagebox.showerror("Error", "No se encontraron dispositivos que cumplan con el filtro de WhatsApp seleccionado.", parent=self.root))
            return
        # --- FIN FILTRADO ---

        num_devices = len(active_devices)
        num_grupos = len(self.manual_inputs_groups)
        num_numeros = len(self.manual_inputs_numbers)
        num_bucles = self.manual_loops
        variant = self.mixto_variant.get()
        
        if len(self.manual_messages_numbers) < 1:
            self.log("Error: Modo Mixto requiere al menos 1 mensaje cargado.", "error")
            messagebox.showerror("Error", "Debes cargar al menos 1 archivo de mensajes para el modo mixto.", parent=self.root)
            return
        
        # Usar índice de inicio aleatorio
        mensaje_index = self.mensaje_start_index
        total_mensajes = len(self.manual_messages_numbers)
        task_counter = 0
        whatsapp_apps = self._get_whatsapp_apps_to_use()
        
        variant_names = {1: "1:1", 2: "2:1", 3: "3:1"}
        self.log(f"Modo Mixto (Variante {variant} - {variant_names[variant]}): {num_bucles} ciclo(s), {num_grupos} grupo(s), {num_numeros} número(s), {num_devices} dispositivo(s)", 'info')
        self.log(f"WhatsApp: {self.whatsapp_mode.get()}", 'info')
        self.log(f"Total de envíos: {self.total_messages}", 'info')
        
        for ciclo in range(num_bucles):
            if self.should_stop: break
            self.log(f"\n--- CICLO {ciclo + 1}/{num_bucles} ---", 'info')
            
            # Crear lista de targets según la variante
            targets = []
            
            if variant == 1:
                # Variante 1: G1→N1→G2→N2 (alternar 1 a 1)
                max_len = max(num_grupos, num_numeros)
                for i in range(max_len):
                    grupo_idx = i % num_grupos
                    numero_idx = i % num_numeros
                    targets.append(('grupo', grupo_idx, self.manual_inputs_groups[grupo_idx]))
                    targets.append(('numero', numero_idx, self.manual_inputs_numbers[numero_idx]))
                    
            elif variant == 2:
                # Variante 2: G1→G2→N1→G3→G4→N2 (2 grupos por número)
                grupo_idx = 0
                for num_idx in range(num_numeros):
                    # Añadir 2 grupos
                    for _ in range(2):
                        targets.append(('grupo', grupo_idx % num_grupos, self.manual_inputs_groups[grupo_idx % num_grupos]))
                        grupo_idx += 1
                    # Añadir 1 número
                    targets.append(('numero', num_idx, self.manual_inputs_numbers[num_idx]))
                # Si sobran grupos, continuar añadiendo en bucle
                if grupo_idx < num_grupos:
                    remaining = num_grupos - grupo_idx
                    for _ in range(remaining):
                        targets.append(('grupo', grupo_idx % num_grupos, self.manual_inputs_groups[grupo_idx % num_grupos]))
                        grupo_idx += 1
                        
            elif variant == 3:
                # Variante 3: G1→G2→G3→N1→G4→G5→G6→N2 (3 grupos por número)
                grupo_idx = 0
                for num_idx in range(num_numeros):
                    # Añadir 3 grupos
                    for _ in range(3):
                        targets.append(('grupo', grupo_idx % num_grupos, self.manual_inputs_groups[grupo_idx % num_grupos]))
                        grupo_idx += 1
                    # Añadir 1 número
                    targets.append(('numero', num_idx, self.manual_inputs_numbers[num_idx]))
                # Si sobran grupos, continuar añadiendo en bucle
                if grupo_idx < num_grupos:
                    remaining = num_grupos - grupo_idx
                    for _ in range(remaining):
                        targets.append(('grupo', grupo_idx % num_grupos, self.manual_inputs_groups[grupo_idx % num_grupos]))
                        grupo_idx += 1
            
            # Procesar cada target en la secuencia
            for target_type, target_idx, target_value in targets:
                if self.should_stop: break
                
                tipo_str = "Grupo" if target_type == 'grupo' else "Número"
                if target_type == 'grupo':
                    self.log(f"\n=== GRUPO {target_idx + 1}/{num_grupos}: {target_value[:50]}... ===", 'info')
                else:
                    self.log(f"\n=== NÚMERO {target_idx + 1}/{num_numeros}: +549{target_value} ===", 'info')
                
                # Por cada dispositivo (todas las líneas procesan la misma secuencia)
                for device in self.devices:
                    if self.should_stop: break
                    
                    # Por cada WhatsApp (Normal, Business, o Ambos)
                    for wa_idx, (wa_name, wa_package) in enumerate(whatsapp_apps):
                        if self.should_stop: break
                        
                        task_counter += 1
                        self.current_index = task_counter
                        self.root.after(0, self.update_stats)
                        
                        # Obtener mensaje rotativo
                        mensaje = self.manual_messages_numbers[mensaje_index % total_mensajes]
                        mensaje_index += 1
                        
                        # Construir link según tipo
                        if target_type == 'grupo':
                            target_link = target_value
                        else:
                            target_link = f"https://wa.me/549{target_value}"
                        
                        # Enviar usando la función auxiliar
                        success = self._send_to_target_with_whatsapp(
                            device, target_link, wa_name, wa_package, mensaje, task_counter
                        )
                        
                        # Pausa entre WhatsApps si hay más de uno y es Business (primero)
                        if success and len(whatsapp_apps) > 1 and wa_idx == 0:
                            wait_between = self.wait_between_messages.get()
                            if wait_between > 0:
                                self.log(f"Esperando {wait_between}s antes del siguiente WhatsApp...", 'info')
                                elapsed = 0
                                while elapsed < wait_between and not self.should_stop:
                                    while self.is_paused and not self.should_stop: time.sleep(0.1)
                                    if self.should_stop: break
                                    time.sleep(0.1)
                                    elapsed += 0.1
                        
                        time.sleep(0.5)  # Pequeña pausa entre envíos
                
                if self.should_stop: break
                self.log(f"\n=== {tipo_str} {target_idx + 1} completado ===", 'success')
            
            if self.should_stop: break
            self.log(f"\n--- CICLO {ciclo + 1} completado ---", 'success')
        
        self.log(f"\nModo Mixto (Variante {variant}) finalizado", 'success')

    def run_uno_a_uno_thread(self):
        """
        Lógica de envío para MODO NÚMEROS - UNO A UNO.
        Las líneas detectadas se emparejan aleatoriamente y conversan entre sí.
        Se repite según la configuración de Bucles y Ciclos.
        """
        self.log("Iniciando MODO NÚMEROS (Uno a uno)...", 'info')
        if len(self.detected_phone_lines) < 2:
            self.log("Se necesitan al menos 2 líneas de WhatsApp para este modo.", 'error')
            messagebox.showerror("Error", "Se necesitan al menos 2 líneas de WhatsApp detectadas para el modo 'Uno a uno'.", parent=self.root)
            return

        lines = self.detected_phone_lines.copy()
        num_lines = len(lines)
        num_bucles = self.manual_loops_var.get()
        num_ciclos = self.manual_cycles_var.get() # NUEVO
        
        task_counter = 0
        mensaje_index = self.mensaje_start_index
        total_mensajes_lib = len(self.manual_messages_numbers)

        # Bucle externo para los Bucles
        for bucle_num in range(num_bucles):
            if self.should_stop: break
            self.log(f"\n--- INICIANDO BUCLE {bucle_num + 1}/{num_bucles} ---", 'success')

            # Bucle interno para los Ciclos
            for ciclo_num in range(num_ciclos):
                if self.should_stop: break
                self.log(f"\n--- Ciclo {ciclo_num + 1}/{num_ciclos} (dentro del Bucle {bucle_num + 1}) ---", 'info')

                random.shuffle(lines)

                # Manejar número impar de líneas
                excluded = None
                current_lines = lines.copy()
                if num_lines % 2 != 0:
                    excluded = current_lines.pop()
                    self.log(f"Número impar de líneas. {excluded['number']} quedará fuera de este ciclo.", 'warning')

                # Crear parejas
                pairs = []
                for i in range(0, len(current_lines), 2):
                    pairs.append((current_lines[i], current_lines[i+1]))

                self.log(f"Se formaron {len(pairs)} parejas para este ciclo.", 'info')

                for i, (line_a, line_b) in enumerate(pairs):
                    if self.should_stop: break
                    self.log(f"\n=== Pareja {i+1}/{len(pairs)}: {line_a['number']} <-> {line_b['number']} ===", 'info')

                    # Conversación de ida y vuelta
                    # A -> B
                    task_counter += 1
                    mensaje_ida = self.manual_messages_numbers[mensaje_index % total_mensajes_lib]; mensaje_index += 1
                    link_ida = f"https://wa.me/{line_b['number'].replace('+', '')}?text={urllib.parse.quote(mensaje_ida, safe='')}"
                    self.run_single_task(line_a['device'], link_ida, None, task_counter, whatsapp_package=line_a['package'])
                    if self.should_stop: break

                    # B -> A
                    task_counter += 1
                    mensaje_vuelta = self.manual_messages_numbers[mensaje_index % total_mensajes_lib]; mensaje_index += 1
                    link_vuelta = f"https://wa.me/{line_a['number'].replace('+', '')}?text={urllib.parse.quote(mensaje_vuelta, safe='')}"
                    self.run_single_task(line_b['device'], link_vuelta, None, task_counter, whatsapp_package=line_b['package'])

            if self.should_stop: break
            self.log(f"\n--- FIN BUCLE {bucle_num + 1}/{num_bucles} ---", 'success')

    def run_uno_a_muchos_thread(self):
        """
        Lógica de envío para MODO NÚMEROS - UNO A MUCHOS.
        Cada línea detectada conversa con todas las demás, una sola vez por par.
        """
        self.log("Iniciando MODO NÚMEROS (Uno a muchos)...", 'info')
        if len(self.detected_phone_lines) < 2:
            self.log("Se necesitan al menos 2 líneas de WhatsApp para este modo.", 'error')
            messagebox.showerror("Error", "Se necesitan al menos 2 líneas de WhatsApp detectadas para el modo 'Uno a muchos'.", parent=self.root)
            return

        lines = self.detected_phone_lines.copy()
        num_lines = len(lines)
        num_bucles = self.manual_loops_var.get()

        task_counter = 0
        mensaje_index = self.mensaje_start_index
        total_mensajes_lib = len(self.manual_messages_numbers)

        for bucle_num in range(num_bucles):
            if self.should_stop: break
            self.log(f"\n--- BUCLE {bucle_num + 1}/{num_bucles} ---", 'info')

            # Crear todas las parejas únicas
            pairs = []
            for i in range(num_lines):
                for j in range(i + 1, num_lines):
                    pairs.append((lines[i], lines[j]))

            self.log(f"Se formaron {len(pairs)} parejas únicas para este bucle.", 'info')

            for i, (line_a, line_b) in enumerate(pairs):
                if self.should_stop: break
                self.log(f"\n=== Conversación {i+1}/{len(pairs)}: {line_a['number']} <-> {line_b['number']} ===", 'info')

                # Conversación de ida y vuelta
                # A -> B
                task_counter += 1
                mensaje_ida = self.manual_messages_numbers[mensaje_index % total_mensajes_lib]; mensaje_index += 1
                link_ida = f"https://wa.me/{line_b['number'].replace('+', '')}?text={urllib.parse.quote(mensaje_ida, safe='')}"
                self.run_single_task(line_a['device'], link_ida, None, task_counter, whatsapp_package=line_a['package'])
                if self.should_stop: break

                # B -> A
                task_counter += 1
                mensaje_vuelta = self.manual_messages_numbers[mensaje_index % total_mensajes_lib]; mensaje_index += 1
                link_vuelta = f"https://wa.me/{line_a['number'].replace('+', '')}?text={urllib.parse.quote(mensaje_vuelta, safe='')}"
                self.run_single_task(line_b['device'], link_vuelta, None, task_counter, whatsapp_package=line_b['package'])
    
    def run_grupos_dual_whatsapp_thread(self):
        """
        Lógica de envío para MODO GRUPOS.
        Por cada grupo, envía con los WhatsApps seleccionados (Normal, Business o Ambos).
        Los mensajes rotan: 1,2,3,4... y cuando se acaban vuelven al 1.
        """
        # --- FILTRADO DE DISPOSITIVOS ---
        active_devices = self._get_filtered_devices()
        if not active_devices:
            self.log("No hay dispositivos que cumplan con el filtro seleccionado.", "error")
            self.root.after(0, lambda: messagebox.showerror("Error", "No se encontraron dispositivos que cumplan con el filtro de WhatsApp seleccionado.", parent=self.root))
            return
        # --- FIN FILTRADO ---

        num_devices = len(active_devices)
        num_grupos = len(self.manual_inputs_groups)
        num_bucles = self.manual_loops_var.get() # <--- CORRECCIÓN: Leer el valor correcto de la UI

        if len(self.manual_messages_groups) < 1:
            self.log("Error: Modo Grupos requiere al menos 1 mensaje cargado.", "error")
            messagebox.showerror("Error", "Debes cargar al menos 1 archivo de mensajes.", parent=self.root)
            return

        # Usar índice de inicio aleatorio
        mensaje_index = self.mensaje_start_index
        total_mensajes_lib = len(self.manual_messages_groups)
        task_counter = 0
        whatsapp_apps = self._get_whatsapp_apps_to_use()

        self.log(f"Modo Grupos: {num_bucles} bucle(s), {num_grupos} grupo(s), {num_devices} dispositivo(s)", 'info')
        self.log(f"WhatsApp: {self.whatsapp_mode.get()}", 'info')
        self.log(f"Total de envíos: {self.total_messages}", 'info')

        # --- Bucle principal corregido ---
        for bucle_num in range(num_bucles):
            if self.should_stop: break
            self.log(f"\n--- INICIANDO BUCLE {bucle_num + 1}/{num_bucles} ---", 'success')

            # Por cada grupo
            for idx_grupo, grupo_link in enumerate(self.manual_inputs_groups):
                if self.should_stop: break
                grupo_display = grupo_link[:50] + "..." if len(grupo_link) > 50 else grupo_link
                self.log(f"\n=== GRUPO {idx_grupo + 1}/{num_grupos}: {grupo_display} ===", 'info')

                # Por cada dispositivo ACTIVO
                for device in active_devices:
                    if self.should_stop: break

                    # Por cada WhatsApp (Normal, Business, etc.)
                    for wa_idx, (wa_name, wa_package) in enumerate(whatsapp_apps):
                        if self.should_stop: break

                        task_counter += 1

                        # Obtener mensaje rotativo
                        mensaje = self.manual_messages_groups[mensaje_index % total_mensajes_lib]
                        mensaje_index += 1

                        # Usar run_single_task que ya tiene toda la lógica de envío + retardo post-tarea
                        success = self.run_single_task(
                            device, grupo_link, mensaje, task_counter, whatsapp_package=wa_package
                        )

                        # Lógica de cambio de cuenta para el modo "TODAS"
                        if wa_name == "Normal" and self.whatsapp_mode.get() == "Todas":
                            self.log(f"[{device}] Cerrando WhatsApp Normal después de enviar...", 'info')
                            self._run_adb_command(['-s', device, 'shell', 'am', 'force-stop', 'com.whatsapp'], timeout=5)
                            time.sleep(1)

                            self.log(f"[{device}] Reabriendo WhatsApp Normal...", 'info')
                            self._run_adb_command(['-s', device, 'shell', 'am', 'start', '-n', 'com.whatsapp/.Main'], timeout=5)
                            time.sleep(3)

                            self.log(f"[{device}] Cambiando de cuenta...", 'info')
                            self._switch_account_for_device(device)
                            time.sleep(1)

                            self.log(f"[{device}] Cerrando WhatsApp Normal después de cambiar cuenta...", 'info')
                            self._run_adb_command(['-s', device, 'shell', 'am', 'force-stop', 'com.whatsapp'], timeout=5)
                            time.sleep(1)

                            self.log(f"[{device}] Reabriendo WhatsApp Normal con nueva cuenta...", 'info')
                            self._run_adb_command(['-s', device, 'shell', 'am', 'start', '-n', 'com.whatsapp/.Main'], timeout=5)
                            time.sleep(2)


                        # Pausa entre diferentes tipos de WhatsApp (ej. Business -> Normal)
                        if success and len(whatsapp_apps) > 1 and wa_idx < len(whatsapp_apps) - 1:
                            wait_between = self.wait_between_messages.get()
                            if wait_between > 0:
                                self.log(f"Esperando {wait_between}s antes del siguiente WhatsApp...", 'info')
                                elapsed = 0
                                while elapsed < wait_between and not self.should_stop:
                                    while self.is_paused and not self.should_stop: time.sleep(0.1)
                                    if self.should_stop: break
                                    time.sleep(0.1)
                                    elapsed += 0.1

                if self.should_stop: break
                self.log(f"\n=== GRUPO {idx_grupo + 1} completado ===", 'success')

            if self.should_stop: break
            self.log(f"\n--- FIN BUCLE {bucle_num + 1}/{num_bucles} ---", 'success')

        self.log(f"\nModo Grupos finalizado", 'success')

    def run_unirse_grupos(self, grupos):
        """
        Función para unirse automáticamente a grupos.
        NUEVA LÓGICA CON THREADING (EJECUCIÓN PARALELA):
        Por cada grupo:
          - TODOS los dispositivos se unen SIMULTÁNEAMENTE según la selección de WhatsApp
        Proceso:
          - Presiona DPAD_DOWN 3 veces (con pausas de 2s)
          - Presiona ENTER dos veces (doble Enter)
          - Presiona BACK para salir
        """
        try:
            self._enter_task_mode()
            num_devices = len(self.devices)
            num_grupos = len(grupos)

            # Obtener qué WhatsApp usar
            wa_mode = self.whatsapp_mode.get()

            # Determinar cuántas uniones totales habrá
            if wa_mode == "Todas":
                total_uniones = num_grupos * num_devices * 3
            elif wa_mode == "Ambas":
                total_uniones = num_grupos * num_devices * 2
            else:
                total_uniones = num_grupos * num_devices

            self.log(f"\n=== UNIRSE A GRUPOS (MODO PARALELO) ===", 'info')
            self.log(f"Grupos: {num_grupos}", 'info')
            self.log(f"Dispositivos: {num_devices}", 'info')
            self.log(f"WhatsApp: {wa_mode}", 'info')
            self.log(f"Total de uniones: {total_uniones}", 'info')

            total = num_grupos * num_devices * 2

            # Función auxiliar para unirse a un grupo en un dispositivo
            def unirse_a_grupo_device(device, grupo_link, whatsapp_package, whatsapp_name):
                """Ejecuta el proceso completo de unión para un dispositivo."""
                try:
                    if self.should_stop:
                        return False

                    # Verificar pausa
                    while self.is_paused and not self.should_stop:
                        time.sleep(0.1)
                    if self.should_stop:
                        return False

                    self.log(f"[{device}] Uniéndose por {whatsapp_name}...", 'info')

                    # Abrir grupo
                    open_args = ['-s', device, 'shell', 'am', 'start', '-a', 'android.intent.action.VIEW',
                                '-d', grupo_link, '-p', whatsapp_package]

                    if not self._run_adb_command(open_args, timeout=20):
                        self.log(f"[{device}] Fallo al abrir grupo en {whatsapp_name}", "error")
                        return False

                    # Esperar 1 segundo (reducido de 2 para acelerar)
                    time.sleep(1)

                    if self.should_stop:
                        return False

                    # Presionar DPAD_DOWN 3 veces
                    for i in range(3):
                        if self.should_stop:
                            return False
                        down_args = ['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_DPAD_DOWN']
                        self._run_adb_command(down_args, timeout=5)
                        time.sleep(1) # Reducido de 2s

                    if self.should_stop:
                        return False

                    # Presionar ENTER (primer Enter)
                    enter_args = ['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_ENTER']
                    self._run_adb_command(enter_args, timeout=10)

                    # Esperar 0.5 segundos entre Enters (reducido de 1s)
                    time.sleep(0.5)

                    # Presionar ENTER (segundo Enter)
                    self._run_adb_command(enter_args, timeout=10)

                    # Esperar 1 segundo (reducido de 2s)
                    time.sleep(1)

                    # Presionar BACK para salir del grupo
                    back_args = ['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_BACK']
                    self._run_adb_command(back_args, timeout=10)
                    self.log(f"[{device}] Presionando BACK para salir...", 'info')

                    # Esperar 0.5 segundos final (reducido de 1s)
                    time.sleep(0.5)

                    self.log(f"[{device}] Unido a grupo por {whatsapp_name}", 'success')
                    return True

                except Exception as e:
                    self.log(f"[{device}] Error en unión: {e}", 'error')
                    return False
            
            # Por cada grupo
            for idx_grupo, grupo_link in enumerate(grupos):
                if self.should_stop:
                    break
                
                grupo_display = grupo_link[:50] + "..." if len(grupo_link) > 50 else grupo_link
                self.log(f"\n--- GRUPO {idx_grupo + 1}/{num_grupos}: {grupo_display} ---", 'info')
                
                # ===== FASE 1: WHATSAPP BUSINESS (si corresponde) =====
                if wa_mode == "Business" or wa_mode == "Ambas" or wa_mode == "Todas":
                    fase_num = 1 if (wa_mode == "Ambas" or wa_mode == "Todas") else 0
                    if fase_num == 1:
                        self.log(f"\n>>> FASE 1: Todos los dispositivos uniéndose SIMULTÁNEAMENTE por WhatsApp Business...", 'info')
                    else:
                        self.log(f"\n>>> Todos los dispositivos uniéndose SIMULTÁNEAMENTE por WhatsApp Business...", 'info')
                    
                    threads_business = []
                    for device in self.devices:
                        if self.should_stop:
                            break
                        thread = threading.Thread(
                            target=unirse_a_grupo_device,
                            args=(device, grupo_link, 'com.whatsapp.w4b', 'WhatsApp Business'),
                            daemon=True
                        )
                        threads_business.append(thread)
                        thread.start()

                    # Esperar a que TODOS los threads de Business terminen
                    for thread in threads_business:
                        thread.join()

                    if fase_num == 1:
                        self.log(f"\n>>> FASE 1 completada: Todos unidos por WhatsApp Business", 'success')
                    else:
                        self.log(f"\n>>> Completado: Todos unidos por WhatsApp Business", 'success')
                    
                    if self.should_stop:
                        break
                
                # ===== FASE 2: WHATSAPP NORMAL (si corresponde) =====
                if wa_mode == "Normal" or wa_mode == "Ambas" or wa_mode == "Todas":
                    fase_num = 2 if (wa_mode == "Ambas" or wa_mode == "Todas") else 0
                    if fase_num == 2:
                        self.log(f"\n>>> FASE 2: Todos los dispositivos uniéndose SIMULTÁNEAMENTE por WhatsApp Normal...", 'info')
                    else:
                        self.log(f"\n>>> Todos los dispositivos uniéndose SIMULTÁNEAMENTE por WhatsApp Normal...", 'info')

                    threads_normal = []
                    for device in self.devices:
                        if self.should_stop:
                            break
                        thread = threading.Thread(
                            target=unirse_a_grupo_device,
                            args=(device, grupo_link, 'com.whatsapp', 'WhatsApp Normal'),
                            daemon=True
                        )
                        threads_normal.append(thread)
                        thread.start()

                    # Esperar a que TODOS los threads de Normal terminen
                    for thread in threads_normal:
                        thread.join()

                    if fase_num == 2:
                        self.log(f"\n>>> FASE 2 completada: Todos unidos por WhatsApp Normal", 'success')
                    else:
                        self.log(f"\n>>> Completado: Todos unidos por WhatsApp Normal", 'success')

                    # Si el modo es "Todas", cambiar de cuenta DESPUÉS de unirse con Normal
                    if wa_mode == "Todas":
                        self.log(f"\n>>> Cambiando de cuenta en todos los dispositivos...", 'info')

                        for device in self.devices:
                            if self.should_stop:
                                break

                            self.log(f"[{device}] Cerrando WhatsApp Normal...", 'info')
                            close_cmd = ['-s', device, 'shell', 'am', 'force-stop', 'com.whatsapp']
                            self._run_adb_command(close_cmd, timeout=5)
                            time.sleep(1)

                            self.log(f"[{device}] Reabriendo WhatsApp Normal...", 'info')
                            open_cmd = ['-s', device, 'shell', 'am', 'start', '-n', 'com.whatsapp/.Main']
                            self._run_adb_command(open_cmd, timeout=5)
                            time.sleep(3)  # Esperar 3 segundos para que WhatsApp se abra completamente

                            self.log(f"[{device}] Cambiando de cuenta...", 'info')
                            self._switch_account_for_device(device)
                            time.sleep(1)

                            self.log(f"[{device}] Cerrando WhatsApp Normal después de cambiar cuenta...", 'info')
                            close_cmd = ['-s', device, 'shell', 'am', 'force-stop', 'com.whatsapp']
                            self._run_adb_command(close_cmd, timeout=5)
                            time.sleep(1)

                            self.log(f"[{device}] Reabriendo WhatsApp Normal con nueva cuenta...", 'info')
                            open_cmd = ['-s', device, 'shell', 'am', 'start', '-n', 'com.whatsapp/.Main']
                            self._run_adb_command(open_cmd, timeout=5)
                            time.sleep(2)

                        if self.should_stop:
                            break
                
                # ===== FASE 3: WHATSAPP NORMAL 2 (si corresponde) =====
                if wa_mode == "Todas":
                    self.log(f"\n>>> FASE 3: Todos los dispositivos uniéndose SIMULTÁNEAMENTE por WhatsApp Normal (cuenta 2)...", 'info')

                    threads_normal2 = []
                    for device in self.devices:
                        if self.should_stop:
                            break
                        thread = threading.Thread(
                            target=unirse_a_grupo_device,
                            args=(device, grupo_link, 'com.whatsapp', 'WhatsApp Normal (cuenta 2)'),
                            daemon=True
                        )
                        threads_normal2.append(thread)
                        thread.start()

                    # Esperar a que TODOS los threads de Normal 2 terminen
                    for thread in threads_normal2:
                        thread.join()

                    self.log(f"\n>>> FASE 3 completada: Todos unidos por WhatsApp Normal (cuenta 2)", 'success')
                
                self.log(f"\n=== GRUPO {idx_grupo + 1} completado ===", 'success')
            
            self.log(f"\n=== PROCESO DE UNIÓN A GRUPOS FINALIZADO ===", 'success')
            messagebox.showinfo("Éxito", f"Proceso completado.\n\nSe unieron a {num_grupos} grupo(s) con {num_devices} dispositivo(s).", parent=self.root)
        finally:
            self._finalize_sending()
    
    def _write_message_with_keyevents(self, device, message):
        """
        Escribe un mensaje carácter por carácter usando input text de ADB.
        La velocidad depende de la configuración del usuario.
        Retorna True si tuvo éxito, False si falló.
        """
        try:
            if self.should_stop:
                return False
            
            while self.is_paused and not self.should_stop:
                time.sleep(0.1)
            if self.should_stop:
                return False
            
            # Obtener delay según velocidad seleccionada
            speed = self.write_speed.get()
            if speed == "Lento":
                char_delay = 0.15  # 150ms por carácter
            elif speed == "Normal":
                char_delay = 0.08  # 80ms por carácter
            else:  # Rápido
                char_delay = 0.03  # 30ms por carácter
            
            # Escribir carácter por carácter
            for char in message:
                if self.should_stop:
                    return False
                
                while self.is_paused and not self.should_stop:
                    time.sleep(0.1)
                if self.should_stop:
                    return False
                
                # Escapar el carácter individual
                if char == ' ':
                    char_escaped = '%s'
                elif char in ['\\', '"', "'", '$', '`', '!', '&', '|', ';', '<', '>', 
                             '(', ')', '[', ']', '{', '}', '*', '?', '#', '~']:
                    char_escaped = f'\\{char}'
                else:
                    char_escaped = char
                
                # Enviar el carácter
                text_args = ['-s', device, 'shell', 'input', 'text', char_escaped]
                
                if not self._run_adb_command(text_args, timeout=5):
                    # Si falla un carácter, intentar continuar
                    self.log(f"Advertencia: fallo al escribir '{char}'", "warning")
                
                # Delay entre caracteres según velocidad
                time.sleep(char_delay)
            
            # Pausa final después de escribir todo
            time.sleep(0.2)
            return True
            
        except Exception as e:
            self.log(f"Error al escribir mensaje: {e}", 'error')
            return False
    
    def run_bucle_blast_thread_V2(self):
        """
        Lógica de envío NUEVA (Modo Bucles G/N Blast V2).
        Definición de 1 Bucle: Recorrer TODA la lista de G y N.
        Repetir esto N veces.
        """
        num_devices = len(self.devices)
        num_bucles = self.manual_loops
        
        # Crear copias de las listas de tareas
        group_targets = list(self.manual_inputs_groups)
        number_targets = list(self.manual_inputs_numbers)
        group_messages = list(self.manual_messages_groups)
        number_messages = list(self.manual_messages_numbers)

        # Índices para rotación de mensajes
        g_msg_idx = 0
        n_msg_idx = 0
        
        # Contador global de tareas (1-based)
        task_counter = 0

        if not group_targets or not number_targets or not group_messages or not number_messages:
            self.log("Error: Modo Bucles Blast requiere Grupos, Números y sus respectivos Mensajes.", "error")
            return
            
        # Longitud del "sub-bucle" (la lista de targets más larga)
        max_len = max(len(group_targets), len(number_targets))

        # --- Funciones helper para rotar mensajes ---
        def get_next_g_msg():
            nonlocal g_msg_idx
            msg = group_messages[g_msg_idx % len(group_messages)]
            g_msg_idx += 1
            return msg

        def get_next_n_msg():
            nonlocal n_msg_idx
            msg = number_messages[n_msg_idx % len(number_messages)]
            n_msg_idx += 1
            return msg
        # --- Fin helpers ---

        # --- Bucle Principal (N Repeticiones) ---
        for b in range(num_bucles):
            if self.should_stop: break
            rep_num = b + 1
            self.log(f"--- Iniciando REPETICIÓN {rep_num} / {num_bucles} ---", 'info')

            # --- Bucle Interno (Recorrer todos los targets) ---
            for i in range(max_len):
                if self.should_stop: break
                
                # --- Etapa 1: Blast de Grupo ---
                # Target rotativo
                current_group_target = group_targets[i % len(group_targets)]
                self.log(f"Repetición {rep_num}, Etapa {i+1}.1: Todos a GRUPO {current_group_target[:40]}...", 'info')
                
                for device in self.devices:
                    if self.should_stop: break
                    task_counter += 1
                    msg_g = get_next_g_msg() # Mensaje rotativo
                    self.run_single_task(device, current_group_target, msg_g, task_counter)
                
                if self.should_stop: break # Check entre etapas

                # --- Etapa 2: Blast de Número ---
                # Target rotativo
                current_number_target = number_targets[i % len(number_targets)]
                self.log(f"Repetición {rep_num}, Etapa {i+1}.2: Todos a NÚMERO {current_number_target}", 'info')
                
                for device in self.devices:
                    if self.should_stop: break
                    task_counter += 1
                    msg_n = get_next_n_msg() # Mensaje rotativo
                    wa_link = f"https://wa.me/549{current_number_target}?text={urllib.parse.quote(msg_n, safe='')}"
                    self.run_single_task(device, wa_link, None, task_counter) # None pork el msg va en el link
            
            if self.should_stop: break # Check al final del bucle interno

            self.log(f"--- Fin REPETICIÓN {rep_num} ---", 'info')
    

    def _finalize_sending(self):
        """Reestablece la UI al finalizar o cancelar el envío."""
        self.is_running = False

        # -- Vista Tradicional --
        self.btn_start.configure(state=tk.NORMAL)
        self.btn_load.configure(state=tk.NORMAL)
        if self.fidelizado_unlock_btn:
            self.fidelizado_unlock_btn.configure(state=tk.NORMAL)
        self.btn_pause.configure(state=tk.DISABLED, text="⏸  PAUSAR")
        self.btn_stop.configure(state=tk.DISABLED)

        # -- Vista Fidelizado --
        if hasattr(self, 'fidelizado_btn_start'):
            # Ocultar botones de control, mostrar los de inicio
            self.control_buttons_frame.grid_remove()
            self.actions_frame.grid()

            # Configurar estado de botones
            self.fidelizado_btn_start.configure(state=tk.NORMAL)
            self.unirse_grupos_btn.configure(state=tk.NORMAL)
            self.fidelizado_btn_pause.configure(state=tk.DISABLED, text="⏸  PAUSAR")
            self.fidelizado_btn_stop.configure(state=tk.DISABLED)

    def _enter_task_mode(self):
        """Configura la UI para un estado de 'tarea en ejecución'."""
        self.is_running = True
        self.is_paused = False
        self.should_stop = False
        self.sent_count = 0
        self.failed_count = 0
        self.current_index = 0
        self.start_time = datetime.now()

        # Actualizar UI
        # -- Vista Tradicional --
        self.btn_start.configure(state=tk.DISABLED)
        self.btn_load.configure(state=tk.DISABLED)
        if self.fidelizado_unlock_btn:
            self.fidelizado_unlock_btn.configure(state=tk.DISABLED)
        self.btn_pause.configure(state=tk.NORMAL)
        self.btn_stop.configure(state=tk.NORMAL)

        # -- Vista Fidelizado --
        if hasattr(self, 'fidelizado_btn_start'):
            # Ocultar botones de inicio, mostrar los de control
            self.actions_frame.grid_remove()
            self.control_buttons_frame.grid()

            # Configurar estado de botones
            self.fidelizado_btn_start.configure(state=tk.DISABLED)
            self.unirse_grupos_btn.configure(state=tk.DISABLED)
            self.fidelizado_btn_pause.configure(state=tk.NORMAL)
            self.fidelizado_btn_stop.configure(state=tk.NORMAL)

    def _update_device_labels(self):
        """Actualiza todas las etiquetas de la UI que muestran la lista de dispositivos."""
        if hasattr(self, 'fidelizado_device_list_label'):
            if self.devices:
                device_text = "Dispositivos Encontrados:\n" + "\n".join(self.devices)
                self.fidelizado_device_list_label.configure(text=device_text)
            else:
                self.fidelizado_device_list_label.configure(text="No hay dispositivos detectados.")

    # --- ################################################################## ---
    # --- send_msg (MODIFICADO para loguear device)
    # --- ################################################################## ---
    
    def _switch_whatsapp_account(self, device):
        """
        Cambia de cuenta en WhatsApp Normal.
        
        Args:
            device: ID del dispositivo
        """
        self.log(f"[{device}] Cambiando de cuenta en WhatsApp Normal...", 'info')
        
        # 1) Cerrar todo
        close_commands = [
            ['-s', device, 'shell', 'am', 'force-stop', 'com.android.settings'],
            ['-s', device, 'shell', 'am', 'force-stop', 'com.whatsapp'],
            ['-s', device, 'shell', 'am', 'force-stop', 'com.whatsapp.w4b']
        ]
        for cmd in close_commands:
            self._run_adb_command(cmd, timeout=3)
        
        # 2) Abrir WhatsApp y cambiar de cuenta
        self._run_adb_command(['-s', device, 'shell', 'am', 'start', '-n', 'com.whatsapp/.Main'], timeout=10)
        time.sleep(3)  # Esperar a que abra
        
        # Navegar al menú de cambio de cuenta
        for _ in range(2):
            self._run_adb_command(['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_DPAD_UP'], timeout=3)
            time.sleep(0.2)
        
        self._run_adb_command(['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_DPAD_RIGHT'], timeout=3)
        time.sleep(0.2)
        self._run_adb_command(['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_ENTER'], timeout=3)
        time.sleep(0.2)
        
        for _ in range(7):
            self._run_adb_command(['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_TAB'], timeout=3)
            time.sleep(0.05)  # Más rápido: 0.05s entre TABs
        
        self._run_adb_command(['-s', device, 'shell', 'input', 'keyevent', 'KEYCODE_ENTER'], timeout=3)
        
        # Esperar 3 segundos con WhatsApp abierto para que carguen los mensajes
        self.log(f"[{device}] Esperando 3s para que carguen los mensajes...", 'info')
        time.sleep(3)
        
        # 3) Cerrar todo nuevamente
        for cmd in close_commands:
            self._run_adb_command(cmd, timeout=3)
        
        self.log(f"[{device}] Cambio de cuenta completado", 'success')
    
    def _run_adb_command(self, args, timeout=10):
        """Ejecuta un comando ADB y maneja errores comunes."""
        # --- FIX: Añadir bucle de pausa ---
        while self.is_paused and not self.should_stop:
            time.sleep(0.1)
        if self.should_stop:
            return False # Indicar que la tarea fue cancelada
        # --- FIN FIX ---

        adb = self.adb_path.get()
        full_args = [adb] + args # Construye la lista completa de argumentos

        # Ocultar ventana de consola de ADB
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        startupinfo.wShowWindow = subprocess.SW_HIDE

        try:
            # Ejecutar SIEMPRE como lista, NUNCA con shell=True si hay rutas
            result = subprocess.run(full_args, capture_output=True, text=True, timeout=timeout, startupinfo=startupinfo, check=False, encoding='utf-8', errors='ignore')
            if result.returncode != 0 and result.stderr:
                # Limpiar errores comunes
                stderr_clean = result.stderr.strip()
                if "Usage: input" in stderr_clean:
                    stderr_clean = "Error Uso Input (argumentos inválidos)"
                elif "NullPointerException" in stderr_clean:
                    stderr_clean = "NullPointerException (procesando texto)"
                elif "Killed" in stderr_clean:
                    stderr_clean = "Proceso Killed (mensaje largo?)"
                elif "unknown command" in stderr_clean:
                     stderr_clean = "Comando ADB desconocido"
                elif "device unauthorized" in stderr_clean:
                     stderr_clean = "Dispositivo no autorizado (revisa el teléfono)"
                elif "device not found" in stderr_clean:
                     stderr_clean = "Dispositivo no encontrado (desconectado?)"

                self.log(f"Error ADB: {stderr_clean}", 'error')
                return False # Indicar fallo
            elif result.returncode != 0:
                 self.log(f"Error ADB (código {result.returncode}, sin stderr)", 'error')
                 return False # Indicar fallo
            return True # Indicar éxito
        except subprocess.TimeoutExpired:
            self.log("Timeout en comando ADB", 'error')
            return False
        except Exception as e:
            self.log(f"Error inesperado ejecutando ADB: {e}", 'error')
            return False

    def send_msg(self, device, link, i, total, message_to_send=None, whatsapp_package="com.whatsapp.w4b", ui_device=None):
        """Ejecuta los comandos para enviar un único mensaje, usando uiautomator2 si está disponible."""
        try:
            num_display = "?"
            is_group = bool(message_to_send)
            if is_group:
                num_display = f"Grupo ({link[:40]}...)"
            elif 'wa.me/' in link:
                num_display = link.split('wa.me/')[1].split('?')[0]

            self.log(f"({i}/{total}) → {num_display} [en {device}]", 'info')

            # 1. Abrir el enlace de WhatsApp (siempre con ADB para fiabilidad)
            self.log(f"Abriendo link en {device} con ADB...", 'info')
            open_args = ['-s', device, 'shell', 'am', 'start', '-a', 'android.intent.action.VIEW', '-d', f'"{link}"', '-p', whatsapp_package]
            if not self._run_adb_command(open_args, timeout=20):
                self.log(f"Fallo al abrir link para {num_display}. Saltando...", "warning")
                return False
            
            time.sleep(self.wait_after_open.get())
            if self.should_stop: return False

            # 2. Determinar el mensaje a enviar
            msg_to_send = message_to_send
            if not msg_to_send and not is_group:
                try:
                    msg_to_send = urllib.parse.unquote(link.split('text=')[1])
                except IndexError:
                    self.log("No se pudo extraer el mensaje del link para uiautomator2.", "error")
                    return False

            # 3. Lógica de envío refactorizada: usar siempre uiautomator2 para escribir y hacer clic en el botón de enviar.
            if not ui_device:
                self.log("Error crítico: la conexión de uiautomator2 no está disponible.", "error")
                return False

            self.log("Usando uiautomator2 para escribir y enviar.", 'info')
            try:
                # Esperar a que aparezca el campo de texto. Si no es un grupo, el texto ya está, así que solo buscamos el botón de enviar.
                if is_group:
                    edit_text = ui_device(className="android.widget.EditText")
                    if not edit_text.wait(timeout=10):
                        self.log("No se encontró el campo de texto para escribir.", "error")
                        return False # Fallo, no se puede escribir.
                    edit_text.set_text(msg_to_send)

                # Esperar y hacer clic en el botón de enviar. Es el único método de envío.
                send_button = ui_device(description="Enviar")
                if not send_button.wait(timeout=7):
                    self.log("No se encontró el botón 'Enviar'. El mensaje no se puede enviar.", "error")
                    return False # Fallo, no se encontró el botón.

                send_button.click()

                self.log("Mensaje enviado con éxito.", 'success')
                time.sleep(1.5) # Pequeña pausa post-envío
                return True

            except Exception as e:
                self.log(f"Falló el envío con uiautomator2: {e}", 'error')
                return False

        except Exception as e:
            self.log(f"Error inesperado en send_msg: {e}", 'error')
            import traceback
            traceback.print_exc()
            return False
    # --- ################################################################## ---
    # --- FIN
    # --- ################################################################## ---
    def run_cambiador(self):
        """Ejecuta la secuencia Cambiador en todos los dispositivos."""
        if not self.devices:
            messagebox.showwarning("Sin dispositivos", "No hay dispositivos conectados. Detecta dispositivos primero.", parent=self.root)
            return
        
        adb_exe = self.adb_path.get()
        if not adb_exe or not os.path.exists(adb_exe):
            messagebox.showerror("Error ADB", "No se encontró la ruta de ADB. Detecta dispositivos primero.", parent=self.root)
            return
        
        confirm = messagebox.askyesno(
            "Cambiador",
            f"Se ejecutará la secuencia Cambiador en {len(self.devices)} dispositivo(s).\n\n"
            "Secuencia:\n"
            "• Abrir configuración de WhatsApp Business\n"
            "• Navegar y ejecutar acciones (13 TABs + ENTERs)\n"
            "• Abrir configuración de WhatsApp Normal\n"
            "• Navegar y ejecutar acciones (13 TABs + ENTERs)\n\n"
            "¿Continuar?",
            parent=self.root
        )
        
        if not confirm:
            return
        
        self.log(f"Iniciando Cambiador en {len(self.devices)} dispositivo(s)...", 'info')
        
        for idx, device in enumerate(self.devices, 1):
            self.log(f"[{idx}/{len(self.devices)}] Procesando: {device}", 'info')
            
            try:
                # Secuencia para WhatsApp Business
                self.log(f"  → Procesando WhatsApp Business en {device}", 'info')
                
                # Abrir configuración de WhatsApp Business
                cmd = f'"{adb_exe}" -s {device} shell am start -a android.settings.APPLICATION_DETAILS_SETTINGS -d package:com.whatsapp.w4b'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(3)  # Esperar 3 segundos para que la app se abra completamente
                
                # 13 TABs
                for i in range(13):
                    cmd = f'"{adb_exe}" -s {device} shell input keyevent KEYCODE_TAB'
                    subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                    time.sleep(0.05)
                
                # ENTER
                cmd = f'"{adb_exe}" -s {device} shell input keyevent KEYCODE_ENTER'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(0.05)
                
                # TAB
                cmd = f'"{adb_exe}" -s {device} shell input keyevent KEYCODE_TAB'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(0.05)
                
                # ENTER
                cmd = f'"{adb_exe}" -s {device} shell input keyevent KEYCODE_ENTER'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(0.05)
                
                # Cerrar configuración
                cmd = f'"{adb_exe}" -s {device} shell am force-stop com.android.settings'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(0.05)
                
                # Secuencia para WhatsApp Normal
                self.log(f"  → Procesando WhatsApp Normal en {device}", 'info')
                
                # Abrir configuración de WhatsApp Normal
                cmd = f'"{adb_exe}" -s {device} shell am start -a android.settings.APPLICATION_DETAILS_SETTINGS -d package:com.whatsapp'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(3)  # Esperar 3 segundos para que la app se abra completamente
                
                # 13 TABs
                for i in range(13):
                    cmd = f'"{adb_exe}" -s {device} shell input keyevent KEYCODE_TAB'
                    subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                    time.sleep(0.05)
                
                # ENTER
                cmd = f'"{adb_exe}" -s {device} shell input keyevent KEYCODE_ENTER'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(0.05)
                
                # TAB
                cmd = f'"{adb_exe}" -s {device} shell input keyevent KEYCODE_TAB'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(0.05)
                
                # ENTER
                cmd = f'"{adb_exe}" -s {device} shell input keyevent KEYCODE_ENTER'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(0.05)
                
                # Cerrar configuración
                cmd = f'"{adb_exe}" -s {device} shell am force-stop com.android.settings'
                subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(0.05)
                
                self.log(f"✓ Dispositivo {device} completado", 'success')
                
            except subprocess.TimeoutExpired:
                self.log(f"✗ Timeout en dispositivo {device}", 'error')
            except Exception as e:
                self.log(f"✗ Error en {device}: {str(e)}", 'error')
        
        self.log("✓ Cambiador completado en todos los dispositivos", 'success')
        messagebox.showinfo("Completado", f"Cambiador ejecutado en {len(self.devices)} dispositivo(s).", parent=self.root)
    
    def switch_whatsapp_account(self):
        """Cambia de cuenta en WhatsApp para todos los dispositivos."""
        if not self.devices:
            messagebox.showwarning("Sin dispositivos", "No hay dispositivos conectados. Detecta dispositivos primero.", parent=self.root)
            return
        
        adb_exe = self.adb_path.get()
        if not adb_exe or not os.path.exists(adb_exe):
            messagebox.showerror("Error ADB", "No se encontró la ruta de ADB. Detecta dispositivos primero.", parent=self.root)
            return
        
        confirm = messagebox.askyesno(
            "Cambiar Cuenta WhatsApp",
            f"Se ejecutará el cambio de cuenta en {len(self.devices)} dispositivo(s).\n\n"
            "Secuencia:\n"
            "• Abrir WhatsApp\n"
            "• Navegar al menú\n"
            "• Cambiar cuenta\n\n"
            "¿Continuar?",
            parent=self.root
        )
        
        if not confirm:
            return
        
        self.log(f"Iniciando cambio de cuenta en {len(self.devices)} dispositivo(s)...", 'info')
        
        for idx, device in enumerate(self.devices, 1):
            self.log(f"[{idx}/{len(self.devices)}] Procesando: {device}", 'info')
            
            # Usar la función _switch_account_for_device con delay de 0.4s (4x más lento)
            success = self._switch_account_for_device(device, delay=0.4)
            
            if success:
                self.log(f"✓ Dispositivo {device} completado", 'success')
            else:
                self.log(f"✗ Error en dispositivo {device}", 'error')
        
        self.log("✓ Cambio de cuenta completado en todos los dispositivos", 'success')
        messagebox.showinfo("Completado", f"Cambio de cuenta ejecutado en {len(self.devices)} dispositivo(s).", parent=self.root)
    
    def _switch_account_for_device(self, device, delay=0.1):
        """Cambia de cuenta en WhatsApp Normal para un dispositivo específico.
        
        Args:
            device: ID del dispositivo
            delay: Tiempo de espera entre comandos (default 0.1s para cambio automático rápido)
        """
        adb_exe = self.adb_path.get()
        if not adb_exe or not os.path.exists(adb_exe):
            self.log(f"Error: No se encontró ADB para cambiar cuenta en {device}", 'error')
            return False
        
        try:
            commands = [
                "shell am start -n com.whatsapp/.Main",
                "shell input keyevent KEYCODE_DPAD_UP",
                "shell input keyevent KEYCODE_DPAD_UP",
                "shell input keyevent KEYCODE_DPAD_RIGHT",
                "shell input keyevent KEYCODE_ENTER",
                "shell input keyevent KEYCODE_TAB",
                "shell input keyevent KEYCODE_TAB",
                "shell input keyevent KEYCODE_TAB",
                "shell input keyevent KEYCODE_TAB",
                "shell input keyevent KEYCODE_TAB",
                "shell input keyevent KEYCODE_TAB",
                "shell input keyevent KEYCODE_TAB",
                "shell input keyevent KEYCODE_ENTER"
            ]
            
            for cmd in commands:
                full_cmd = f'"{adb_exe}" -s {device} {cmd}'
                subprocess.run(full_cmd, shell=True, capture_output=True, text=True, timeout=5)
                time.sleep(delay)  # Tiempo configurable según el contexto
            
            self.log(f"✓ Cuenta cambiada en {device}", 'success')
            return True
            
        except subprocess.TimeoutExpired:
            self.log(f"✗ Timeout al cambiar cuenta en {device}", 'error')
            return False
        except Exception as e:
            self.log(f"✗ Error al cambiar cuenta en {device}: {str(e)}", 'error')
            return False

    def open_adb_injector(self):
        """Abre una ventana para inyectar comandos ADB a todos los dispositivos."""
        if not self.devices:
            messagebox.showwarning("Sin dispositivos", "No hay dispositivos conectados. Detecta dispositivos primero.", parent=self.root)
            return
        
        # Crear ventana de inyector
        injector_window = ctk.CTkToplevel(self.root)
        injector_window.title("HΞЯMΞS V1 - Inyector ADB")
        injector_window.geometry("900x700")
        injector_window.transient(self.root)

        # Centrar ventana
        injector_window.update_idletasks()
        root_x = self.root.winfo_rootx()
        root_y = self.root.winfo_rooty()
        root_w = self.root.winfo_width()
        root_h = self.root.winfo_height()
        x = root_x + (root_w // 2) - 450
        y = root_y + (root_h // 2) - 350
        injector_window.geometry(f"900x700+{x}+{y}")
        injector_window.after(100, injector_window.focus_force)
        
        # Contenedor principal
        main_cont = ctk.CTkFrame(injector_window, fg_color=self.colors['bg'], corner_radius=0)
        main_cont.pack(fill=tk.BOTH, expand=True)
        
        # Header
        header = ctk.CTkFrame(main_cont, fg_color=self.colors['action_detect'], height=80, corner_radius=0)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        ctk.CTkLabel(header, text="Inyector ADB - Comandos Multiples", font=('Inter', 22, 'bold'), text_color=self.colors['text_header']).pack(expand=True)
        
        # Contenido
        content = ctk.CTkFrame(main_cont, fg_color=self.colors['bg'], corner_radius=0)
        content.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
        
        # Info de dispositivos
        info_card = ctk.CTkFrame(content, fg_color=self.colors['bg_card'], corner_radius=15)
        info_card.pack(fill=tk.X, pady=(0, 15))
        info_frame = ctk.CTkFrame(info_card, fg_color="transparent")
        info_frame.pack(fill=tk.X, padx=20, pady=15)
        ctk.CTkLabel(info_frame, text=f"📱 Dispositivos conectados: {len(self.devices)}", 
                     font=self.fonts['card_title'], text_color=self.colors['text']).pack(anchor='w')
        devices_text = ", ".join(self.devices)
        ctk.CTkLabel(info_frame, text=devices_text, 
                     font=self.fonts['setting_label'], text_color=self.colors['text_light'], wraplength=800).pack(anchor='w', pady=(5, 0))
        
        # Campo de comando
        cmd_card = ctk.CTkFrame(content, fg_color=self.colors['bg_card'], corner_radius=15)
        cmd_card.pack(fill=tk.X, pady=(0, 15))
        cmd_frame = ctk.CTkFrame(cmd_card, fg_color="transparent")
        cmd_frame.pack(fill=tk.X, padx=20, pady=15)
        
        ctk.CTkLabel(cmd_frame, text="Comando ADB (sin 'adb -s <device>'):", 
                     font=self.fonts['setting_label'], text_color=self.colors['text']).pack(anchor='w', pady=(0, 5))
        
        cmd_var = tk.StringVar()
        cmd_entry = ctk.CTkEntry(cmd_frame, textvariable=cmd_var, font=('Consolas', 12), 
                                 corner_radius=10, height=40, placeholder_text="Ejemplo: shell input tap 500 1000")
        cmd_entry.pack(fill=tk.X, pady=(0, 10))
        cmd_entry.focus_set()
        
        # Ejemplos
        examples_label = ctk.CTkLabel(cmd_frame, 
                                      text="Ejemplos: shell input tap 500 1000 | shell input text Hola | shell input keyevent KEYCODE_HOME",
                                      font=('Inter', 10), text_color=self.colors['text_light'])
        examples_label.pack(anchor='w')
        
        # Botón ejecutar
        def execute_command():
            command = cmd_var.get().strip()
            if not command:
                messagebox.showwarning("Comando vacío", "Ingresa un comando ADB para ejecutar.", parent=injector_window)
                return
            
            # Ejecutar en todos los dispositivos
            log_output.configure(state=tk.NORMAL)
            log_output.insert(tk.END, f"\n{'='*80}\n", 'info')
            log_output.insert(tk.END, f"Ejecutando: {command}\n", 'info')
            log_output.insert(tk.END, f"{'='*80}\n", 'info')
            log_output.see(tk.END)
            log_output.configure(state=tk.DISABLED)
            
            for device in self.devices:
                log_output.configure(state=tk.NORMAL)
                log_output.insert(tk.END, f"\n[{device}] Ejecutando...\n", 'device')
                log_output.see(tk.END)
                log_output.configure(state=tk.DISABLED)
                
                # Construir comando completo
                cmd_parts = ['-s', device] + command.split()
                
                try:
                    result = subprocess.run(
                        [self.adb_path.get()] + cmd_parts,
                        capture_output=True,
                        text=True,
                        timeout=10
                    )
                    
                    log_output.configure(state=tk.NORMAL)
                    if result.returncode == 0:
                        log_output.insert(tk.END, f"[{device}] ✓ Éxito\n", 'success')
                        if result.stdout.strip():
                            log_output.insert(tk.END, f"Output: {result.stdout.strip()}\n", 'output')
                    else:
                        log_output.insert(tk.END, f"[{device}] ✗ Error (código {result.returncode})\n", 'error')
                        if result.stderr.strip():
                            log_output.insert(tk.END, f"Error: {result.stderr.strip()}\n", 'error')
                    log_output.see(tk.END)
                    log_output.configure(state=tk.DISABLED)
                    
                except subprocess.TimeoutExpired:
                    log_output.configure(state=tk.NORMAL)
                    log_output.insert(tk.END, f"[{device}] ✗ Timeout (>10s)\n", 'error')
                    log_output.see(tk.END)
                    log_output.configure(state=tk.DISABLED)
                except Exception as e:
                    log_output.configure(state=tk.NORMAL)
                    log_output.insert(tk.END, f"[{device}] ✗ Excepción: {e}\n", 'error')
                    log_output.see(tk.END)
                    log_output.configure(state=tk.DISABLED)
            
            log_output.configure(state=tk.NORMAL)
            log_output.insert(tk.END, f"\nComando completado en todos los dispositivos.\n", 'success')
            log_output.see(tk.END)
            log_output.configure(state=tk.DISABLED)
        
        btn_frame = ctk.CTkFrame(cmd_card, fg_color="transparent")
        btn_frame.pack(fill=tk.X, padx=20, pady=(0, 15))
        
        exec_btn = ctk.CTkButton(btn_frame, text="▶ EJECUTAR EN TODOS", command=execute_command,
                                 fg_color=self.colors['action_start'], hover_color=self.hover_colors['action_start'],
                                 text_color=self.colors['text_header'], font=self.fonts['button'], corner_radius=10, height=45)
        exec_btn.pack(fill=tk.X)
        
        # Log de salida
        log_card = ctk.CTkFrame(content, fg_color=self.colors['bg_card'], corner_radius=15)
        log_card.pack(fill=tk.BOTH, expand=True)
        
        log_header = ctk.CTkFrame(log_card, fg_color="transparent")
        log_header.pack(fill=tk.X, padx=20, pady=(15, 10))
        ctk.CTkLabel(log_header, text="📝 Registro de Ejecución", 
                     font=self.fonts['card_title'], text_color=self.colors['text']).pack(anchor='w')
        
        log_container = ctk.CTkFrame(log_card, fg_color="transparent")
        log_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        
        log_output = ctk.CTkTextbox(log_container, fg_color=self.colors['bg_log'], 
                                    text_color=self.colors['log_text'], font=('Consolas', 10),
                                    corner_radius=10, activate_scrollbars=True, border_width=1, border_color="#444851")
        log_output.pack(fill=tk.BOTH, expand=True)
        
        # Configurar tags de color
        log_output.tag_config('success', foreground=self.colors['log_success'])
        log_output.tag_config('error', foreground=self.colors['log_error'])
        log_output.tag_config('info', foreground=self.colors['log_info'])
        log_output.tag_config('device', foreground=self.colors['log_warning'])
        log_output.tag_config('output', foreground='#98c379')
        
        log_output.configure(state=tk.DISABLED)
        
        # Mensaje inicial
        log_output.configure(state=tk.NORMAL)
        log_output.insert(tk.END, "Inyector ADB iniciado\n", 'success')
        log_output.insert(tk.END, f"Dispositivos disponibles: {len(self.devices)}\n", 'info')
        log_output.insert(tk.END, "Ingresa un comando y presiona EJECUTAR\n\n", 'info')
        log_output.configure(state=tk.DISABLED)
        
        # Bind Enter para ejecutar
        cmd_entry.bind('<Return>', lambda e: execute_command())

    def detect_phone_numbers_thread(self):
        """Inicia la detección de números de teléfono en un hilo separado."""
        if not self.devices:
            messagebox.showwarning("Sin dispositivos", "No hay dispositivos conectados. Detecta dispositivos primero.", parent=self.root)
            return
        threading.Thread(target=self._detect_and_prepare_phone_lines, args=(True,), daemon=True).start()

    def _detect_and_prepare_phone_lines(self, show_results_popup=False):
        """
        Detecta todos los números de WA y WA Business en los dispositivos conectados
        y los almacena en self.detected_phone_lines.
        Retorna True si se encontraron líneas, False en caso contrario.
        """
        self.log("Iniciando detección de números de teléfono...", 'info')
        self.root.after(0, self.detect_numbers_btn.configure, {'state': tk.DISABLED, 'text': "Detectando..."})

        self.detected_phone_lines = []
        temp_dir = tempfile.mkdtemp(prefix="hermes_ui_dump_")

        try:
            all_numbers = {}
            for device_serial in self.devices:
                self.log(f"Procesando dispositivo: {device_serial}", 'info')
                numbers = self._get_number_with_uiautomator2(device_serial, temp_dir)
                all_numbers[device_serial] = numbers

                if numbers.get("WhatsApp") not in ["No encontrado", "Error"]:
                    self.detected_phone_lines.append({"device": device_serial, "type": "WhatsApp", "number": numbers["WhatsApp"], "package": "com.whatsapp"})
                if numbers.get("WhatsApp Business") not in ["No encontrado", "Error"]:
                    self.detected_phone_lines.append({"device": device_serial, "type": "WhatsApp Business", "number": numbers["WhatsApp Business"], "package": "com.whatsapp.w4b"})

            # Actualizar la UI
            display_text = ""
            if self.detected_phone_lines:
                display_text = "Líneas detectadas:\n"
                # Ordenar para consistencia
                sorted_lines = sorted(self.detected_phone_lines, key=lambda x: (x['device'], x['type']))
                for line in sorted_lines:
                    display_text += f"· {line['type']} ({line['device']}): {line['number']}\n"
            else:
                display_text = "No se encontraron líneas de WhatsApp válidas."

            # Actualizar la etiqueta en la vista de Fidelizado
            if hasattr(self, 'fidelizado_device_list_label'):
                self.root.after(0, self.fidelizado_device_list_label.configure, {'text': display_text})

            if show_results_popup:
                self.log("Detección de números finalizada.", 'success')
                # Usar un formato ligeramente diferente para el popup para mejor legibilidad
                popup_text = display_text.replace("·", "\n·").strip()
                self.root.after(0, lambda: messagebox.showinfo("Detección Finalizada", popup_text, parent=self.root))

            return bool(self.detected_phone_lines)

        except Exception as e:
            self.log(f"Error durante la detección de números: {e}", 'error')
            if show_results_popup:
                self.root.after(0, lambda: messagebox.showerror("Error", f"Ocurrió un error inesperado:\n{e}", parent=self.root))
            return False
        finally:
            try:
                shutil.rmtree(temp_dir)
            except Exception as e:
                self.log(f"Error al eliminar el directorio temporal: {e}", "warning")
            self.root.after(0, self.detect_numbers_btn.configure, {'state': tk.NORMAL, 'text': "Detectar Números"})

    def _get_number_with_uiautomator2(self, device_serial, temp_dir):
        """Usa uiautomator2 para encontrar los números de teléfono en un dispositivo."""
        device_numbers = {"WhatsApp": "No encontrado", "WhatsApp Business": "No encontrado"}
        try:
            d = u2.connect(device_serial)
            # Desbloquear si es necesario (mejor si el usuario lo hace)
            d.unlock()

            # 1. Listar paquetes de WhatsApp según la selección del usuario
            wa_mode = self.whatsapp_mode.get()
            self.log(f"  Optimizando detección para el modo: {wa_mode}", 'info')

            packages_to_scan = []
            if wa_mode == "Normal":
                packages_to_scan.append("com.whatsapp")
            elif wa_mode == "Business":
                packages_to_scan.append("com.whatsapp.w4b")
            else: # Ambas o Todas
                packages_to_scan.append("com.whatsapp")
                packages_to_scan.append("com.whatsapp.w4b")

            # Verificar si los paquetes realmente existen en el dispositivo
            installed_packages_raw = d.shell('pm list packages').output
            installed_packages = installed_packages_raw.splitlines()

            final_packages_to_scan = [f"package:{p}" for p in packages_to_scan if f"package:{p}" in installed_packages]

            if not final_packages_to_scan:
                self.log(f"  No se encontraron las apps de WhatsApp seleccionadas en {device_serial}.", 'warning')
                return {"WhatsApp": "No encontrado", "WhatsApp Business": "No encontrado"}

            for pkg_full in final_packages_to_scan:
                pkg = pkg_full.split(':')[1]
                self.log(f"  Analizando paquete: {pkg} en {device_serial}", 'info')
                number = "No encontrado"
                try:
                    # 2. Abrir la app
                    d.app_start(pkg, stop=True)
                    d.wait_activity(".Main", timeout=12)
                    self.log(f"    {pkg} iniciado.", 'info')

                    # --- PASO A: Buscar en la pantalla principal de chats ---
                    self.log("    Paso A: Buscando número en la pantalla principal...", 'info')
                    time.sleep(2) # Esperar a que carguen los chats
                    main_screen_text = " ".join([elem.get_text() for elem in d(className="android.widget.TextView") if elem.exists and elem.get_text()])
                    match = re.search(r'(\+[\d\s\-\(\)]+)', main_screen_text)
                    if match:
                        number = re.sub(r'[\s\-\(\)]', '', match.group(1))
                        self.log(f"    ¡Número encontrado en la pantalla principal!: {number}", 'success')

                        # Asignar y saltar al siguiente paquete
                        if 'w4b' in pkg:
                            device_numbers['WhatsApp Business'] = number
                        else:
                            device_numbers['WhatsApp'] = number
                        d.app_stop(pkg)
                        continue
                        # --- FIN PASO A ---
                    else:
                        self.log("    Número no encontrado en la pantalla principal. Continuando...", 'info')
                        # --- PASO B: Navegar a "Nuevo Chat" -> Contacto Propio -> Info ---
                        self.log("    Paso B: Navegando a 'Nuevo Chat'...", 'info')
                        new_chat_button = d(description="Nuevo chat")
                        if new_chat_button.wait(timeout=7):
                            new_chat_button.click()
                        time.sleep(1) # Reducido para acelerar

                        self.log("    Buscando contacto propio (Tú)...", 'info')
                        own_contact_element = d(textContains="Tú")

                        if own_contact_element.wait(timeout=7):
                            # Estrategia 1: Intentar extraer el número directamente de la lista
                            contact_text = own_contact_element.get_text(timeout=3)
                            if contact_text:
                                match = re.search(r'(\+[\d\s\-\(\)]+)', contact_text)
                                if match:
                                    number = re.sub(r'[\s\-\(\)]', '', match.group(1)) # Limpiar todos los símbolos no numéricos
                                    self.log(f"    Número encontrado directamente en la lista de contactos: {number}", 'success')
                                    # Asignar el número encontrado y saltar al siguiente paquete
                                    if 'w4b' in pkg:
                                        device_numbers['WhatsApp Business'] = number
                                    else:
                                        device_numbers['WhatsApp'] = number
                                    d.press("back") # Volver de la lista de contactos
                                    continue      # Ir al siguiente paquete de whatsapp

                            # Estrategia 2: Si no se encuentra el número, navegar a la pantalla de información
                            self.log("    Número no visible en la lista. Navegando a la info de contacto...", 'info')
                            own_contact_element.click()
                            time.sleep(2) # Esperar a que se abra el chat

                            self.log("    Abriendo pantalla de información...", 'info')

                            # Estrategia de clic modificada para ser más robusta y evitar crashes.
                            clicked_successfully = False

                            # 1. El método más fiable: buscar por el texto visible "(Tú)".
                            self.log("    Intento 1: Buscando por texto '(Tú)'...", 'info')
                            contact_text_element = d(textContains="(Tú)")
                            if contact_text_element.wait(timeout=7):
                                try:
                                    contact_text_element.click()
                                    clicked_successfully = True
                                    self.log("    Éxito al hacer clic en elemento con texto '(Tú)'.", 'success')
                                except Exception as e:
                                    self.log(f"    Elemento de texto encontrado, pero el clic falló: {e}", 'warning')

                            # 2. Si no funciona, intentar el selector original por resourceId.
                            if not clicked_successfully:
                                self.log("    Intento 2: Buscando por resourceId 'conversation_contact_name'...", 'info')
                                contact_name_in_toolbar = d(resourceId=f"{pkg}:id/conversation_contact_name")
                                if contact_name_in_toolbar.wait(timeout=5):
                                    try:
                                        contact_name_in_toolbar.click()
                                        clicked_successfully = True
                                        self.log("    Éxito al hacer clic en 'conversation_contact_name'.", 'success')
                                    except Exception as e:
                                        self.log(f"    Elemento 'conversation_contact_name' encontrado, pero el clic falló: {e}", 'warning')

                            # 3. Como último recurso, hacer clic en toda la barra de herramientas.
                            if not clicked_successfully:
                                self.log("    Intento 3: Buscando por resourceId 'toolbar'...", 'info')
                                toolbar = d(resourceId=f"{pkg}:id/toolbar")
                                if toolbar.wait(timeout=5):
                                    try:
                                        toolbar.click()
                                        clicked_successfully = True
                                        self.log("    Éxito al hacer clic en 'toolbar'.", 'success')
                                    except Exception as e:
                                        self.log(f"    Elemento 'toolbar' encontrado, pero el clic falló: {e}", 'warning')

                            if clicked_successfully:
                                self.log("    En la pantalla de información, buscando número...", 'info')
                                time.sleep(2) # Esperar a que cargue la pantalla de info

                                # Bucle de desplazamiento inteligente con detección de final de página
                                number_found_in_text = False
                                max_scrolls = 10  # Límite de seguridad para evitar bucles infinitos
                                for i in range(max_scrolls):
                                    self.log(f"    Intento de búsqueda y scroll #{i+1}...", 'info')

                                    # Capturar textos ANTES de desplazarse
                                    before_scroll_texts = {elem.get_text() for elem in d(className="android.widget.TextView") if elem.exists and elem.get_text()}
                                    current_view_text = " ".join(before_scroll_texts)

                                    # Buscar el número en la vista actual
                                    match = re.search(r'(\+[\d\s\-\(\)]+)', current_view_text)
                                    if match:
                                        number = re.sub(r'[\s\-\(\)]', '', match.group(1))
                                        self.log(f"    ¡Número encontrado!: {number}", 'success')
                                        number_found_in_text = True
                                        break

                                    # Si no se encuentra, realizar el desplazamiento
                                    self.log("    Número no encontrado, realizando 'fling'...", 'info')
                                    scrollable_view = d(scrollable=True)
                                    if scrollable_view.exists:
                                        scrollable_view.fling.forward()
                                    else:
                                        # Fallback a swipe si no hay un elemento "scrollable"
                                        width, height = d.info['displayWidth'], d.info['displayHeight']
                                        d.swipe(width / 2, height * 0.8, width / 2, height * 0.2, 0.5)

                                    time.sleep(1.5)

                                    # Capturar textos DESPUÉS de desplazarse y comparar
                                    after_scroll_texts = {elem.get_text() for elem in d(className="android.widget.TextView") if elem.exists and elem.get_text()}
                                    if before_scroll_texts == after_scroll_texts:
                                        self.log("    El contenido no cambió. Se ha llegado al final de la página.", 'info')
                                        # Realizar una última búsqueda por si acaso
                                        final_text = " ".join(after_scroll_texts)
                                        final_match = re.search(r'(\+[\d\s\-\(\)]+)', final_text)
                                        if final_match:
                                            number = re.sub(r'[\s\-\(\)]', '', final_match.group(1))
                                            self.log(f"    Número encontrado en la última vista: {number}", 'success')
                                            number_found_in_text = True
                                        break

                                if not number_found_in_text:
                                    self.log("    Búsqueda finalizada. No se encontró el número en la página.", 'warning')
                                    number = "No encontrado"
                            else:
                                self.log("    No se pudo encontrar un elemento para abrir la pantalla de info.", 'error')
                        else:
                            self.log("    No se encontró el contacto '(Tú)' en la lista.", 'error')

                except Exception as e:
                    self.log(f"    Error procesando {pkg}: {e}", 'error')

                # Asignar nombre amigable
                if 'w4b' in pkg:
                    device_numbers['WhatsApp Business'] = number
                else:
                    device_numbers['WhatsApp'] = number

                d.app_stop(pkg) # Limpiar para la siguiente app

        except Exception as e:
            self.log(f"  Fallo grave al conectar o procesar {device_serial}: {e}", 'error')
            return {"WhatsApp": "Error", "WhatsApp Business": "Error"}

        return device_numbers

    def close_all_apps(self, device):
        """Fuerza el cierre de WhatsApp y Google (MOD 25)."""
        self.log(f"Cerrando apps en {device}", 'info')
        targets = ["com.whatsapp.w4b", "com.whatsapp", "com.google.android.googlequicksearchbox"]
        for package in targets:
            close_args = ['-s', device, 'shell', 'am', 'force-stop', package]
            self._run_adb_command(close_args, timeout=5) # Usar la función helper, ignorar resultado

# --- Main y Login ---
def main():
    """Función principal: Configura CTk y muestra la ventana de login."""
    ctk.set_appearance_mode("Light")
    ctk.set_default_color_theme("blue")
    root = ctk.CTk()
    root.title("HΞЯMΞS - Autenticación")

    # Colores y fuentes específicos para el login
    colors = {
        'bg': '#f0f2f5', 'bg_card': '#ffffff', 'text': '#202124',
        'text_light': '#5f6368', 'blue': '#4285F4', 'action_start': '#16A34A'
    }
    fonts = {
        'header': ('Big Russian', 64, 'bold'),
        'card_title': ('Inter', 16, 'bold'),
        'button': ('Inter', 13, 'bold'),
        'setting_label': ('Inter', 12)
    }

    def show_main_app():
        """Destruye los widgets de login y construye la app principal."""
        root.unbind('<Return>')  # --- FIX: Desactivar el atajo de teclado de login ---
        for w in root.winfo_children():
            w.destroy()
        app = Hermes(root) # Reutiliza la ventana raíz

    def create_login_window():
        """Construye la UI de la ventana de login (MOD 37)."""
        root.configure(fg_color=colors['bg'])
        width, height = 450, 550
        root.geometry(f"{width}x{height}")
        root.resizable(False, False)

        # Centrar ventana
        root.update_idletasks()
        sw = root.winfo_screenwidth()
        sh = root.winfo_screenheight()
        x = (sw // 2) - (width // 2)
        y = (sh // 2) - (height // 2)
        root.geometry(f'{width}x{height}+{x}+{y}')
        root.attributes('-topmost', True)
        root.after(100, root.focus_force)

        # Frame contenedor transparente
        content_frame = ctk.CTkFrame(root, fg_color="transparent")
        content_frame.pack(expand=True, fill="x", padx=40)

        # Logo
        try:
            logo_p = os.path.join(BASE_DIR, 'logo_left.png')
            logo_i = Image.open(logo_p).resize((150, 150), Image.Resampling.LANCZOS)
            logo_img = ctk.CTkImage(light_image=logo_i, dark_image=logo_i, size=(150, 150))
            logo_label = ctk.CTkLabel(content_frame, image=logo_img, text="", fg_color="transparent")
            logo_label.pack(pady=(20, 10))
        except Exception as e:
            print(f"Error cargando logo en login: {e}")
            logo_label = ctk.CTkLabel(content_frame, text="🦶", font=('Inter', 60, 'bold'), fg_color="transparent")
            logo_label.pack(pady=(20, 10))

        # Título HΞЯMΞS
        title_label = ctk.CTkLabel(content_frame, text="HΞЯMΞS", font=fonts['header'], text_color=colors['text'], fg_color="transparent")
        title_label.pack(pady=(0, 5))

        # Subtítulo
        subtitle_label = ctk.CTkLabel(content_frame, text="Ingrese la contraseña", font=fonts['setting_label'], text_color=colors['text_light'], fg_color="transparent")
        subtitle_label.pack(pady=(0, 30))

        # Campo Contraseña
        pwd_frame = ctk.CTkFrame(content_frame, fg_color='transparent')
        pwd_frame.pack(fill="x", pady=(0, 5))

        ctk.CTkLabel(pwd_frame, text="Contraseña", font=fonts['button'], text_color=colors['text']).pack(anchor='w')
        pwd_var = tk.StringVar()
        pwd_entry = ctk.CTkEntry(pwd_frame, textvariable=pwd_var, font=('Inter', 14), show='*', corner_radius=10, height=40, border_color="#cccccc", border_width=1)
        pwd_entry.pack(fill=tk.X, pady=(5, 0))
        pwd_entry.focus_set()

        # Mensaje de Estado (para error)
        status_l = ctk.CTkLabel(content_frame, text="", font=('Inter', 10, 'bold'), text_color='red', fg_color="transparent")
        status_l.pack(pady=(5, 10))

        # Botón INGRESAR
        def check_pwd(e=None):
            if pwd_var.get() == "1234": # Contraseña de login
                root.attributes('-topmost', False)
                show_main_app()
            else:
                status_l.configure(text="Contraseña incorrecta.")
                pwd_var.set("")

        login_btn = ctk.CTkButton(content_frame, text="INGRESAR", command=check_pwd,
                                  fg_color=colors['action_start'],
                                  hover_color=darken_color(colors['action_start'], 0.18),
                                  text_color='#ffffff', font=fonts['button'],
                                  corner_radius=30, height=50)
        login_btn.pack(fill="x", pady=(20, 20))

        root.bind('<Return>', check_pwd)
        root.protocol("WM_DELETE_WINDOW", root.destroy)

    # Iniciar la ventana de login
    create_login_window()
    root.mainloop()

if __name__ == "__main__":
    main()